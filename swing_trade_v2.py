"""
台股隔日沖選股與回測系統 - swing_trade.py
==========================================
功能：
  1. 選股   -- 收盤後掃描，輸出明日候選股（終端機 + CSV）
  2. 回測   -- 對歷史資料跑策略，計算勝率 / 期望值 / 最大回撤
  3. 歷史查詢 -- 從 SQLite 查詢過去選股與實際損益

策略（三十六個，原有十六個 + 新增二十個）：
  原有策略：
    A多 均線突破+爆量   AS空 均線死亡+爆量
    B多 開盤跳空向上   BS空 開盤跳空向下
    C多 RSI超賣反彈    CS空 RSI超買反轉
    D多 突破近5日高點  DS空 跌破前低
    E多 強勢連漲       ES空 弱勢連跌
  均量類：
    F多 均量擴張+爆量上漲   FS空 均量萎縮+爆量下跌
    G多 縮量後爆量上漲      GS空 縮量後爆量下跌
  K棒型態類：
    H多 鎚子K（下影線≥2×實體）   HS空 射擊之星（上影線≥2×實體）
    I多 吞噬陽線                  IS空 吞噬陰線
  技術指標類（新增）：
    J多 MACD黃金交叉             JS空 MACD死亡交叉
    K多 布林下軌反彈             KS空 布林上軌反壓
    L多 KD超賣黃金交叉(<20)      LS空 KD超買死亡交叉(>70)
    M多 威廉%R超賣反彈(<-80)     MS空 威廉%R超買回落(>-20)
    N多 多頭排列回測MA5站回      NS空 空頭排列反彈MA5跌破
  進階K棒型態（新增）：
    O多 晨星三K棒               OS空 黃昏之星三K棒
    P多 紅三兵                  PS空 黑三兵
    Q多 Inside Bar向上突破      QS空 Inside Bar向下跌破
  均值回歸（新增）：
    R多 BIAS乖離率<-10%反彈     RS空 BIAS乖離率>+8%回落

進出場邏輯（隔日沖）：
  進場：今日收盤訊號成立 → 明日開盤價買入（或融券賣出）
  出場：明日收盤價
  停損：若設定 --stop-loss，用明日 Low（多）/ High（空）估算觸損（估算值）

SQLite 儲存：
  - swing_trade.db（與腳本同目錄）
  - 每次選股結果自動存入 scans 資料表
  - 回測結果存入 backtests 資料表
  - 可跨日查詢歷史選股與實際走勢

用法：
  # 選股（今日收盤後執行）
  python swing_trade.py --scan --min-hit 2 --save

  # 回測（60 日）
  python swing_trade.py --backtest --days 60 --min-hit 2 --stop-loss 1.5

  # 回測（30 日，比較兩個停損設定）
  python swing_trade.py --backtest --days 30 --stop-loss 0
  python swing_trade.py --backtest --days 30 --stop-loss 1.5

  # 查詢歷史選股
  python swing_trade.py --history --limit 10

  # 查詢特定股票歷史
  python swing_trade.py --history --code 2330
"""

import os
import sys
import math
import time
import sqlite3
import argparse
import datetime
import unicodedata

import numpy as np
import threading
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import requests

try:
    import certifi
except ImportError:
    certifi = None

try:
    import yfinance as yf
    HAS_YF = True
except ImportError:
    HAS_YF = False

try:
    import twstock
    HAS_TWSTOCK = True
except ImportError:
    HAS_TWSTOCK = False

try:
    import shioaji as sj
    HAS_SHIOAJI = True
except ImportError:
    HAS_SHIOAJI = False


# ══════════════════════════════════════════════
# 常數
# ══════════════════════════════════════════════
DB_PATH      = os.environ.get("SWING_DB_PATH",
               os.path.join(os.path.dirname(os.path.abspath(__file__)), "swing_trade.db"))
TRADE_COST   = 0.00435   # 手續費 0.1425%×2 + 證交稅 0.15%
DEFAULT_DAYS = 60
RSI_PERIOD   = 14
VOL_MULT     = 1.5
MIN_AVG_VOL  = 3000      # 5 日均量最低門檻（張）

STRATEGY_NAMES = {
    # ── 原有十個策略 ──
    "A":  "多 均線突破",        "AS": "空 均線死亡+爆量",
    "B":  "多 開盤跳空向上",    "BS": "空 開盤跳空向下",
    "C":  "多 RSI超賣反彈",     "CS": "空 RSI超買反轉",
    "D":  "多 突破近5日高點",   "DS": "空 跌破近5日低點",
    "E":  "多 強勢連漲",        "ES": "空 弱勢連跌",
    # ── 均量類 ──
    "F":  "多 均量擴張+上漲",   "FS": "空 均量萎縮+下跌",
    "G":  "多 縮量後上漲",      "GS": "空 縮量後爆量下跌",
    # ── K棒型態 ──
    "H":  "多 鎚子K",           "HS": "空 射擊之星",
    "I":  "多 吞噬陽線",        "IS": "空 吞噬陰線",
    # ── 技術指標類（新增）──
    "J":  "多 MACD黃金交叉",    "JS": "空 MACD死亡交叉",
    "K":  "多 布林下軌反彈",    "KS": "空 布林上軌反壓",
    "L":  "多 KD超賣黃金交叉",  "LS": "空 KD超買死亡交叉",
    "M":  "多 威廉%R超賣反彈",  "MS": "空 威廉%R超買回落",
    "N":  "多 多頭排列回測站回", "NS": "空 空頭排列反彈跌破",
    # ── 進階K棒型態（新增）──
    "O":  "多 晨星三K棒",       "OS": "空 黃昏之星三K棒",
    "P":  "多 紅三兵",          "PS": "空 黑三兵",
    "Q":  "多 InsideBar突破",   "QS": "空 InsideBar跌破",
    # ── 均值回歸（新增）──
    "R":  "多 BIAS超跌反彈",    "RS": "空 BIAS超漲回落",
    # ── 大跳空低量（新增）──
    "B2": "多 大跳空(≥5%)+低量(≥0.8x)",
    # ── 頸線型態（新增）──
    "T":  "多 W底頸線突破",
    "TS": "空 M頭頸線跌破",
}

# 預先建立策略順序索引（向量化與多執行緒共用）
_STRAT_KEYS = list(STRATEGY_NAMES.keys())            # ['A','AS','B','BS',...]
_IS_SHORT   = [k.endswith("S") for k in _STRAT_KEYS] # True 表示空方策略
_N_STRATS   = len(_STRAT_KEYS)

# ──────────────────────────────────────────────────────────────────────────────
# 訊號分類與條件說明（供 --list 使用）
# ──────────────────────────────────────────────────────────────────────────────

# 六大類別定義
SIGNAL_CATEGORIES: list[tuple[str, str, list[str]]] = [
    # (類別名稱, 英文名, [訊號碼...])
    ("① 缺口動能", "Gap Momentum",          ["B", "B2", "BS"]),
    ("② 超賣反轉", "Oversold Reversal",      ["R", "K", "L", "M", "C", "T",
                                               "RS", "KS", "LS", "MS", "CS", "TS"]),
    ("③ 趨勢延續", "Trend Following",        ["A", "N", "F",
                                               "AS", "NS", "FS"]),
    ("④ 動能突破", "Momentum Breakout",      ["D", "E",
                                               "DS", "ES"]),
    ("⑤ K線型態", "Candlestick Pattern",     ["H", "I", "O", "P", "Q",
                                               "HS", "IS", "OS", "PS", "QS"]),
    ("⑥ 其他振盪", "Other Oscillator",       ["G", "J",
                                               "GS", "JS"]),
]

# 每個訊號的觸發條件說明
SIGNAL_COND: dict[str, str] = {
    "B":  "gap≥+2%，收漲（量限已移除，量大反降 EV）",
    "B2": "gap≥+5%，量≥0.8x均量（B 的強化版，約佔 B 的 15%）",
    "BS": "gap≤-2%，收跌，爆量（量≥1.5x均量）",
    "C":  "RSI<30，close<MA20",
    "CS": "RSI>70，close>MA20",
    "D":  "收盤>5日最高，MA5>MA20",
    "DS": "收盤<5日最低，MA5<MA20",
    "E":  "連漲≥3日，MA5>MA10>MA20",
    "ES": "連跌≥3日，MA5<MA10<MA20",
    "F":  "成交量>均量，收漲",
    "FS": "成交量萎縮<均量，收跌",
    "G":  "量<均量（縮量），MA5>MA20，收漲",
    "GS": "量縮後爆量（≥1.5x），MA5<MA20，收跌",
    "H":  "鎚子K線（下影≥2倍實體），close<MA20",
    "HS": "射擊之星（上影≥2倍實體），close>MA20",
    "I":  "多頭吞噬（大陽吞前陰）",
    "IS": "空頭吞噬（大陰吞前陽）",
    "J":  "MACD 黃金交叉（DIF 上穿 DEA），MA5>MA20，收漲",
    "JS": "MACD 死亡交叉（DIF 下穿 DEA），MA5<MA20，收跌",
    "K":  "收盤接近布林下軌（±2%），close<MA20",
    "KS": "收盤接近布林上軌（±2%），close>MA20",
    "L":  "KD<20 黃金交叉（30→20 收緊，EV 最優點）",
    "LS": "KD>70 死亡交叉，close>MA20",
    "M":  "威廉%R<−80（超賣），MA5<MA20，收跌",
    "MS": "威廉%R>−20（超買），MA5>MA20，收漲",
    "N":  "MA5>MA10>MA20，回測 MA5 後收回 MA5 上",
    "NS": "MA5<MA10<MA20，反彈至 MA5 下方後再跌破",
    "O":  "晨星型態（陰→小實體→大陽），close<MA20",
    "OS": "黃昏之星（陽→小實體→大陰），close>MA20",
    "P":  "紅三兵（連三根陽線+收盤遞增+實體≥50%+上影<50%實體），MA5>MA20",
    "PS": "黑三兵（連三根實體遞增陰線），MA5<MA20",
    "Q":  "InsideBar 突破（前日內包，今日向上突破），MA5>MA20",
    "QS": "InsideBar 跌破（前日內包，今日向下跌破），MA5<MA20",
    "R":  "BIAS<−10%（超跌，收盤偏離 MA20 超過 10%）",
    "RS": "BIAS>+10%（超漲，收盤偏離 MA20 超過 10%）",
    "T":  "W底頸線突破：60根K棒內找兩個相近低點（差距<5%），收盤突破頸線（兩低點間高點）",
    "TS": "M頭頸線跌破：60根K棒內找兩個相近高點（差距<5%），收盤跌破頸線（兩高點間低點）",
    "A":  "MA5>MA20（多頭排列），昨收>前日收（上漲）",
    "AS": "MA5<MA20（空頭排列），爆量（≥1.5x均量），收跌",
}

# ══════════════════════════════════════════════════════════════
# 策略型態分類（供 --scan 輸出型態標籤）
# ══════════════════════════════════════════════════════════════
# 均值回歸：超賣/逆勢型，short hold（exit-day=3）
_REVERSION_SIGS = frozenset({
    "C","CS","H","HS","I","IS","K","KS","L","LS","M","MS","O","OS","R","RS","T","TS"
})
# 趨勢跟風：順勢/動能型，long hold（exit-day=15）
_TREND_SIGS = frozenset({
    "A","AS","D","DS","E","ES","F","FS","G","GS","J","JS","N","NS","P","PS","Q","QS"
})
# 缺口動能：中性，跟隨配對的其他訊號決定型態
_GAP_SIGS = frozenset({"B","BS","B2"})

def _classify_style(triggered_long: list) -> str:
    """
    輸入當日觸發的多方策略代碼清單，回傳型態標籤字串。
    格式："📉回歸 D3" / "📈跟風 D15" / "🔀混合 D3~15" / "⚡缺口 D3"

    判斷邏輯：
    ① 無非缺口訊號 → ⚡缺口 D3（純缺口）
    ② 有缺口訊號（B/B2）時，缺口為主要進場催化劑，出場以 D3 為準：
       - 非缺口訊號全為回歸 → 📉回歸 D3（如 B+K+R, B+L）
       - 非缺口訊號全為趨勢 → ⚡缺口 D3（如 B+J, B+Q；缺口+趨勢確認，仍是D3組合）
       - 回歸 > 趨勢      → 🔀混合↘ D3
       - 趨勢 > 回歸      → 🔀混合↗ D3（缺口+趨勢混合，保守取 D3）
       - 趨勢 == 回歸     → 🔀混合 D3~15
    ③ 無缺口訊號時，純由非缺口訊號決定：
       - 全趨勢 → 📈跟風 D15
       - 全回歸 → 📉回歸 D3
       - 趨勢 > 回歸 → 🔀混合↗ D15
       - 回歸 > 趨勢 → 🔀混合↘ D3
       - 相等   → 🔀混合 D3~15
    """
    non_gap = [k for k in triggered_long if k not in _GAP_SIGS]
    has_gap = any(k in _GAP_SIGS for k in triggered_long)
    rev   = sum(1 for k in non_gap if k in _REVERSION_SIGS)
    trend = sum(1 for k in non_gap if k in _TREND_SIGS)

    # ① 純缺口（沒有其他訊號）
    if not non_gap:
        return "⚡缺口 D3" if has_gap else "—"

    # ② 有缺口 → 缺口優先，出場日不超過 D3
    if has_gap:
        if rev == 0:                   # 缺口 + 趨勢確認（B+J, B+Q, B+F...）
            return "⚡缺口 D3"
        if trend == 0:                 # 缺口 + 回歸訊號（B+K+R, B+L...）
            return "📉回歸 D3"
        if trend > rev:
            return "🔀混合↗ D3"
        # rev >= trend（含平手）：gap 存在時偏 D3
        return "🔀混合↘ D3"

    # ③ 無缺口 → 純訊號決定
    if trend > 0 and rev == 0:
        return "📈跟風 D15"
    if rev > 0 and trend == 0:
        return "📉回歸 D3"
    if trend > rev:
        return "🔀混合↗ D15"
    if rev > trend:
        return "🔀混合↘ D3"
    return "🔀混合 D3~15"


def print_strategy_list() -> None:
    """--list：印出所有訊號分類與 preset 組合清單。"""
    W = 78
    line = "═" * W

    print(f"\n{line}")
    print(f"  📋  台股隔日沖  訊號代碼 & 策略組合 總覽")
    print(f"{line}\n")

    # ── Part 1：訊號分類 ──────────────────────────────────────
    print("  【訊號分類】\n")
    for cat_zh, cat_en, codes in SIGNAL_CATEGORIES:
        print(f"  {cat_zh}  {cat_en}")
        print(f"  {'─' * (W - 2)}")
        long_codes  = [c for c in codes if not c.endswith("S") or c == "BS"]
        short_codes = [c for c in codes if c.endswith("S") and c != "BS"]

        # 先印多方
        for code in long_codes:
            name = STRATEGY_NAMES.get(code, "")
            cond = SIGNAL_COND.get(code, "")
            direction = "🔴多" if not code.endswith("S") else "🔵空"
            if code == "BS":
                direction = "🔵空"
            label = f"  {code:<4} {direction}  {name:<18}  {cond}"
            print(label)
        # 再印空方（若有）
        if short_codes:
            print()
            for code in short_codes:
                name = STRATEGY_NAMES.get(code, "")
                cond = SIGNAL_COND.get(code, "")
                label = f"  {code:<4} 🔵空  {name:<18}  {cond}"
                print(label)
        print()

    # ── Part 2：Preset 組合清單 ───────────────────────────────
    print(f"\n{line}")
    print(f"  📦  Preset 組合清單  （--preset NAME 使用）")
    print(f"{line}\n")

    for pname, combos in COMBO_PRESETS.items():
        exit_day = PRESET_EXIT_DAY.get(pname)
        exit_str = f"exit-day={exit_day}" if exit_day else "exit-day=2（預設）"
        direction = "多方" if pname.startswith("long") else ("空方" if pname.startswith("short") else "多空")
        n = len(combos)
        if n == 0:
            status = "（已廢除）"
        else:
            status = f"{n} 個組合"
        print(f"  ── {pname:<16} {direction}  {status}  {exit_str}")

        if not combos:
            print()
            continue

        tiers = PRESET_TIERS.get(pname)
        if tiers:
            # 有 tier 定義：按 EV 分組顯示，含 WR/EV/n 指標
            tier_idx = 0
            cur_threshold = tiers[0][1]
            print(f"       {'組合':<14} {'勝率':>6}  {'期望值':>8}  {'樣本':>5}  備注")
            print(f"       {'─'*60}")
            # 先印 tier header
            print(f"       {tiers[0][0]}")
            for combo in combos:
                meta = COMBO_META.get(combo)
                if meta:
                    wr, ev, sn, note = meta
                    # 若 EV 低於當前 tier 門檻，切換到下一個 tier
                    while tier_idx + 1 < len(tiers) and ev < tiers[tier_idx][1]:
                        tier_idx += 1
                        cur_threshold = tiers[tier_idx][1]
                        print(f"       {tiers[tier_idx][0]}")
                    warn = " ⚠️少" if sn < 100 else ""
                    note_str = f"  {note}" if note else ""
                    print(f"       {combo:<14} {wr:>5.1f}%  EV{ev:>+7.3f}%  n={sn:<5}{warn}{note_str}")
                else:
                    print(f"       {combo}")
        else:
            # 無 tier 定義：簡單列出
            for combo in combos:
                meta = COMBO_META.get(combo)
                if meta:
                    wr, ev, sn, note = meta
                    warn = " ⚠️少" if sn < 100 else ""
                    print(f"       {combo:<14} {wr:>5.1f}%  EV{ev:>+7.3f}%  n={sn:<5}{warn}")
                else:
                    print(f"       {combo}")
        print()

    print(line)
    print()


# ══════════════════════════════════════════════
# 組合預設清單（--preset 用）
# 新增自訂 preset：在 COMBO_PRESETS 加入一行即可
# ══════════════════════════════════════════════
COMBO_PRESETS: dict[str, list[str]] = {
# ── 空方 ────────────────────────────────────
    "short1": [
        "BS+FS+GS",   # 75.0%  EV+2.254%  n=36
        "AS+BS+GS",   # 72.7%  EV+2.555%  n=88
        "BS+DS+GS",   # 67.0%  EV+2.145%  n=100
        "BS+GS",      # 64.8%  EV+1.948%  n=108
        "BS+ES+GS",   # 60.0%  EV+0.340%  n=5  (⚠️少)
        "BS+ES+FS",   # 57.8%  EV+0.533%  n=64
        "AS+DS+GS",   # 56.5%  EV+0.850%  n=216
        "AS+BS+FS",   # 54.5%  EV+0.727%  n=442
        "AS+BS+DS",   # 53.0%  EV+0.634%  n=1118
        "AS+BS",      # 52.6%  EV+0.603%  n=1308
    ],
    # ── 多方 ─────────────────────────────────────
    "long1": [
        "B+E+F",   # 41.2%  EV-0.554%  n=413
        "B+E",     # 40.9%  EV-0.606%  n=711
        "B+D+E",   # 40.1%  EV-0.705%  n=568
        "B+C+D",   # 39.6%  EV-0.500%  n=164
        "A+B+E",   # 39.6%  EV-0.697%  n=626
        "B+D+F",   # 39.0%  EV-0.772%  n=1692
        "B+C",     # 38.1%  EV-0.540%  n=462
        "B+F",     # 37.9%  EV-0.727%  n=2190
        "A+B+F",   # 37.7%  EV-0.763%  n=2037
        "A+B+D",   # 37.3%  EV-0.815%  n=3217
        "B+D",     # 37.2%  EV-0.778%  n=3873
        "C+D+F",   # 37.0%  EV-0.475%  n=46
        "D+E+F",   # 36.6%  EV-0.664%  n=1085
        "A+B",     # 36.3%  EV-0.805%  n=4165
    ],
    # ── 空方（v2 新版，含 J~R 策略，90日回測 2026-04）──────────
    "short2": [
        # ▸ BS 錨點策略群（跳空向下 + 超買確認）
        "BS+MS+RS",   # 80.0%  EV+2.797%  n=30
        "BS+MS",      # 75.0%  EV+2.561%  n=44
        "BS+JS",      # 82.4%  EV+2.287%  n=17
        "BS+RS",      # 78.1%  EV+2.420%  n=32
        "BS+KS",      # 70.8%  EV+2.132%  n=24
        "BS+CS",      # 65.0%  EV+1.388%  n=20
        "BS+FS",      # 59.4%  EV+1.474%  n=32
        # ▸ FS/CS 策略群（量縮/RSI 超買）
        "CS+FS",      # 66.7%  EV+2.682%  n=15
        "FS+MS+RS",   # 70.6%  EV+2.123%  n=17
        "FS+RS",      # 72.7%  EV+1.728%  n=22
        "FS+MS",      # 63.9%  EV+1.380%  n=36
        # ▸ 純超買確認群（無跳空錨點，樣本較大）
        "IS+KS+RS",   # 59.2%  EV+1.250%  n=49
        "KS+MS+RS",   # 54.9%  EV+0.550%  n=384  ← 樣本最穩定
        "LS+MS+RS",   # 55.3%  EV+0.539%  n=190
    ],
    # ── 多方（v2 新版，含 J~R 策略，90日回測 2026-04）──────────
    "long2": [
        # ▸ 動能突破類
        "B+C+D",      # 63.6%  EV+1.633%  n=22
        "B+M+O",      # 54.5%  EV+1.027%  n=11
        "B+C",        # 48.9%  EV+0.589%  n=47
        "B+M",        # 51.2%  EV+0.369%  n=82
        "B+D+M",      # 52.6%  EV+0.429%  n=19
        # ▸ 超跌多重確認類
        "L+R",        # 60.0%  EV+1.566%  n=35
        "L+M+R",      # 57.6%  EV+1.440%  n=33
        "K+L+M",      # 47.6%  EV+1.050%  n=21
        "K+L",        # 47.8%  EV+0.992%  n=23
        "C+R",        # 55.2%  EV+0.606%  n=29
        "C+M+R",      # 53.8%  EV+0.324%  n=26
        # ▸ 大樣本穩健型
        "M+R",        # 49.5%  EV+0.066%  n=208
    ],
    # ── 多方（v3：2400 日長期回測 + 停損敏感度驗證，2026-04-24）──────────
    # 篩選門檻：n ≥ 100 樣本充足、正 EV、含停損 -7% 驗證
    # 建議搭配 --stop-loss 7（預設值），EV 幾乎等同無停損但有尾端保護
    "long3": [
        # ▸ TIER 1 明星組合（實戰首選）
        "K+L+R",      # 49.4%  EV+0.576%  n=332  (無停損) | -7%: EV+0.714%  n=256
        "C+K+R",      # 56.3%  EV+0.458%  n=174  (無停損) | -7%: EV+0.684%  n=126
        # ▸ TIER 2 輔助組合
        "B+R",        # 51.4%  EV+0.252%  n=208  (無停損) | -7%: EV+0.274%  n=150
        "B+M+R",      # 49.7%  EV+0.173%  n=165  (無停損)
        # ▸ 大樣本基礎型（訊號頻繁但邊際正 EV）
        "K+M+R",      # 48.4%  EV+0.163%  n=1616
        "K+R",        # 48.3%  EV+0.162%  n=1662
    ],
    # ── 空方（v3：2400 日長期回測 + 停損敏感度驗證，2026-04-24）──────────
    # 篩選門檻：n ≥ 100 樣本充足、正 EV、尾端風險 < 15%
    # ✅ 停損敏感度已測（compare_stoploss.py --direction short）
    #   ▸ BS 家族：-7% 停損 EV 幾乎等同無停損（台股 ±10% 限制下尾端自然封頂 -10.3%）
    #   ▸ FS 家族：-7% 會砍到正常賠錢單，建議 -10% 或不設停損
    # ⚠️ 排除含 GS/KS+RS 的高尾端風險組合（單筆虧損可達 -26% ~ -34%）
    "short3": [
        # ▸ TIER 1 高 EV 明星組合（BS+R 家族，勝率 65%+，-7% 停損最佳）
        "BS+GS",       # 62.7%  EV+1.905%  n=102  max -10.30%  ⭐EV 之王    | -7%: EV+1.823%
        "BS+MS+RS",    # 68.8%  EV+1.672%  n=173  max -10.34%  ⭐勝率之王  | -7%: EV+1.695%
        "BS+RS",       # 67.7%  EV+1.550%  n=189  max -10.34%              | -7%: EV+1.534%
        # ▸ TIER 1 低尾端風險優選（NS 訊號降低尾端風險，-7% 停損不觸發）
        "BS+FS+NS",    # 60.2%  EV+0.971%  n=166  max  -5.66%  ⭐風控最佳  | -7% 等同無停損
        # ▸ TIER 2 FS 家族（與 BS 互補，建議 -10% 或不設停損）
        "FS+MS+RS",    # 61.5%  EV+0.887%  n=117  max -12.30%              | -7%: EV+0.689% ⚠ 建議 -10%
        "FS+RS",       # 62.8%  EV+0.886%  n=145  max -12.30%              | -7%: EV+0.756% ⚠ 建議 -10%
        # ▸ TIER 3 大樣本穩健型（EV 中等但訊號頻繁，-7% 與無停損相近）
        "BS+KS",       # 56.9%  EV+0.589%  n=274  max -10.34%              | -7%: EV+0.583%
        "AS+BS+FS",    # 53.8%  EV+0.570%  n=474  max -11.80%              | -7%: EV+0.458%
    ],
    # ── long3_lean：多方精簡版（2400日回測，參數優化後，2026-05-05）──
    # 參數更新：B 移除 vol_ratio；R BIAS −8→−10；L KD<30→<20
    # 注意：B2多觸發時 B多必然同時觸發（B2 gap≥5% ⊃ B gap≥2%）
    #       故 B2+R = B+R+B2（同一批交易），preset 只保留 B2+X 形式，避免重複
    # 所有組合 EV 均正，2400日樣本充足（⚠️少=樣本<100，統計不可靠供參考）
    "long3_lean": [
        # ▸ TIER 1  EV ≥ 3%（exit_day=3 最佳）
        "B+K+R",      # 77.8%  EV+5.608%  n=482   ⭐充足樣本EV之王
        "B+K+L",      # 79.9%  EV+4.917%  n=134   ⭐⭐勝率最高
        "B+F+L",      # 80.8%  EV+4.532%  n=130   ⭐⭐均量＋KD超賣，勝率王
        "M+R+B2",     # 61.5%  EV+4.269%  n=104   威廉超賣三共振
        "K+L+R",      # 72.9%  EV+4.180%  n=203   純指標三共振
        "B+L+R",      # 73.9%  EV+4.047%  n=253   高勝率三指標
        "B2+K",       # 71.9%  EV+3.979%  n=96    ⚠️少  BB下軌確認（≡ B+K+B2）
        "B+F+R",      # 68.6%  EV+3.698%  n=471   均量擴張＋BIAS超跌
        "B+F+K",      # 67.5%  EV+3.592%  n=391   均量擴張＋布林下軌
        "R+B2",       # 58.0%  EV+3.569%  n=119   （≡ B+R+B2）
        "B+M+R",      # 68.5%  EV+3.412%  n=1111  大樣本穩健
        "B+R",        # 66.7%  EV+3.210%  n=1233  最大樣本基線
        # ▸ TIER 2  EV 2–3%（高頻觸發，大樣本）
        "B+K",        # 63.3%  EV+2.695%  n=1081  高頻觸發
        "K+R",        # 63.4%  EV+2.642%  n=1108  純指標大樣本
        # ▸ TIER 3  大樣本基線（EV 1.5–2%）
        "B2+M",       # 56.3%  EV+1.938%  n=183   ≡ B+M+B2
    ],
    # ── long3_pattern：跨類別精選（2400日回測，stop=0，exit-day=3，2026-05-05）────────
    # 探索訊號分類交叉組合（① × ⑤，② × ⑤⑥，① × ⑥）後篩出正期望值組合
    # 核心特性：納入 K線型態（O/Q）與 MACD/縮量振盪（J/G），補充 long3_lean 盲區
    # ⚠️ 停損設定說明：
    #   B+Q 屬「讓利奔跑」型：不設停損（stop=0）時正EV；加 -7% 停損反轉為負
    #              → 建議搭配寬停損或不設停損使用
    #   R+O / M+G 搭配 -7% 停損 EV 更佳（尾端切割有效）
    # ★ 本 preset 建議 --stop-loss 0（或 --stop-loss 10 折衷）
    # ── P/B2+J 更新說明（2026-05-05）──────────────────────────────
    # P 策略新增條件：④ 實體≥50% K棒範圍；⑤ 上影線<50% 實體
    # B+P EV 由 +0.284% → -0.445%（gap+強三兵=過度延伸）→ 已移除
    # B2+J 更新後 EV 由 +0.110% → -0.083%（轉負）→ 已移除
    "long3_pattern": [
        # ▸ TIER 1  EV ≥ 1%（超賣 × K線型態/振盪）
        "R+O",    # 62.1%  EV+1.592%  n=124   BIAS超跌 × 晨星（② × ⑤）⭐勝率王
        "M+G",    # 44.6%  EV+0.803%  n=177   威廉超賣 × 縮量反彈（② × ⑥）
        # ▸ TIER 2  EV 0.25–1%（缺口 × K線型態/MACD）
        "B+O",    # 54.6%  EV+0.452%  n=141   缺口 × 晨星（① × ⑤）
        "B+J",    # 46.9%  EV+0.353%  n=733   缺口 × MACD黃金交叉（① × ⑥）
        "B+Q",    # 46.1%  EV+0.234%  n=584   缺口 × InsideBar突破（① × ⑤）
    ],
    # long_trend（趨勢延續類）已廢除（2026-05-04）
    # 原因：缺口動能×趨勢延續組合在 2400日回測中 EV 全面偏低（+0.02～+0.61%），
    #       勝率天花板 48%，與 long3_lean 差距過大，不具實用性。
    #       保留空 list 以防舊指令報錯，使用者改用 long3_lean。
    "long_trend": [],
    # ── long_momentum：趨勢跟風精選（2400日回測，stop=0，exit-day=15，2026-05-06）──────
    # 與 long3_lean 的根本差異：
    #   long3_lean  = 超賣反彈型（跌多等漲），最佳出場 exit-day=3，EV 高但持有期短
    #   long_momentum = 趨勢跟風型（跟風慣性），最佳出場 exit-day=15，EV 低但需讓趨勢跑
    # 特性說明：
    #   ① exit-day=3 時所有組合幾乎全為負 EV（-0.03～-0.52%）
    #   ② EV 隨持有天數單調遞增，損益平衡點約在 exit-day=5
    #   ③ 與 long3_lean 為互補策略，建議雙軌並行，各自使用各自的 exit-day
    # 等效關係（觸發時必然同步）：
    #   A+D+N ≡ D+N（多頭排列觸發時，均線突破幾乎必然同步）
    #   A+D+F ≡ D+F（同上）
    # ★ 特例 N+J：EV 在 exit-day=10 達到高峰（+0.367%），超過後回落
    #             → 若使用 N+J，改 --exit-day 10
    # ★ 建議搭配 --stop-loss 7（趨勢型停損較寬；比 long3_lean 的 stop=0 稍緊）
    "long_momentum": [
        # ▸ TIER 1  EV ≥ 1.4%（exit-day=15，核心趨勢組合）
        "D+N",    # 48.4%  EV+1.680%  n=5680   突破5日高×多頭排列（≡ A+D+N）⭐
        "N+F",    # 48.6%  EV+1.585%  n=473    多頭排列×均量擴張  ⚠️少
        "D+F",    # 46.6%  EV+1.432%  n=11186  突破5日高×均量擴張（≡ A+D+F）
        "D+J",    # 47.0%  EV+1.420%  n=2510   突破5日高×MACD黃金交叉
        # ▸ TIER 2  EV 1.0–1.4%（大樣本穩健）
        "A+D",    # 48.1%  EV+1.308%  n=26588  均線突破×突破近5日高（大樣本）
        "A+J",    # 45.6%  EV+1.248%  n=4239   均線突破×MACD黃金交叉
        "A+N",    # 47.6%  EV+1.164%  n=20963  均線突破×多頭排列（大樣本）
        "D+Q",    # 47.4%  EV+1.138%  n=2224   突破5日高×InsideBar突破
        "E+N",    # 52.5%  EV+1.095%  n=398    強勢連漲×多頭排列  ⚠️少 勝率最高
        "A+F",    # 46.3%  EV+1.098%  n=37983  均線突破×均量擴張（最大樣本）
        "D+N+F",  # 48.1%  EV+1.005%  n=426    突破×趨勢×均量三共振  ⚠️少
        # ▸ TIER 3  特殊：exit-day=10 最佳（不適用15日出場）
        "N+J",    # 46.2%  EV+0.367%@d10  多頭排列×MACD（★改用 --exit-day 10）
    ],
    # ── short3_lean：空方精簡版（從 short3 2400日回測去除多餘組合後保留 14 個）──
    # 移除原則：
    #   ① 三策略 EV < 子兩策略 EV（加了策略反而變差）
    #      → FS+MS+RS、BS+KS+MS、BS+CS+KS、BS+DS+FS、BS+CS+DS、AS+FS+GS、KS+MS+RS、KS+LS+RS 全移除
    #   ② 兩策略已被 EV 更高的三策略版本涵蓋
    #      → BS+FS/BS+NS/BS+DS/AS+GS/DS+GS 移除（被 BS+FS+NS、BS+DS+NS、AS+DS+GS 取代）
    # 停損建議：BS 家族 -7%；FS 家族建議 -10% 或不設停損（尾端可達 -12%）
    "short3_lean": [
        # ▸ TIER 1 高 EV 明星組合（EV ≥ 1%，勝率 60%+）
        "BS+GS",       # 63.0%  EV+1.878%  n=108  ⭐EV 之王
        "BS+MS+RS",    # 65.3%  EV+1.676%  n=170  ⭐勝率之王
        "BS+RS",       # 64.3%  EV+1.556%  n=185  BS+MS+RS 較寬版（無 MS 條件）
        "BS+FS+NS",    # 60.5%  EV+0.999%  n=172  ⭐風控最佳（max -5.66%）
        # ▸ TIER 2 良好組合（EV 0.5–1.0%）
        "BS+FS+JS",    # 55.2%  EV+0.839%  n=105  BS+FS 強化版（+JS 過濾）
        "AS+DS+GS",    # 55.3%  EV+0.780%  n=228  取代 AS+GS / DS+GS
        "FS+RS",       # 61.1%  EV+0.767%  n=144  ⚠ 建議 -10% 停損（尾端 -12%）
        "BS+KS",       # 56.3%  EV+0.561%  n=279  +MS/+CS 版均劣於此
        "BS+OS",       # 55.2%  EV+0.531%  n=116
        # ▸ TIER 3 選擇性保留（EV 0.3–0.5%，有特定優勢）
        "BS+DS+NS",    # 56.6%  EV+0.460%  n=389  取代 BS+DS / BS+NS
        "CS+FS+JS",    # 47.1%  EV+0.415%  n=136
        "AS+BS+DS",    # 51.6%  EV+0.407%  n=1256 最大樣本，統計穩健
        "BS+CS+JS",    # 53.1%  EV+0.371%  n=147
        "BS+MS",       # 49.9%  EV+0.333%  n=678  大樣本基線
    ],
}

# ── all2：long2 + short2 + 本次回測所有正EV且樣本充足的補充組合 ────
# （long2 / short2 保留精選核心；all2 擴展為完整正EV清單，掃股用）
_long2_extra = [
    # 動能 + 超跌 補充
    "C+L+R",      # 60.0%  EV+1.529%  n=10
    "F+N+O",      # 52.6%  EV+0.819%  n=19
    "B+C+M",      # 51.7%  EV+0.269%  n=29
    "B+G",        # 40.7%  EV+0.132%  n=27
    "B+L",        # 50.0%  EV+0.093%  n=26
    "M+Q",        # 46.2%  EV+0.078%  n=13
    "B+L+M",      # 50.0%  EV+0.063%  n=24
    "D+N+O",      # 44.4%  EV+0.017%  n=45
    "A+D+I",      # 47.5%  EV+0.016%  n=40
]
_short2_extra = [
    # BS 補充
    "BS+CS+JS",   # 80.0%  EV+2.433%  n=10
    "BS+KS+RS",   # 69.2%  EV+2.413%  n=13
    "BS+KS+MS",   # 66.7%  EV+2.165%  n=21
    "BS+DS+FS",   # 50.0%  EV+0.870%  n=12
    # FS/CS 補充
    "CS+FS+MS",   # 70.0%  EV+2.099%  n=10
    "FS+JS",      # 64.3%  EV+1.545%  n=14
    "FS+OS",      # 81.8%  EV+1.145%  n=11
    # 純超買確認補充
    "KS+LS+OS",   # 80.0%  EV+1.641%  n=10
    "KS+OS+RS",   # 66.7%  EV+1.486%  n=12
    "IS+KS+LS",   # 73.7%  EV+1.305%  n=19
    "MS+OS+RS",   # 63.2%  EV+1.146%  n=38
    "JS+MS",      # 58.1%  EV+0.802%  n=43
    "LS+MS+OS",   # 55.8%  EV+0.768%  n=43
    "AS+MS",      # 53.8%  EV+0.724%  n=13
    "LS+OS+RS",   # 58.3%  EV+0.674%  n=24
    "CS+JS+MS",   # 60.9%  EV+0.551%  n=23
]
COMBO_PRESETS["all2"] = (
    COMBO_PRESETS["long2"]  + _long2_extra +
    COMBO_PRESETS["short2"] + _short2_extra
)

# ── all3_lean：long3_lean + short3_lean 精簡合集 ──
COMBO_PRESETS["all3_lean"] = (
    COMBO_PRESETS["long3_lean"] + COMBO_PRESETS["short3_lean"]
)

# ══════════════════════════════════════════════
# 各 preset 的建議出場日（--combo 模式下使用）
# 來源：2400 日回測 exit-day 1/2/3 敏感度分析（2026-04-25）
# --exit-day 手動指定時永遠優先於此設定
# ══════════════════════════════════════════════
PRESET_EXIT_DAY: dict[str, int] = {
    # 多方：exit-day=3 全面最佳（K+L+R EV +0.63→+3.01%，K+R +0.14→+1.78%）
    "long3_lean":    3,
    # 跨類別：exit-day=3，stop=0 建議（B+P/B+Q 型態類讓利型）
    "long3_pattern": 3,
    # long_trend 已廢除，保留 key 避免舊指令報錯
    "long_trend":  3,
    # 跟風趨勢：exit-day=15 最佳；N+J 例外需 exit-day=10（手動指定）
    # EV 在 exit-day=3 全為負，必須持有才能兌現趨勢慣性
    "long_momentum": 15,
    # 空方：exit-day=1 整體最穩
    # （BS+GS/AS+DS+GS 在 D2/D3 變負；D1 收盤含漲跌停板邏輯最能保留獲利）
    "short3_lean": 1,
    # 混合：交由長空各自決定（見下方 _run_long/_run_short 分別查表）
    "all3_lean":   None,
    # 其他 preset 未設定 → fallback 到 --exit-day 預設值（2）
}


# ── allforall：多空精選合集（2026-05-05 更新）────────────────────────
# 包含：long3_lean（最新精選）+ long3 + long2 + short3_lean + short3 + short2 + short1
# ⚠️ 移除 long1：全為負 EV 的 v1 舊版組合，且大量與已廢除的 long_trend 重疊
#                （B+E+F, B+F, B+E, B+D 等趨勢型負 EV 組合一併清除）
# ⚠️ 移除 long_trend：EV 偏低（+0.03～+0.61%），已於 2026-05-04 廢除
# 自動去重，保留首次出現順序（越新越優先）
def _dedup_keep_order(seq: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out

COMBO_PRESETS["allforall"] = _dedup_keep_order(
    # 多方：momentum（跟風15日）→ pattern（跨類別）→ lean（精選）→ v3 → v2
    COMBO_PRESETS["long_momentum"] + COMBO_PRESETS["long3_pattern"] + COMBO_PRESETS["long3_lean"] + COMBO_PRESETS["long3"] + COMBO_PRESETS["long2"] +
    # 空方：lean（最新精選）→ v3 → v2 → v1
    COMBO_PRESETS["short3_lean"] + COMBO_PRESETS["short3"] + COMBO_PRESETS["short2"] + COMBO_PRESETS["short1"]
)


# ══════════════════════════════════════════════════════════════
# COMBO_META：各組合回測指標（WR%, EV%, n, 備注）
# 來源：2400日回測，stop=0，exit-day=3（多方）/ exit-day=1（空方），2026-05-05
# 格式：combo -> (win_rate_pct, ev_pct, n, note)
# ══════════════════════════════════════════════════════════════
COMBO_META: dict[str, tuple] = {
    # ── long3_lean（參數：B無量限，R BIAS<-10，L KD<20）──────────────
    "B+K+R":   (77.8, 5.608,  482, "充足樣本EV之王"),
    "B+K+L":   (79.9, 4.917,  134, "勝率最高"),
    "B+F+L":   (80.8, 4.532,  130, "均量＋KD超賣，勝率王"),
    "M+R+B2":  (61.5, 4.269,  104, "威廉超賣三共振"),
    "K+L+R":   (72.9, 4.180,  203, "純指標三共振"),
    "B+L+R":   (73.9, 4.047,  253, "高勝率三指標"),
    "B2+K":    (71.9, 3.979,   96, "BB下軌確認（≡ B+K+B2）"),
    "B+F+R":   (68.6, 3.698,  471, "均量擴張＋BIAS超跌"),
    "B+F+K":   (67.5, 3.592,  391, "均量擴張＋布林下軌"),
    "R+B2":    (58.0, 3.569,  119, "≡ B+R+B2"),
    "B+M+R":   (68.5, 3.412, 1111, "大樣本穩健"),
    "B+R":     (66.7, 3.210, 1233, "最大樣本基線"),
    "B+K":     (63.3, 2.695, 1081, "高頻觸發"),
    "K+R":     (63.4, 2.642, 1108, "純指標大樣本"),
    "B2+M":    (56.3, 1.938,  183, "≡ B+M+B2"),
    # ── long3_pattern（stop=0 建議）────────────────────────────────
    "R+O":     (62.1, 1.592,  124, "BIAS超跌×晨星"),
    "M+G":     (44.6, 0.803,  177, "威廉超賣×縮量反彈"),
    "B+O":     (54.6, 0.452,  141, "缺口×晨星"),
    "B+J":     (46.9, 0.353,  733, "缺口×MACD黃金交叉"),
    "B+Q":     (46.1, 0.234,  584, "缺口×InsideBar突破"),
    # ── long_momentum（exit-day=15，除 N+J 用 exit-day=10）───────────
    "D+N":   (48.4, 1.680, 5680,  "突破5日高×多頭排列（≡ A+D+N）"),
    "N+F":   (48.6, 1.585,  473,  "多頭排列×均量擴張"),
    "D+F":   (46.6, 1.432, 11186, "突破5日高×均量擴張（≡ A+D+F）"),
    "D+J":   (47.0, 1.420, 2510,  "突破5日高×MACD黃金交叉"),
    "A+D":   (48.1, 1.308, 26588, "均線突破×突破近5日高"),
    "A+J":   (45.6, 1.248, 4239,  "均線突破×MACD黃金交叉"),
    "A+N":   (47.6, 1.164, 20963, "均線突破×多頭排列"),
    "D+Q":   (47.4, 1.138, 2224,  "突破5日高×InsideBar突破"),
    "E+N":   (52.5, 1.095,  398,  "強勢連漲×多頭排列，勝率最高"),
    "A+F":   (46.3, 1.098, 37983, "均線突破×均量擴張，最大樣本"),
    "D+N+F": (48.1, 1.005,  426,  "突破×趨勢×均量三共振"),
    "N+J":   (46.2, 0.367,  340,  "多頭排列×MACD（exit-day=10 最佳）"),
    # ── short3_lean（exit-day=1）────────────────────────────────────
    "BS+GS":    (63.0, 1.878, 108, "EV之王"),
    "BS+MS+RS": (65.3, 1.676, 170, "勝率之王"),
    "BS+RS":    (64.3, 1.556, 185, "BS+MS+RS寬版"),
    "BS+FS+NS": (60.5, 0.999, 172, "風控最佳"),
    "BS+FS+JS": (55.2, 0.839, 105, "BS+FS強化版"),
    "AS+DS+GS": (55.3, 0.780, 228, "取代AS+GS/DS+GS"),
    "FS+RS":    (61.1, 0.767, 144, "建議-10%停損"),
    "BS+KS":    (56.3, 0.561, 279, ""),
    "BS+OS":    (55.2, 0.531, 116, ""),
    "BS+DS+NS": (56.6, 0.460, 389, "取代BS+DS/BS+NS"),
    "CS+FS+JS": (47.1, 0.415, 136, ""),
    "AS+BS+DS": (51.6, 0.407, 1256, "最大樣本"),
    "BS+CS+JS": (53.1, 0.371, 147, ""),
    "BS+MS":    (49.9, 0.333, 678, "大樣本基線"),
}

# ── 各 preset 的 tier 分界（供 --list 輸出用）────────────────────────
# 格式：[(tier_label, ev_threshold), ...]  由高到低，最後一個 threshold=0
PRESET_TIERS: dict[str, list[tuple[str, float]]] = {
    "long3_lean":    [
        ("▸ TIER 1  EV ≥ 3.0%  （精選，exit-day=3）", 3.0),
        ("▸ TIER 2  EV 2–3%    （高頻大樣本）",        2.0),
        ("▸ TIER 3  EV < 2%    （基線參考）",          0.0),
    ],
    "long3_pattern": [
        ("▸ TIER 1  EV ≥ 1.0%  （超賣×型態）",         1.0),
        ("▸ TIER 2  EV < 1.0%  （缺口×型態/振盪）",    0.0),
    ],
    "long_momentum": [
        ("▸ TIER 1  EV ≥ 1.4%  （核心跟風，exit-day=15）", 1.4),
        ("▸ TIER 2  EV 1.0–1.4%（大樣本穩健）",             1.0),
        ("▸ TIER 3  特殊        （exit-day=10 最佳）",       0.0),
    ],
    "short3_lean": [
        ("▸ TIER 1  EV ≥ 1.0%  （高EV精選）",          1.0),
        ("▸ TIER 2  EV 0.5–1%  （良好組合）",          0.5),
        ("▸ TIER 3  EV < 0.5%  （選擇性保留）",        0.0),
    ],
}


def resolve_preset(preset: str) -> list[str]:
    """
    將 --preset 名稱展開成組合清單。
    支援逗號分隔的多 preset，回傳合併（去重保序）清單。
    例如：--preset long3_lean,short3_lean
    不存在時印出可用清單並回傳空 list。
    """
    names = [p.strip().lower() for p in preset.split(",") if p.strip()]
    merged: list[str] = []
    seen:   set[str]  = set()
    ok = True
    for name in names:
        if name not in COMBO_PRESETS:
            print(f"  ❌ 找不到 preset '{name}'，可用清單：")
            for pname in COMBO_PRESETS:
                print(f"     {pname}")
            ok = False
            continue
        for combo in COMBO_PRESETS[name]:
            if combo not in seen:
                merged.append(combo)
                seen.add(combo)
    return merged if ok else []


# ══════════════════════════════════════════════
# 漲停 / 跌停價計算（台股最小升降單位）
# ══════════════════════════════════════════════
def _tick_size(price: float) -> float:
    """台股最小升降單位（tick）"""
    if price <   10: return 0.01
    if price <   50: return 0.05
    if price <  100: return 0.10
    if price <  500: return 0.50
    if price < 1000: return 1.00
    return 5.00

def calc_limit_up(prev_close: float) -> float:
    """漲停價：前收 × 1.10，無條件捨去到最小升降單位"""
    raw  = prev_close * 1.10
    tick = _tick_size(raw)
    return round(math.floor(raw / tick) * tick, 2)

def calc_limit_down(prev_close: float) -> float:
    """跌停價：前收 × 0.90，無條件進位到最小升降單位"""
    raw  = prev_close * 0.90
    tick = _tick_size(raw)
    return round(math.ceil(raw / tick) * tick, 2)


# ══════════════════════════════════════════════
# SQLite 資料庫
# ══════════════════════════════════════════════
def init_db() -> sqlite3.Connection:
    """初始化資料庫，建立資料表（如不存在）"""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS scans (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        scan_date   TEXT NOT NULL,          -- 選股日期 YYYY-MM-DD
        code        TEXT NOT NULL,
        name        TEXT,
        close       REAL,
        chg_pct     REAL,
        vol_k       INTEGER,
        vol_ratio   TEXT,
        strategies  TEXT,                   -- 命中策略，逗號分隔
        hit_count   INTEGER,
        direction   TEXT,                   -- 多 / 空
        min_hit     INTEGER,
        -- 次日實際走勢（事後填入）
        next_open   REAL,
        next_close  REAL,
        next_high   REAL,
        next_low    REAL,
        pnl_pct     REAL,                   -- 實際損益%（次日收盤）
        created_at  TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE (scan_date, code) ON CONFLICT REPLACE
    );

    CREATE TABLE IF NOT EXISTS backtests (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        run_date      TEXT NOT NULL,
        strategy      TEXT NOT NULL,
        days          INTEGER,
        min_hit       INTEGER,
        stop_loss     REAL,
        signals       INTEGER,
        win_rate      REAL,
        avg_win       REAL,
        avg_loss      REAL,
        expectancy    REAL,
        max_loss      REAL,
        stop_rate     REAL,
        created_at    TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE INDEX IF NOT EXISTS idx_scans_date ON scans(scan_date);
    CREATE INDEX IF NOT EXISTS idx_scans_code ON scans(code);

    -- 日K快取（避免重複抓 sinopac/twse；可用 sinopac 一次抓滿，後續用 twse 增量補）
    CREATE TABLE IF NOT EXISTS kbars_daily (
        code   TEXT NOT NULL,
        date   TEXT NOT NULL,                  -- YYYY-MM-DD
        open   REAL,
        high   REAL,
        low    REAL,
        close  REAL,
        vol_k  REAL,                           -- 千股 / 張
        volume REAL,                           -- 股數
        source TEXT,                           -- 'sinopac' / 'twse' / 'yfinance'
        updated_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY (code, date)
    );
    CREATE INDEX IF NOT EXISTS idx_kbars_code_date
        ON kbars_daily(code, date);
    """)
    conn.commit()

    # ── Migration：確保 scans 有 UNIQUE(scan_date, code) ──────────────
    # 若現有 DB 無此 index（舊版建立），先清除重複資料再建立
    has_idx = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='index' AND name='idx_scans_unique_date_code'"
    ).fetchone()
    if not has_idx:
        # 保留每組 (scan_date, code) 中 id 最大（最新）的那筆，刪除其餘重複
        conn.execute("""
            DELETE FROM scans
            WHERE id NOT IN (
                SELECT MAX(id) FROM scans GROUP BY scan_date, code
            )
        """)
        conn.execute(
            "CREATE UNIQUE INDEX idx_scans_unique_date_code ON scans(scan_date, code)"
        )
        conn.commit()

    return conn


def save_scan(conn: sqlite3.Connection, scan_date: str, rows: list, min_hit: int):
    """儲存選股結果到 scans 資料表（同日同代號以最新掃描覆寫）"""
    conn.executemany("""
        INSERT OR REPLACE INTO scans
            (scan_date, code, name, close, chg_pct, vol_k, vol_ratio,
             strategies, hit_count, direction, min_hit)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, [
        (scan_date,
         r["代號"], r.get("名稱",""), r["收盤"], r["漲跌幅(%)"],
         r["成交量(張)"], r["量/均量"], r["策略清單"],
         r["命中數"], r["方向"], min_hit)
        for r in rows
    ])
    conn.commit()


def save_backtest_csv(stats: list, days: int, min_hit: int,
                      stop_loss: float) -> str:
    """
    儲存回測結果到 CSV。
    檔名包含日期、天數、門檻、停損，方便比較多次回測。
    """
    date_str = datetime.date.today().strftime("%Y%m%d")
    sl_str   = f"sl{stop_loss}" if stop_loss > 0 else "nosl"
    fname    = f"回測結果_{date_str}_d{days}_hit{min_hit}_{sl_str}.csv"
    df = pd.DataFrame(stats)
    # 加入執行資訊欄
    df.insert(0, "回測日期", datetime.date.today().isoformat())
    df.insert(1, "回測天數", days)
    df.insert(2, "命中門檻", min_hit)
    df.insert(3, "停損(%)",  stop_loss)
    df.to_csv(fname, index=False, encoding="utf-8-sig")
    return fname


def query_history(conn: sqlite3.Connection, code: str = None,
                  limit: int = 20) -> pd.DataFrame:
    """查詢歷史選股記錄"""
    if code:
        df = pd.read_sql(
            "SELECT scan_date,code,name,close,chg_pct,strategies,direction,pnl_pct "
            "FROM scans WHERE code=? ORDER BY scan_date DESC LIMIT ?",
            conn, params=(code, limit)
        )
    else:
        df = pd.read_sql(
            "SELECT scan_date,code,name,close,chg_pct,strategies,direction,hit_count,pnl_pct "
            "FROM scans ORDER BY scan_date DESC, hit_count DESC LIMIT ?",
            conn, params=(limit,)
        )
    return df


# ══════════════════════════════════════════════
# 股票名稱對照
# ══════════════════════════════════════════════
_NAME_MAP: dict = {}

def get_name(code: str) -> str:
    global _NAME_MAP
    if not _NAME_MAP and HAS_TWSTOCK:
        try:
            _NAME_MAP = {c: getattr(s, "name", c)
                         for c, s in twstock.codes.items() if c.isdigit()}
        except Exception:
            pass
    return _NAME_MAP.get(str(code), code)


# ══════════════════════════════════════════════
# 資料新鮮度檢查（共用）
# ══════════════════════════════════════════════
def _check_data_freshness(results: dict, source: str = ""):
    """檢查並印出資料新鮮度警告"""
    last_dates = []
    for df in results.values():
        try:
            last_dates.append(df.index[-1].date())
        except Exception:
            pass
    if not last_dates:
        return
    data_date = max(last_dates)
    today_d   = datetime.date.today()
    delta     = (today_d - data_date).days
    # 台股週一至週五開市；週一 delta==3 表示資料是上週五，屬正常
    is_weekend_gap = (today_d.weekday() == 0 and delta == 3)
    if delta == 0:
        print(f"  📅 資料截至：{data_date}（今日，✅ 最新）")
    elif is_weekend_gap or delta <= 1:
        print(f"  📅 資料截至：{data_date}（上個交易日，✅ 正常）")
    else:
        src_hint = f"{source} 尚未更新今日收盤" if source else "資料可能過時"
        print(f"  ⚠️  資料截至：{data_date}（{delta} 天前）"
              f"— {src_hint}，選股結果為 {data_date} 的訊號！")


# ══════════════════════════════════════════════
# 資料下載
# ══════════════════════════════════════════════
def _get_top_codes(top_n: int = 300) -> list:
    """從 TWSE 取得成交量前 N 名的股票代號（三層 fallback）"""
    # TWSE 憑證有 Missing Subject Key Identifier 問題
    # 優先用 certifi，失敗時改用 verify=False
    ssl_opt = certifi.where() if certifi else True
    ssl_fallback = False  # 若 certifi 仍失敗，切換到 verify=False

    # Layer 1：TWSE 盤後 CSV（最穩定）
    try:
        url = "https://www.twse.com.tw/exchangeReport/STOCK_DAY_ALL?response=open_data"
        for verify in [ssl_opt, False]:  # 先用 certifi，失敗再用 verify=False
            try:
                r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"},
                                 timeout=20, verify=verify)
                break
            except Exception:
                if verify is False:
                    raise
                continue
        df  = pd.read_csv(__import__("io").StringIO(r.text))
        # 欄位：證券代號, 證券名稱, 成交股數, ...
        code_col = next((c for c in df.columns if "代號" in c or "code" in c.lower()), df.columns[0])
        vol_col  = next((c for c in df.columns if "成交" in c and "股" in c), None)
        if vol_col:
            df[vol_col] = pd.to_numeric(df[vol_col].astype(str).str.replace(",",""), errors="coerce")
            df = df.dropna(subset=[vol_col])
            df = df[df[code_col].astype(str).str.match(r"^\d{4}$")]
            df = df.sort_values(vol_col, ascending=False)
            codes = df[code_col].astype(str).str.strip().head(top_n).tolist()
            if codes:
                print(f"  ✅ TWSE CSV：取得 {len(codes)} 檔")
                return codes
    except Exception as e:
        print(f"  ⚠️  TWSE CSV 失敗：{e}")

    # Layer 2：TWSE JSON API
    try:
        url  = "https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?type=ALLBUT0999&response=json"
        for verify in [ssl_opt, False]:
            try:
                r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"},
                                 timeout=20, verify=verify)
                break
            except Exception:
                if verify is False:
                    raise
                continue
        data = r.json()
        rows = data.get("data9", data.get("data5", data.get("data", [])))
        codes = []
        for row in rows:
            try:
                code = str(row[0]).strip()
                vol  = int(str(row[2]).replace(",", ""))
                if code.isdigit() and len(code) == 4:
                    codes.append((code, vol))
            except Exception:
                continue
        if codes:
            codes.sort(key=lambda x: x[1], reverse=True)
            result = [c for c, _ in codes[:top_n]]
            print(f"  ✅ TWSE JSON：取得 {len(result)} 檔")
            return result
    except Exception as e:
        print(f"  ⚠️  TWSE JSON 失敗：{e}")

    # Layer 3：twstock 內建清單（以常見大型股優先排序）
    if HAS_TWSTOCK:
        try:
            # 優先選常見上市股（避免冷門股 yfinance 下載失敗）
            priority = ["2330","2317","2454","2382","2412","2308","2303","3711",
                        "2881","2882","2886","2891","2892","2884","2885","2880",
                        "1301","1303","1326","2002","2886","6505","5880","2207"]
            all_codes = [c for c in twstock.codes if c.isdigit() and len(c) == 4]
            # 把 priority 排前面，其餘補齊
            rest   = [c for c in all_codes if c not in priority]
            result = (priority + rest)[:top_n]
            print(f"  ✅ twstock 備援：取得 {len(result)} 檔（優先大型股）")
            return result
        except Exception:
            pass

    print("  ❌ 三層 fallback 均失敗，請使用 --codes 指定股票")
    return []


def fetch_data(codes: list, days: int = DEFAULT_DAYS) -> dict:
    """下載歷史 K 線資料"""
    if not HAS_YF:
        print("  ❌ 請安裝：pip install yfinance")
        return {}

    # 盤中保護：yfinance end 設為今日即可；dropna 本來就會過濾掉盤中不完整列
    _tz_tw    = datetime.timezone(datetime.timedelta(hours=8))
    _now_tw   = datetime.datetime.now(_tz_tw)
    _open_tw  = _now_tw.replace(hour=9,  minute=0,  second=0, microsecond=0)
    _close_tw = _now_tw.replace(hour=13, minute=30, second=0, microsecond=0)
    if _open_tw <= _now_tw < _close_tw:
        print(f"  ⏰ 現在 {_now_tw.strftime('%H:%M')} 台灣時間，盤中執行"
              f"——使用昨日收盤訊號選股")

    buf     = days + 30
    end     = datetime.date.today() + datetime.timedelta(days=1)
    start   = datetime.date.today() - datetime.timedelta(days=buf)
    def _to_ticker(code: str) -> str:
        """判斷上市(.TW)或上櫃(.TWO)"""
        if HAS_TWSTOCK:
            try:
                s = twstock.codes.get(code)
                if s:
                    mtype = getattr(s, "market_type", "") or ""
                    if "OTC" in mtype or "上櫃" in mtype:
                        return f"{code}.TWO"
                    return f"{code}.TW"
            except Exception:
                pass
        # fallback：代號首字 4 或 0 通常是上櫃
        return f"{code}.TWO" if code.startswith(("4","0")) else f"{code}.TW"

    tickers  = [_to_ticker(c) for c in codes]
    # ticker → code 反查表
    t2c = {t: t.replace(".TW","").replace("O","") if t.endswith(".TWO")
           else t.replace(".TW","") for t in tickers}
    t2c = {t: t.replace(".TWO","").replace(".TW","") for t in tickers}
    # 抑制 yfinance 下市股票警告（如 "$3128.TW: possibly delisted"）
    import warnings, logging
    logging.getLogger("yfinance").setLevel(logging.CRITICAL)
    warnings.filterwarnings("ignore", category=UserWarning)

    results = {}
    batch   = 50

    def _download_batch(sub_tickers, sub_t2c):
        """下載一批 ticker，回傳 {code: df}"""
        partial = {}
        try:
            raw = yf.download(sub_tickers, start=str(start), end=str(end),
                              group_by="ticker", auto_adjust=False,
                              progress=False, threads=True)
            for ticker in sub_tickers:
                code = sub_t2c.get(ticker, ticker.replace(".TWO","").replace(".TW",""))
                try:
                    if isinstance(raw.columns, pd.MultiIndex):
                        if ticker in raw.columns.get_level_values(0):
                            df = raw[ticker].copy()
                        else:
                            continue
                    else:
                        df = raw.copy()
                    df.dropna(subset=["Close"], inplace=True)
                    if len(df) >= 22:
                        df["Vol_K"] = df["Volume"] / 1000
                        partial[code] = df
                except Exception:
                    pass
        except Exception as e:
            print(f"  ⚠️  批次下載失敗：{e}")
        return partial

    print(f"  📥 下載 {len(codes)} 檔資料（{start} ~ {end}）...")
    for i in range(0, len(tickers), batch):
        sub = tickers[i:i+batch]
        sub_t2c = {t: t2c[t] for t in sub if t in t2c}
        results.update(_download_batch(sub, sub_t2c))
        time.sleep(0.3)

    # ── 自動換 suffix 重試（修正 .TW / .TWO 誤判，如 4540 等）──
    missing = [c for c in codes if c not in results]
    if missing:
        def _alt_ticker(code: str) -> str:
            orig = _to_ticker(code)
            return f"{code}.TW" if orig.endswith(".TWO") else f"{code}.TWO"

        alt_tickers = [_alt_ticker(c) for c in missing]
        alt_t2c     = {t: t.replace(".TWO","").replace(".TW","") for t in alt_tickers}
        print(f"  🔄 換 suffix 重試 {len(missing)} 檔（.TW ↔ .TWO）：{', '.join(missing[:10])}"
              + ("..." if len(missing) > 10 else ""))
        for i in range(0, len(alt_tickers), batch):
            sub = alt_tickers[i:i+batch]
            sub_t2c = {t: alt_t2c[t] for t in sub}
            results.update(_download_batch(sub, sub_t2c))
            time.sleep(0.3)

    print(f"  ✅ 成功取得 {len(results)} 檔")
    _check_data_freshness(results, source="yfinance")
    return results


# ══════════════════════════════════════════════
# TWSE / TPEX 官方日K 備援下載
# ══════════════════════════════════════════════
import warnings as _ssl_warn
_ssl_warn.filterwarnings("ignore", message="Unverified HTTPS")

def _fetch_twse_month(code: str, yyyymm: str,
                      session: "requests.Session | None" = None) -> list:
    """抓 TWSE 上市股票單月日K，回傳 list of row-dict"""
    url = (f"https://www.twse.com.tw/exchangeReport/STOCK_DAY"
           f"?response=json&date={yyyymm}01&stockNo={code}")
    getter = session.get if session else requests.get
    try:
        r = getter(url, timeout=12, verify=False)
        data = r.json()
        if data.get("stat") != "OK":
            return []
        rows = []
        for row in data.get("data", []):
            try:
                p = row[0].replace("/", "-").split("-")
                rows.append({
                    "date":  f"{int(p[0])+1911}-{p[1]}-{p[2]}",
                    "Open":  float(row[3].replace(",", "")),
                    "High":  float(row[4].replace(",", "")),
                    "Low":   float(row[5].replace(",", "")),
                    "Close": float(row[6].replace(",", "")),
                    "Vol_K": int(row[1].replace(",", "")) // 1000,
                })
            except Exception:
                continue
        return rows
    except Exception:
        return []


def _fetch_tpex_month(code: str, yyyymm: str,
                      session: "requests.Session | None" = None) -> list:
    """抓 TPEX 上櫃股票單月日K，回傳 list of row-dict"""
    roc_y = int(yyyymm[:4]) - 1911
    mon   = yyyymm[4:]
    url = (f"https://www.tpex.org.tw/web/stock/aftertrading/daily_trading_info/"
           f"st43_result.php?l=zh-tw&d={mon}/{roc_y}&s={code}&o=json")
    getter = session.get if session else requests.get
    try:
        r = getter(url, timeout=12, verify=False)
        data = r.json()
        if not data.get("iTotalRecords", 0):
            return []
        rows = []
        for row in data.get("aaData", []):
            try:
                p = row[0].replace("/", "-").split("-")
                rows.append({
                    "date":  f"{int(p[0])+1911}-{p[1]}-{p[2]}",
                    "Open":  float(row[2].replace(",", "")),
                    "High":  float(row[3].replace(",", "")),
                    "Low":   float(row[4].replace(",", "")),
                    "Close": float(row[5].replace(",", "")),
                    "Vol_K": int(row[1].replace(",", "")) // 1000,
                })
            except Exception:
                continue
        return rows
    except Exception:
        return []


def _fetch_official_month(code: str, yyyymm: str,
                          session: "requests.Session | None" = None) -> list:
    """TWSE 優先，失敗改 TPEX（單月）"""
    rows = _fetch_twse_month(code, yyyymm, session=session)
    return rows if rows else _fetch_tpex_month(code, yyyymm, session=session)


def _rows_to_df(rows: list, start_date: datetime.date,
                end_date: datetime.date):
    """把 row-dict list 整理成 DatetimeIndex DataFrame，失敗回傳 None"""
    if not rows:
        return None
    df = pd.DataFrame(rows)
    df = df[(df["date"] >= str(start_date)) & (df["date"] <= str(end_date))]
    if df.empty:
        return None
    df = df.sort_values("date").drop_duplicates("date").reset_index(drop=True)
    df.index = pd.to_datetime(df["date"])
    df.index.name = None
    df = df.drop(columns=["date"])
    df["Volume"] = df["Vol_K"] * 1000
    return df


def _fetch_official_daily(code: str,
                           start_date: datetime.date,
                           end_date: datetime.date):
    """
    單一股票：循序抓所有月份後整理成 DataFrame。
    （fetch_data_sinopac 的 Step 3 改用扁平並行；此函式保留給 --diagnose 等單股情境）
    """
    months: list = []
    cur = start_date.replace(day=1)
    while cur <= end_date:
        months.append(cur.strftime("%Y%m"))
        cur = (cur + datetime.timedelta(days=32)).replace(day=1)

    all_rows: list = []
    for ym in months:
        all_rows.extend(_fetch_official_month(code, ym))

    return _rows_to_df(all_rows, start_date, end_date)


# ══════════════════════════════════════════════
# TWSE/TPEX 純官方資料下載（不需 shioaji 帳號、不依賴 yfinance）
# ══════════════════════════════════════════════
def fetch_data_twse(codes: list, days: int = DEFAULT_DAYS,
                    workers: int = 20) -> dict:
    """
    使用 TWSE / TPEX 官方公開 API 下載歷史日 K。
      - 上市股 → twse.com.tw exchangeReport/STOCK_DAY
      - 上櫃股 → tpex.org.tw st43_result.php
      - 不需登入帳號，無需 yfinance
      - API 限制：每月 1 次請求 / 檔，故大量股票會以 (code × month) 扁平並行加速

    Args:
      codes:    股票代號清單（不含 .TW / .TWO）
      days:     回測/分析天數（會額外加 buffer）
      workers:  並行 worker 數，預設 20（不要太高，TWSE 會擋）

    Returns:
      dict[code -> DataFrame(OHLCV)]
    """
    if not codes:
        return {}

    _tz_tw    = datetime.timezone(datetime.timedelta(hours=8))
    _now_tw   = datetime.datetime.now(_tz_tw)
    _today    = _now_tw.date()
    _open_tw  = _now_tw.replace(hour=9,  minute=0,  second=0, microsecond=0)
    _close_tw = _now_tw.replace(hour=14, minute=30, second=0, microsecond=0)

    # TWSE 通常 14:00 後才有完整當日資料；保險起見以昨日為終點
    if _open_tw <= _now_tw < _close_tw:
        print(f"  ⏰ 現在 {_now_tw.strftime('%H:%M')} 台灣時間，盤中執行"
              f"——使用昨日（{_today - datetime.timedelta(days=1)}）作為資料終點")
        hist_end = _today - datetime.timedelta(days=1)
    else:
        hist_end = _today

    buf        = days + 40
    hist_start = _today - datetime.timedelta(days=buf)

    # 列出所有需要抓的月份
    months: list = []
    cur = hist_start.replace(day=1)
    while cur <= hist_end:
        months.append(cur.strftime("%Y%m"))
        cur = (cur + datetime.timedelta(days=32)).replace(day=1)

    n_codes  = len(codes)
    n_months = len(months)
    n_tasks  = n_codes * n_months
    _tw      = max(1, min(workers, n_tasks))
    est_sec  = n_tasks * 0.4 / _tw

    print(f"  📥 TWSE/TPEX 官方 API：{n_codes} 檔 × {n_months} 個月 = "
          f"{n_tasks} 筆請求（{_tw} workers，估計 {est_sec:.0f} 秒）")
    print(f"     區間 {hist_start} ~ {hist_end}")

    from collections import defaultdict
    code_rows: dict = defaultdict(list)
    counter = {"done": 0}
    lock    = threading.Lock()
    report_every = max(1, n_tasks // 20)

    # thread-local Session：每 thread 一條 keep-alive，減少 TLS handshake
    _tls = threading.local()

    def _get_session() -> requests.Session:
        if not hasattr(_tls, "session"):
            _tls.session = requests.Session()
            _tls.session.headers.update({"User-Agent": "Mozilla/5.0"})
            _tls.session.verify = False
        return _tls.session

    def _fetch_task(task: tuple) -> None:
        code, ym = task
        sess = _get_session()
        rows = _fetch_official_month(code, ym, session=sess)
        with lock:
            counter["done"] += 1
            done_now = counter["done"]
        if rows:
            code_rows[code].extend(rows)
        if done_now % report_every == 0 or done_now == n_tasks:
            pct = done_now * 100 // n_tasks
            print(f"  ── TWSE 進度：{done_now}/{n_tasks} 筆（{pct}%）")

    # 月份優先排序：避免對同一股票連續轟炸
    tasks = [(c, ym) for ym in months for c in codes]
    with ThreadPoolExecutor(max_workers=_tw) as pool:
        list(pool.map(_fetch_task, tasks))

    # 彙整成 dict[code -> DataFrame]
    results: dict = {}
    for code in codes:
        df = _rows_to_df(code_rows.get(code, []), hist_start, hist_end)
        if df is not None and len(df) >= 22:
            results[code] = df

    print(f"  ✅ TWSE/TPEX 完成：{len(results)}/{n_codes} 檔成功")
    if results:
        _check_data_freshness(results, source="TWSE/TPEX")
    return results


# ══════════════════════════════════════════════
# SQLite K bar 快取（kbars_daily 表）
# 用法：先用 sinopac 抓滿，後續用 twse 增量補
# ══════════════════════════════════════════════
def load_kbars_cache(conn: sqlite3.Connection, codes: list,
                     start_date: datetime.date,
                     end_date: datetime.date) -> dict:
    """
    從 SQLite 讀取 [start_date, end_date] 區間的快取日 K，
    回傳 dict[code -> DataFrame]（DatetimeIndex，欄位 OHLC + Vol_K + Volume）。
    沒命中或無資料的 code 不會出現在 dict 中。
    """
    if not codes:
        return {}
    results: dict = {}
    # SQLite IN 子句長度限制（SQLITE_LIMIT_VARIABLE_NUMBER），分批讀
    chunk = 500
    for i in range(0, len(codes), chunk):
        sub = codes[i:i+chunk]
        ph  = ",".join("?" * len(sub))
        rows = conn.execute(
            f"""
            SELECT code, date, open, high, low, close, vol_k, volume
              FROM kbars_daily
             WHERE code IN ({ph})
               AND date >= ?
               AND date <= ?
             ORDER BY code, date
            """,
            sub + [str(start_date), str(end_date)],
        ).fetchall()
        if not rows:
            continue
        from collections import defaultdict
        bucket: dict = defaultdict(list)
        for r in rows:
            bucket[r[0]].append(r)
        for code, code_rows in bucket.items():
            df = pd.DataFrame(
                code_rows,
                columns=["code", "date", "Open", "High", "Low",
                         "Close", "Vol_K", "Volume"],
            )
            df.index = pd.to_datetime(df["date"])
            df.index.name = None
            df = df.drop(columns=["code", "date"])
            results[code] = df
    return results


def save_kbars_cache(conn: sqlite3.Connection, results: dict,
                     source: str = "") -> int:
    """
    把 fetch 回來的 dict[code -> df] 寫入 kbars_daily（INSERT OR REPLACE）。
    回傳寫入筆數。
    """
    rows: list = []
    for code, df in results.items():
        if df is None or len(df) == 0:
            continue
        for idx, row in df.iterrows():
            try:
                rows.append((
                    str(code),
                    idx.strftime("%Y-%m-%d"),
                    float(row.get("Open",  float("nan"))),
                    float(row.get("High",  float("nan"))),
                    float(row.get("Low",   float("nan"))),
                    float(row.get("Close", float("nan"))),
                    float(row.get("Vol_K", 0) or 0),
                    float(row.get("Volume", 0) or 0),
                    source,
                ))
            except Exception:
                continue
    if rows:
        conn.executemany(
            """
            INSERT OR REPLACE INTO kbars_daily
                (code, date, open, high, low, close, vol_k, volume, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
    return len(rows)


def _merge_into_cache(cached: dict, fresh: dict) -> dict:
    """把新抓的 fresh 合併進 cached（同日期以 fresh 為準），回傳合併後的 dict。"""
    out = dict(cached)
    for code, df in fresh.items():
        if df is None or len(df) == 0:
            continue
        if code in out and out[code] is not None and len(out[code]) > 0:
            combined = pd.concat([out[code], df])
            combined = combined[~combined.index.duplicated(keep="last")]
            out[code] = combined.sort_index()
        else:
            out[code] = df.sort_index()
    return out


def fetch_data_cached(codes: list, days: int, datasource: str = "sinopac",
                      *,
                      sj_key: str = "", sj_secret: str = "",
                      cache_only: bool = False,
                      refresh: bool = False) -> dict:
    """
    SQLite 快取 + 增量補抓的協調函式。

    流程：
      1. 從 kbars_daily 讀取 [today - days - 30, today] 的快取
      2. 分類每個 code：
         - 完全沒資料 / 不足  → 全量補（呼叫 datasource 全抓）
         - 末日 < 昨日       → 增量補（只抓 max(latest)+1 → today）
         - 已最新            → 不動
      3. 補完的資料寫回快取，與快取合併後回傳

    Args:
      codes:      股票代號清單
      days:       回測/分析所需的天數
      datasource: 'sinopac' / 'twse' / 'yfinance'
      sj_key/secret: sinopac 金鑰（datasource=sinopac 時必填）
      cache_only: 只讀快取，不打遠端（即使有缺口）
      refresh:    忽略快取，強制重抓全部

    Returns:
      dict[code -> DataFrame]
    """
    if not codes:
        return {}

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")

    today        = datetime.date.today()
    target_start = today - datetime.timedelta(days=days + 30)

    # ── Step 1：讀快取 ──────────────────────────
    if refresh:
        cached: dict = {}
        print(f"  🔄 --refresh-cache：忽略快取，全部重抓")
    else:
        cached = load_kbars_cache(conn, codes, target_start, today)
        if cached:
            sample_latest = max(
                df.index[-1].date() for df in cached.values()
            )
            print(f"  📦 SQLite 快取命中 {len(cached)}/{len(codes)} 檔"
                  f"  最新日期 {sample_latest}")
        else:
            print(f"  📦 SQLite 快取未命中（共 {len(codes)} 檔）")

    # ── Step 2：分類 ─────────────────────────────
    full_codes: list = []      # 缺資料（要全量補）
    inc_codes: list  = []      # 只缺最近幾天（要增量補）
    inc_from: datetime.date | None = None  # 增量區間起點
    min_required_rows = max(22, min(days, 250) // 2)  # 太少就視為不足

    # 門檻設為「今日」：只要沒有今天的資料就嘗試增量補抓。
    # 這樣收盤後執行 --scan 可自動抓當天資料，無需 --refresh-cache。
    # 若市場尚未收盤或今天是非交易日，遠端 API 回傳空結果，快取保持不變。
    for code in codes:
        df = cached.get(code)
        if df is None or len(df) < min_required_rows:
            full_codes.append(code)
            continue
        latest = df.index[-1].date()
        if latest < today:          # 改：只要不含今日資料就增量補
            inc_codes.append(code)
            cand = latest + datetime.timedelta(days=1)
            if inc_from is None or cand < inc_from:
                inc_from = cand

    already_fresh = len(codes) - len(full_codes) - len(inc_codes)
    print(f"  └── 全量補：{len(full_codes)} 檔  │  "
          f"增量補：{len(inc_codes)} 檔"
          + (f"（從 {inc_from}）" if inc_from else "")
          + f"  │  今日已最新：{already_fresh} 檔")

    if cache_only:
        if full_codes or inc_codes:
            print(f"  ⏸  --cache-only：不打遠端，直接回傳快取（缺 "
                  f"{len(full_codes)+len(inc_codes)} 檔不補）")
        conn.close()
        return cached

    # ── Step 3：全量補 ──────────────────────────
    if full_codes:
        if datasource == "sinopac":
            if not sj_key or not sj_secret:
                print(f"  ⚠️  缺 sinopac 金鑰，{len(full_codes)} 檔全量補改用 TWSE")
                fresh = fetch_data_twse(full_codes, days)
                src_tag = "twse"
            else:
                fresh = fetch_data_sinopac(full_codes, days, sj_key, sj_secret)
                src_tag = "sinopac"
        elif datasource == "twse":
            fresh = fetch_data_twse(full_codes, days)
            src_tag = "twse"
        else:
            fresh = fetch_data(full_codes, days)
            src_tag = "yfinance"

        n_saved = save_kbars_cache(conn, fresh, source=src_tag)
        print(f"  💾 全量補寫入快取 {n_saved} 筆 K bar（source={src_tag}）")
        cached = _merge_into_cache(cached, fresh)

    # ── Step 4：增量補 ──────────────────────────
    if inc_codes and inc_from is not None:
        gap_days = (today - inc_from).days + 5  # 抓多一點點 buffer
        print(f"  🔄 增量補抓 {len(inc_codes)} 檔 × 約 {gap_days} 天 "
              f"（{inc_from} ~ {today}）")

        if datasource == "sinopac":
            if not sj_key or not sj_secret:
                print(f"  ⚠️  缺 sinopac 金鑰，增量補改用 TWSE")
                fresh = fetch_data_twse(inc_codes, gap_days)
                src_tag = "twse"
            else:
                fresh = fetch_data_sinopac(inc_codes, gap_days, sj_key, sj_secret)
                src_tag = "sinopac"
        elif datasource == "twse":
            fresh = fetch_data_twse(inc_codes, gap_days)
            src_tag = "twse"
        else:
            fresh = fetch_data(inc_codes, gap_days)
            src_tag = "yfinance"

        n_saved = save_kbars_cache(conn, fresh, source=src_tag)
        print(f"  💾 增量補寫入快取 {n_saved} 筆 K bar（source={src_tag}）")
        cached = _merge_into_cache(cached, fresh)

    conn.close()

    # 最後把太舊的截掉，只回傳 [target_start, today] 的部分
    final: dict = {}
    for code, df in cached.items():
        if df is None or len(df) == 0:
            continue
        df2 = df[(df.index.date >= target_start) & (df.index.date <= today)]
        if len(df2) >= 22:
            final[code] = df2

    if final:
        _check_data_freshness(final, source=f"快取+{datasource}")
    return final


# ══════════════════════════════════════════════
# 永豐金 Shioaji 資料下載（kbars 主源 + TWSE 備援）
# ══════════════════════════════════════════════
def fetch_data_sinopac(codes: list, days: int,
                       api_key: str, secret_key: str) -> dict:
    """
    全新架構（yfinance 已完全移除）：
      歷史資料  → 永豐金 kbars（分鐘K聚合日K，與 TWSE 官方完全一致）
      今日 OHLCV → 永豐金 snapshots（即時，已修正 snap.ts 日期）
      備援      → TWSE / TPEX 官方 API（kbars 失敗時自動切換）

    需求：pip install shioaji  ＋ 永豐金 API 金鑰
    """
    if not HAS_SHIOAJI:
        print("  ❌ 請先安裝：pip install shioaji")
        return {}

    # ── 台灣時間 ──────────────────────────────────────────────────
    _tz_tw       = datetime.timezone(datetime.timedelta(hours=8))
    _now_tw      = datetime.datetime.now(_tz_tw)
    _today_tw    = _now_tw.date()
    _open_tw     = _now_tw.replace(hour=9,  minute=0,  second=0, microsecond=0)
    _close_tw    = _now_tw.replace(hour=13, minute=30, second=0, microsecond=0)
    _market_open = (_open_tw <= _now_tw < _close_tw)
    # TWSE 通常 14:00 後才有完整當日資料；14:30 後確認可安全取今日
    # （與 fetch_data_twse 的邊界邏輯一致）
    _twse_close  = _now_tw.replace(hour=14, minute=30, second=0, microsecond=0)
    _twse_ready  = (_now_tw >= _twse_close)   # 盤後且 TWSE 已更新完畢

    # 歷史區間（kbars 不含當日盤中，故 end = 昨日）
    buf        = days + 40
    hist_end   = _today_tw - datetime.timedelta(days=1)
    hist_start = _today_tw - datetime.timedelta(days=buf)
    start_str  = str(hist_start)
    end_str    = str(hist_end)

    # ── Step 1：登入 ──────────────────────────────────────────────
    print(f"  🔑 永豐金登入中...")
    try:
        api = sj.Shioaji()
        api.login(api_key=api_key, secret_key=secret_key, contracts_timeout=30000)
        print(f"  ✅ 永豐金登入成功（api_key: {api_key[:6]}...）")
    except Exception as e:
        print(f"  ❌ 永豐金登入失敗：{e}")
        return {}

    # 輪詢等待合約就緒（與測試腳本相同的做法）
    print(f"  ⏳ 等待合約載入完成...")
    for _wi in range(60):
        try:
            if api.Contracts.Stocks["2330"] is not None:
                break
        except Exception:
            pass
        time.sleep(1)
    # 合約物件就緒後，再等 5 秒讓 SOLACE session 完全穩定，才能接受 kbars 請求
    time.sleep(5)
    print(f"  ✅ 合約載入完成（已等待穩定）")

    def _get_contract(code: str):
        for mkt in ("TSE", "OTC"):
            try:
                c = getattr(api.Contracts.Stocks, mkt)[code]
                if c is not None:
                    return c
            except Exception:
                pass
        try:
            return api.Contracts.Stocks[code]
        except Exception:
            return None

    # ── Step 2：kbars smoke test → 確認此帳號可用 kbars ─────────────
    print(f"  📥 [1/2] 永豐金 kbars 下載歷史日K（{start_str} ~ {end_str}）...")
    if _market_open:
        print(f"  ⏰ 現在 {_now_tw.strftime('%H:%M')} 台灣時間，尚未收盤"
              f"——以昨日（{hist_end}）收盤資料選股")

    results: dict = {}
    fallback_codes: list = []
    _kbars_err_sample: list = []

    # 盤中：kbars Request-Reply 頻道不可用，直接略過 smoke test 走備援
    if _market_open:
        print(f"  ⏩ 盤中（{_now_tw.strftime('%H:%M')}）略過 kbars，直接走 TWSE/TPEX 備援")
        fallback_codes = list(codes)
    else:
        # 先用 2330 試一筆，確認 kbars 可用，否則直接跳備援
        _smoke_ok = False
        _smoke_contract = _get_contract("2330")
        if _smoke_contract:
            try:
                _smoke_kb = api.kbars(
                    contract=_smoke_contract,
                    start=str(hist_end - datetime.timedelta(days=5)),
                    end=str(hist_end),
                )
                if _smoke_kb and len(pd.DataFrame({**_smoke_kb})) > 0:
                    _smoke_ok = True
                    print(f"  ✅ kbars smoke test 通過，開始下載 {len(codes)} 檔...")
            except Exception as e:
                print(f"  ⚠️  kbars smoke test 失敗（{type(e).__name__}: {e}）"
                      f"，全部切換 TWSE 備援")

        if not _smoke_ok:
            fallback_codes = list(codes)

    # shioaji api.kbars() 非 thread-safe，必須循序呼叫
    # smoke test 失敗時 fallback_codes 已含全部 codes，跳過迴圈
    total = len(codes)
    for idx, code in enumerate(codes, 1):
        if code in fallback_codes:       # smoke test 失敗時直接跳過
            continue
        contract = _get_contract(code)
        if contract is None:
            fallback_codes.append(code)
            continue

        kb = None
        last_err = None
        # 外層 retry：SOLACE "Not ready" 時重試（最多 3 次，間隔 1s）
        for _attempt in range(3):
            for freq_kwarg in [{"frequency": "D"}, {}]:
                try:
                    kb = api.kbars(contract=contract,
                                   start=start_str, end=end_str,
                                   **freq_kwarg)
                    break
                except TypeError as te:
                    if "frequency" in str(te):
                        continue
                    last_err = te
                    break
                except Exception as e:
                    last_err = e
                    break
            if kb is not None:
                break
            # Not ready 或暫時性錯誤，等一秒重試
            if last_err and "Not ready" in str(last_err):
                time.sleep(1)
            else:
                break   # 非暫時性錯誤，不重試

        if kb is None:
            fallback_codes.append(code)
            if last_err and len(_kbars_err_sample) < 3:
                _kbars_err_sample.append(f"{code}: {type(last_err).__name__}: {last_err}")
            continue

        try:
            df_min = pd.DataFrame({**kb})
            if len(df_min) == 0:
                fallback_codes.append(code)
                continue

            df_min["ts"]   = pd.to_datetime(df_min["ts"])
            df_min["date"] = df_min["ts"].dt.date

            df_day = df_min.groupby("date").agg(
                Open   = ("Open",   "first"),
                High   = ("High",   "max"),
                Low    = ("Low",    "min"),
                Close  = ("Close",  "last"),
                Vol_K  = ("Volume", "sum"),   # kbars volume 單位：張
            ).reset_index()

            df_day.index = pd.to_datetime(df_day["date"])
            df_day.index.name = None
            df_day = df_day.drop(columns=["date"])
            df_day["Volume"] = df_day["Vol_K"] * 1000   # 張 → 股（相容下游）

            df_day = df_day[(df_day.index.date >= hist_start) &
                            (df_day.index.date <= hist_end)]

            if len(df_day) < 5:
                fallback_codes.append(code)
                continue

            results[code] = df_day

        except Exception as e:
            fallback_codes.append(code)
            if len(_kbars_err_sample) < 3:
                _kbars_err_sample.append(f"{code}(parse): {type(e).__name__}: {e}")

        if idx % 50 == 0 or idx == total:
            print(f"  ── kbars 進度：{idx}/{total} 檔"
                  f"（成功 {len(results)} / 備援 {len(fallback_codes)}）")

    # 若全部走備援，印出錯誤樣本供診斷
    if _kbars_err_sample:
        print(f"  ⚠️  kbars 錯誤樣本（前 {len(_kbars_err_sample)} 筆）：")
        for s in _kbars_err_sample:
            print(f"     {s}")

    print(f"  ✅ kbars 完成：{len(results)} 檔成功  |  {len(fallback_codes)} 檔切換備援")

    # ── Step 3：TWSE / TPEX 備援（(stock×month) 扁平並行） ──────────
    # 把所有 (code, month) 組合攤平成一個大任務池，workers 同時打
    if fallback_codes:
        _fb_months: list = []
        _cur = hist_start.replace(day=1)
        while _cur <= hist_end:
            _fb_months.append(_cur.strftime("%Y%m"))
            _cur = (_cur + datetime.timedelta(days=32)).replace(day=1)

        n_fb     = len(fallback_codes)
        n_months = len(_fb_months)
        n_tasks  = n_fb * n_months
        _tw      = min(20, n_tasks)
        est_sec  = n_tasks * 0.4 / max(_tw, 1)
        print(f"  🔄 TWSE/TPEX 備援：{n_fb} 檔 × {n_months} 個月 = {n_tasks} 筆請求"
              f"（{_tw} workers，估計 {est_sec:.0f} 秒）...")

        from collections import defaultdict
        _code_rows: dict = defaultdict(list)
        _t_done = 0
        _t_lock = threading.Lock()
        _report_every = max(1, n_tasks // 20)

        # thread-local Session：每個 thread 各持一個 TCP 連線池，避免重複 handshake
        _tls = threading.local()

        def _get_session() -> requests.Session:
            if not hasattr(_tls, "session"):
                _tls.session = requests.Session()
                _tls.session.headers.update({"User-Agent": "Mozilla/5.0"})
                _tls.session.verify = False
            return _tls.session

        def _fetch_task(task: tuple) -> None:
            code, ym = task
            # 用 thread-local session 發請求
            sess = _get_session()
            rows = _fetch_official_month(code, ym, session=sess)
            # Lock 只保護計數器和進度輸出，extend 在 lock 外執行
            with _t_lock:
                nonlocal _t_done
                _t_done += 1
                done_now = _t_done
            # extend 不在 lock 內（defaultdict 各 key 獨立，不需鎖）
            if rows:
                _code_rows[code].extend(rows)
            if done_now % _report_every == 0 or done_now == n_tasks:
                pct = done_now * 100 // n_tasks
                print(f"  ── 備援進度：{done_now}/{n_tasks} 筆（{pct}%）")

        # 月份優先排序：(月1,股1),(月1,股2),...  避免對同一支股票集中打
        tasks = [(c, ym) for ym in _fb_months for c in fallback_codes]
        with ThreadPoolExecutor(max_workers=_tw) as pool:
            list(pool.map(_fetch_task, tasks))

        # 盤後且 TWSE 已更新（14:30 後）→ 允許讀取今日資料；盤中仍截至昨日
        _fb_end = _today_tw if _twse_ready else hist_end
        official_ok = 0
        for code in fallback_codes:
            df_off = _rows_to_df(_code_rows.get(code, []), hist_start, _fb_end)
            if df_off is not None and len(df_off) >= 5:
                results[code] = df_off
                official_ok += 1

        print(f"  ✅ TWSE/TPEX 備援：{official_ok}/{n_fb} 檔成功")

    if not results:
        try:
            api.logout()
        except Exception:
            pass
        print("  ❌ 無法取得任何股票資料")
        return {}

    # ── Step 4：snapshots 補上今日收盤（盤後才執行）──────────────
    if _market_open:
        # 盤中：kbars 已截至昨日，訊號以昨日為準，不取今日不完整資料
        print(f"  ⏸  [2/2] 盤中跳過 snapshot，以昨日（{hist_end}）收盤作為選股基準")
    else:
        print(f"  📡 [2/2] 永豐金 snapshots 補上今日 OHLCV...")

        contracts, c2code = [], {}
        for code in list(results.keys()):
            c = _get_contract(code)
            if c is not None:
                contracts.append(c)
                c2code[c.code] = code

        if contracts:
            updated     = 0
            zero_vol    = 0
            stale_codes: list = []

            for bi in range(0, len(contracts), 200):
                batch = contracts[bi:bi+200]
                try:
                    snaps = api.snapshots(batch)
                    for snap in snaps:
                        code = c2code.get(snap.code, snap.code)
                        if code not in results:
                            continue
                        close = float(snap.close)
                        if close <= 0:
                            continue

                        # 從 snap.ts（奈秒）取得資料的實際日期
                        snap_ts = getattr(snap, "ts", None)
                        if snap_ts and snap_ts > 0:
                            snap_date = datetime.datetime.fromtimestamp(
                                snap_ts / 1e9, tz=_tz_tw
                            ).date()
                        else:
                            snap_date = _today_tw

                        if snap_date != _today_tw:
                            stale_codes.append(
                                f"{code}（snap={snap_date}, 今日={_today_tw}）"
                            )

                        # snap.total_volume = 當日累計總量（張）← 正確欄位
                        tv    = float(snap.total_volume) if snap.total_volume else 0.0
                        lv    = float(snap.volume)       if snap.volume       else 0.0
                        vol_k = tv or lv

                        if vol_k <= 0:
                            zero_vol += 1
                            continue

                        today_bar = pd.DataFrame({
                            "Open":   [float(snap.open)],
                            "High":   [float(snap.high)],
                            "Low":    [float(snap.low)],
                            "Close":  [close],
                            "Vol_K":  [vol_k],
                            "Volume": [vol_k * 1000],
                        }, index=[pd.Timestamp(snap_date)])

                        df = results[code]
                        # 同日期的 kbars 舊資料替換成 snapshot 即時值
                        if len(df) > 0 and df.index[-1].date() == snap_date:
                            df = df.iloc[:-1]
                        results[code] = pd.concat([df, today_bar])
                        updated += 1
                except Exception as e:
                    print(f"  ⚠️  snapshots 批次失敗：{e}")

            print(f"  ✅ 永豐金今日收盤更新：{updated}/{len(contracts)} 檔"
                  f"  （零成交量跳過：{zero_vol} 檔）")

            if stale_codes:
                print(f"\n  ⚠️  【資料日期異常】以下 {len(stale_codes)} 檔 snapshot 日期"
                      f"與今日（{_today_tw}）不符，訊號可能基於昨日舊資料：")
                for s in stale_codes[:10]:
                    print(f"     • {s}")
                if len(stale_codes) > 10:
                    print(f"     ...（共 {len(stale_codes)} 檔）")
                print(f"  💡 建議：收盤後 5 分鐘再執行，或等盤後 snapshot 更新完畢")
        else:
            print(f"  ⚠️  找不到可用合約，跳過今日更新")

    try:
        api.logout()
        print(f"  🔓 已登出永豐金 API")
    except Exception:
        pass

    # ── Step 5：過濾不足資料 & 輸出 ───────────────────────────────
    results = {c: df for c, df in results.items() if len(df) >= 22}
    print(f"  📊 最終有效股票（≥22日K）：{len(results)} 檔")

    _check_data_freshness(results, source="永豐金 kbars")
    return results


# ══════════════════════════════════════════════
# 技術指標
# ══════════════════════════════════════════════
def _rsi(series: pd.Series, period: int = RSI_PERIOD) -> pd.Series:
    delta = series.diff()
    gain  = delta.clip(lower=0).rolling(period).mean()
    loss  = (-delta.clip(upper=0)).rolling(period).mean()
    return 100 - 100 / (1 + gain / loss.replace(0, float("nan")))


def _macd(series: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9):
    """回傳 (macd_line, signal_line) 兩個 Series"""
    ema_fast   = series.ewm(span=fast,   adjust=False).mean()
    ema_slow   = series.ewm(span=slow,   adjust=False).mean()
    macd_line  = ema_fast - ema_slow
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    return macd_line, signal_line


def _bollinger(series: pd.Series, period: int = 20, n_std: float = 2.0):
    """回傳 (upper, mid, lower) 三個 Series"""
    mid   = series.rolling(period).mean()
    std   = series.rolling(period).std()
    return mid + n_std * std, mid, mid - n_std * std


def _stochastic(high: pd.Series, low: pd.Series, close: pd.Series,
                period: int = 9, smooth_k: int = 3, smooth_d: int = 3):
    """KD 隨機指標，回傳 (K, D) 兩個 Series"""
    hi  = high.rolling(period).max()
    lo  = low.rolling(period).min()
    rsv = ((close - lo) / (hi - lo).replace(0, float("nan")) * 100).fillna(50)
    k   = rsv.ewm(com=smooth_k - 1, adjust=False).mean()
    d   = k.ewm(com=smooth_d - 1, adjust=False).mean()
    return k, d


def _williams_r(high: pd.Series, low: pd.Series, close: pd.Series,
                period: int = 14) -> pd.Series:
    """Williams %R，值域 -100 ~ 0"""
    hi = high.rolling(period).max()
    lo = low.rolling(period).min()
    return ((hi - close) / (hi - lo).replace(0, float("nan")) * -100).fillna(-50)


def _detect_neckline_patterns(df: pd.DataFrame, i: int) -> tuple:
    """
    在第 i 根 K 棒偵測 W底（多方）和 M頭（空方）頸線型態。
    回傳 (w_bottom: bool, m_top: bool)

    演算法：
    - 往回看 LOOKBACK 根 K 棒
    - 用 WIN 根滑動窗口找局部高低點
    - W底：兩個相近低點（差距 ≤ SIMILARITY），今日收盤突破頸線 且 昨日收盤在頸線下
           （突破首日條件：避免同一型態重複觸發多天）
    - M頭：兩個相近高點（差距 ≤ SIMILARITY），今日收盤跌破頸線 且 昨日收盤在頸線上
    - 頸線需有實質意義：至少比平均低點高 MIN_NECK_PCT%（W底）
    - 兩低/高點最小間距 MIN_SPAN 根（排除噪音）
    - 第二個低/高點需在最近 RECENCY 根 K 棒內（避免型態過舊）
    """
    LOOKBACK     = 60    # 最多往回看 60 根 K 棒
    SIMILARITY   = 0.05  # 兩個低/高點差距 ≤ 5%
    RECENCY      = 20    # 第二個低/高點需在最近 20 根 K 棒內
    WIN          = 3     # 局部高低點滑動窗口（前後各 WIN 根）
    MIN_SPAN     = 5     # 兩低/高點最小間距（根），避免過窄型態
    MIN_NECK_PCT = 0.01  # 頸線需比平均低點高 ≥ 1%（W底）/ 比平均高點低 ≥ 1%（M頭）

    start = max(0, i - LOOKBACK)
    sub   = df.iloc[start : i + 1].reset_index(drop=True)
    n     = len(sub)

    if n < WIN * 2 + 5:
        return False, False

    hi_arr   = sub["High"].values.astype(float)
    lo_arr   = sub["Low"].values.astype(float)
    cl_arr   = sub["Close"].values.astype(float)
    today_cl = cl_arr[-1]
    prev_cl  = cl_arr[-2] if n >= 2 else today_cl
    today_i  = n - 1

    def find_swings(arr: np.ndarray, is_min: bool) -> list:
        """回傳 (index, value) 局部極值清單（排除最後一根，那是今日）"""
        result = []
        for j in range(WIN, today_i - WIN):
            window = arr[j - WIN : j + WIN + 1]
            if is_min and arr[j] == window.min():
                result.append((j, arr[j]))
            elif not is_min and arr[j] == window.max():
                result.append((j, arr[j]))
        return result

    lows  = find_swings(lo_arr, True)
    highs = find_swings(hi_arr, False)

    # ── W底多（T）────────────────────────────────────────────────────
    w_bottom = False
    if len(lows) >= 2:
        for a in range(len(lows) - 1):
            if w_bottom:
                break
            for b in range(a + 1, len(lows)):
                i1, v1 = lows[a]
                i2, v2 = lows[b]
                if today_i - i2 > RECENCY:
                    continue
                if i2 - i1 < MIN_SPAN:                      # ★ 間距過窄排除
                    continue
                if abs(v1 - v2) / max(v1, v2) > SIMILARITY:
                    continue
                # 找兩低點之間的最高點（頸線）
                between_highs = [hv for (hi_i, hv) in highs if i1 < hi_i < i2]
                if between_highs:
                    neckline = max(between_highs)
                elif i2 > i1 + 1:
                    neckline = hi_arr[i1 + 1 : i2].max()
                else:
                    continue
                # ★ 頸線需比平均低點高至少 MIN_NECK_PCT（確保 W 有足夠幅度）
                if neckline < (v1 + v2) / 2 * (1 + MIN_NECK_PCT):
                    continue
                # ★ 突破首日：昨日在頸線下，今日突破（避免連續多天重複觸發）
                if today_cl > neckline and prev_cl <= neckline:
                    w_bottom = True
                    break

    # ── M頭空（TS）───────────────────────────────────────────────────
    m_top = False
    if len(highs) >= 2:
        for a in range(len(highs) - 1):
            if m_top:
                break
            for b in range(a + 1, len(highs)):
                i1, v1 = highs[a]
                i2, v2 = highs[b]
                if today_i - i2 > RECENCY:
                    continue
                if i2 - i1 < MIN_SPAN:
                    continue
                if abs(v1 - v2) / max(v1, v2) > SIMILARITY:
                    continue
                # 找兩高點之間的最低點（頸線）
                between_lows = [lv for (lo_i, lv) in lows if i1 < lo_i < i2]
                if between_lows:
                    neckline = min(between_lows)
                elif i2 > i1 + 1:
                    neckline = lo_arr[i1 + 1 : i2].min()
                else:
                    continue
                if neckline > (v1 + v2) / 2 * (1 - MIN_NECK_PCT):
                    continue
                # ★ 跌破首日
                if today_cl < neckline and prev_cl >= neckline:
                    m_top = True
                    break

    return w_bottom, m_top


def _check(df: pd.DataFrame, i: int,
           min_avg_vol: int, vol_mult: float) -> dict:
    """
    在第 i 日計算所有策略訊號。
    回傳 {A: bool, B: bool, ... ES: bool}
    """
    empty = {k: False for k in STRATEGY_NAMES}
    if i < 20:
        return empty

    win    = df.iloc[:i+1]
    today  = win.iloc[-1]
    prev   = win.iloc[-2] if len(win) >= 2 else today

    close   = float(today["Close"])
    open_p  = float(today["Open"])
    prev_cl = float(prev["Close"])
    vol_k   = float(today["Vol_K"])
    avg5    = float(win["Vol_K"].iloc[-6:-1].mean())
    chg_pct = (close - prev_cl) / prev_cl * 100 if prev_cl else 0

    if avg5 < min_avg_vol:
        return empty

    vol_ratio = vol_k / avg5 if avg5 > 0 else 0
    ma5   = float(win["Close"].iloc[-5:].mean())
    ma10  = float(win["Close"].iloc[-10:].mean())
    ma20  = float(win["Close"].iloc[-20:].mean())
    rsi   = _rsi(win["Close"], 14)
    rsi_t = float(rsi.iloc[-1]) if not pd.isna(rsi.iloc[-1]) else 50
    rsi_p = float(rsi.iloc[-2]) if not pd.isna(rsi.iloc[-2]) else 50
    gap   = (open_p - prev_cl) / prev_cl * 100 if prev_cl else 0
    hi5   = float(win["High"].iloc[-6:-1].max())
    lo5   = float(win["Low"].iloc[-6:-1].min())

    three_up = all(
        float(win["Close"].iloc[-(j+1)]) > float(win["Open"].iloc[-(j+1)])
        for j in range(1, 4)
    ) if len(win) >= 4 else False
    three_dn = all(
        float(win["Close"].iloc[-(j+1)]) < float(win["Open"].iloc[-(j+1)])
        for j in range(1, 4)
    ) if len(win) >= 4 else False

    # ── 新策略計算材料 ──────────────────────────
    # 均量類：近5日均量 vs 近20日均量
    avg20    = float(win["Vol_K"].iloc[-21:-1].mean()) if len(win) >= 22 else avg5
    avg3     = float(win["Vol_K"].iloc[-4:-1].mean())  if len(win) >= 4  else avg5
    # 均量擴張：近5日均量 > 近20日均量 × 1.2
    vol_expand  = avg5 > avg20 * 1.2 if avg20 > 0 else False
    # 均量萎縮：近5日均量 < 近20日均量 × 0.8
    vol_shrink  = avg5 < avg20 * 0.8 if avg20 > 0 else False
    # 潮汐：連續3日縮量（近3日各自 < 前一日量）
    tide_shrink = all(
        float(win["Vol_K"].iloc[-(j+1)]) < float(win["Vol_K"].iloc[-(j+2)])
        for j in range(1, 4)
    ) if len(win) >= 5 else False

    # K棒型態類
    body     = abs(close - open_p)
    upper    = float(today["High"]) - max(close, open_p)
    lower    = min(close, open_p) - float(today["Low"])
    is_bull  = close > open_p   # 今日陽線
    is_bear  = close < open_p   # 今日陰線

    prev_body  = abs(float(prev["Close"]) - float(prev["Open"]))
    prev_bull  = float(prev["Close"]) > float(prev["Open"])
    prev_bear  = float(prev["Close"]) < float(prev["Open"])
    prev_close_p = float(prev["Close"])
    prev_open_p  = float(prev["Open"])

    # 鎚子K：下影線 ≥ 2×實體，上影線 ≤ 實體，前日收跌
    hammer      = (body > 0 and lower >= 2 * body and upper <= body
                   and chg_pct < 0 and prev_bear)
    # 射擊之星：上影線 ≥ 2×實體，下影線 ≤ 實體，前日收漲
    shoot_star  = (body > 0 and upper >= 2 * body and lower <= body
                   and chg_pct > 0 and prev_bull)
    # 吞噬陽線：今日陽線完全包覆前日陰線（今開 < 前收，今收 > 前開）
    engulf_bull = (is_bull and prev_bear
                   and open_p <= prev_close_p and close >= prev_open_p
                   and body > prev_body * 0.8)
    # 吞噬陰線：今日陰線完全包覆前日陽線（今開 > 前收，今收 < 前開）
    engulf_bear = (is_bear and prev_bull
                   and open_p >= prev_close_p and close <= prev_open_p
                   and body > prev_body * 0.8)

    # ── 技術指標類（新增）────────────────────────────────
    # J/JS: MACD 黃金/死亡交叉 (12/26/9)
    _macd_line, _macd_sig = _macd(win["Close"])
    macd_t    = float(_macd_line.iloc[-1])   if not pd.isna(_macd_line.iloc[-1])  else 0.0
    macd_p    = float(_macd_line.iloc[-2])   if len(_macd_line) >= 2 and not pd.isna(_macd_line.iloc[-2]) else 0.0
    msig_t    = float(_macd_sig.iloc[-1])    if not pd.isna(_macd_sig.iloc[-1])   else 0.0
    msig_p    = float(_macd_sig.iloc[-2])    if len(_macd_sig) >= 2 and not pd.isna(_macd_sig.iloc[-2])  else 0.0
    macd_golden = (macd_p < msig_p) and (macd_t > msig_t)
    macd_death  = (macd_p > msig_p) and (macd_t < msig_t)

    # K/KS: 布林通道下軌反彈 / 上軌反壓 (20,2)
    _bb_up, _bb_mid, _bb_lo = _bollinger(win["Close"])
    bb_lower_t  = float(_bb_lo.iloc[-1])  if not pd.isna(_bb_lo.iloc[-1])  else close
    bb_upper_t  = float(_bb_up.iloc[-1])  if not pd.isna(_bb_up.iloc[-1])  else close
    bb_lower_p  = float(_bb_lo.iloc[-2])  if len(_bb_lo) >= 2 and not pd.isna(_bb_lo.iloc[-2]) else bb_lower_t
    bb_upper_p  = float(_bb_up.iloc[-2])  if len(_bb_up) >= 2 and not pd.isna(_bb_up.iloc[-2]) else bb_upper_t
    bb_bounce    = (prev_close_p <= bb_lower_p) and (close > bb_lower_t)
    bb_rejection = (prev_close_p >= bb_upper_p) and (close < bb_upper_t)

    # L/LS: KD 超賣黃金交叉(<30) / 超買死亡交叉(>70)
    _kd_k, _kd_d = _stochastic(win["High"], win["Low"], win["Close"])
    kk_t = float(_kd_k.iloc[-1]) if not pd.isna(_kd_k.iloc[-1]) else 50.0
    kd_t = float(_kd_d.iloc[-1]) if not pd.isna(_kd_d.iloc[-1]) else 50.0
    kk_p = float(_kd_k.iloc[-2]) if len(_kd_k) >= 2 and not pd.isna(_kd_k.iloc[-2]) else 50.0
    kd_p = float(_kd_d.iloc[-2]) if len(_kd_d) >= 2 and not pd.isna(_kd_d.iloc[-2]) else 50.0
    kd_oversold_cross   = (kk_t < 20) and (kk_p < kd_p) and (kk_t > kd_t)  # ★ 30→20（EV 0.09→0.19%，見參數分析 2026-05-05）
    kd_overbought_cross = (kk_t > 70) and (kk_p > kd_p) and (kk_t < kd_t)

    # M/MS: 威廉%R 超賣(<-80)反彈 / 超買(>-20)回落
    _wr  = _williams_r(win["High"], win["Low"], win["Close"])
    wr_t = float(_wr.iloc[-1]) if not pd.isna(_wr.iloc[-1]) else -50.0
    wr_p = float(_wr.iloc[-2]) if len(_wr) >= 2 and not pd.isna(_wr.iloc[-2]) else -50.0
    wr_oversold    = (wr_p < -80) and (wr_t > wr_p) and (chg_pct > 0)
    wr_overbought  = (wr_p > -20) and (wr_t < wr_p) and (chg_pct < 0)

    # N/NS: 均線多頭排列+昨收跌破ma5今收站回 / 空頭排列+昨收突破ma5今收跌破
    ma_bull_align  = (ma5 > ma10 > ma20)
    ma_bear_align  = (ma5 < ma10 < ma20)
    ma_pull_bull   = ma_bull_align and (prev_close_p < ma5) and (close >= ma5)
    ma_pull_bear   = ma_bear_align and (prev_close_p > ma5) and (close <= ma5)

    # ── 進階K棒型態（新增）────────────────────────────────
    # O/OS: 晨星 / 黃昏之星（三K棒）
    if len(win) >= 3:
        day1 = win.iloc[-3]
        day2 = win.iloc[-2]
        d1_c  = float(day1["Close"]); d1_o  = float(day1["Open"])
        d1_body = abs(d1_c - d1_o)
        d2_body = abs(float(day2["Close"]) - float(day2["Open"]))
        d1_mid  = (d1_c + d1_o) / 2
        morning_star = (
            d1_c < d1_o and d1_body > 0 and          # Day1 大陰線
            d2_body < d1_body * 0.4 and               # Day2 小實體（猶豫）
            is_bull and                               # Day3 今日陽線
            close > d1_mid and                        # Day3 收盤超過Day1中點
            body >= d1_body * 0.5                     # Day3 實體夠大
        )
        evening_star = (
            d1_c > d1_o and d1_body > 0 and          # Day1 大陽線
            d2_body < d1_body * 0.4 and               # Day2 小實體
            is_bear and                               # Day3 今日陰線
            close < d1_mid and                        # Day3 收盤低於Day1中點
            body >= d1_body * 0.5                     # Day3 實體夠大
        )
    else:
        morning_star = False
        evening_star = False

    # P/PS: 紅三兵 / 黑三兵（連3根，方向一致＋收盤遞進＋開盤在前根實體內＋實體厚實＋影線短）
    if len(win) >= 3:
        d1 = win.iloc[-3]; d2 = win.iloc[-2]
        d1_c2 = float(d1["Close"]); d1_o2 = float(d1["Open"])
        d2_c2 = float(d2["Close"]); d2_o2 = float(d2["Open"])
        d1_h2 = float(d1["High"]);  d1_l2 = float(d1["Low"])
        d2_h2 = float(d2["High"]);  d2_l2 = float(d2["Low"])
        d1_body2  = d1_c2 - d1_o2                              # 陽線實體（正值）
        d2_body2  = d2_c2 - d2_o2
        d1_rng2   = max(d1_h2 - d1_l2, 1e-9)
        d2_rng2   = max(d2_h2 - d2_l2, 1e-9)
        today_rng = max(float(today["High"]) - float(today["Low"]), 1e-9)
        three_soldiers = (
            # ① 三根都是陽線
            d1_c2 > d1_o2 and d2_c2 > d2_o2 and is_bull and
            # ③ 收盤逐步墊高
            d2_c2 > d1_c2 and close > d2_c2 and
            # ② 開盤梯步遞升（每根開盤不低於前根開盤，允許跳空）
            d2_o2 >= d1_o2 and open_p >= d2_o2 and
            # ④ 每根實體 ≥ 50% 整根K棒高低範圍（排除長影假陽線）
            d1_body2 / d1_rng2 >= 0.5 and
            d2_body2 / d2_rng2 >= 0.5 and
            body / today_rng >= 0.5 and
            # ⑤ 上影線 < 實體 × 50%（排除上影過長的弱勢陽線）
            (d1_h2 - d1_c2) < d1_body2 * 0.5 and
            (d2_h2 - d2_c2) < d2_body2 * 0.5 and
            upper < body * 0.5
        )
        d1_bbody2 = d1_o2 - d1_c2                              # 黑三兵陰線實體（正值）
        d2_bbody2 = d2_o2 - d2_c2
        three_crows = (
            # ① 三根都是陰線
            d1_c2 < d1_o2 and d2_c2 < d2_o2 and is_bear and
            # ③ 收盤逐步下沉
            d2_c2 < d1_c2 and close < d2_c2 and
            # ② 開盤梯步遞降（每根開盤不高於前根開盤）
            d2_o2 <= d1_o2 and open_p <= d2_o2 and
            # ④ 實體 ≥ 50%
            d1_bbody2 / d1_rng2 >= 0.5 and
            d2_bbody2 / d2_rng2 >= 0.5 and
            body / today_rng >= 0.5 and
            # ⑤ 下影線 < 實體 × 50%
            (d1_c2 - d1_l2) < d1_bbody2 * 0.5 and
            (d2_c2 - d2_l2) < d2_bbody2 * 0.5 and
            lower < body * 0.5
        )
    else:
        three_soldiers = False
        three_crows    = False

    # Q/QS: Inside Bar 突破/跌破（前日為 Inside Bar，今日突破/跌破前前日高低）
    if len(win) >= 3:
        prev2      = win.iloc[-3]
        prev2_h    = float(prev2["High"]); prev2_l = float(prev2["Low"])
        prev_h_q   = float(prev["High"]);  prev_l_q = float(prev["Low"])
        is_inside  = (prev_h_q < prev2_h) and (prev_l_q > prev2_l)
        inside_breakout  = is_inside and (close > prev2_h)                          # Q多：移除爆量
        inside_breakdown = is_inside and (close < prev2_l) and (vol_ratio >= vol_mult)  # QS空：保留爆量
    else:
        inside_breakout  = False
        inside_breakdown = False

    # ── 均值回歸（新增）──────────────────────────────────
    # R/RS: BIAS 乖離率（對 MA20）過大後反向
    bias          = (close - ma20) / ma20 * 100 if ma20 > 0 else 0
    bias_oversold   = (bias < -10) and (chg_pct > 0)   # ★ -8→-10（EV +0.26%，樣本仍充足，見參數分析 2026-05-05）
    bias_overbought = (bias > +8) and (chg_pct < 0)

    # ── 頸線型態（非向量化，逐根計算）───────────────────
    w_neckline, m_neckline = _detect_neckline_patterns(df, i)

    return {
        # 原有十個
        # ★ 多方已移除爆量條件（量多反而降低多方 EV，見量比分析 2026-04-25）
        "A":  bool(ma5 > ma20 and chg_pct > 0),                     # 移除 vol_ratio
        "AS": bool(ma5 < ma20 and vol_ratio >= vol_mult and chg_pct < 0),
        "B":  bool(gap >= 2.0 and chg_pct > 0),  # ★ vol_ratio 已移除（量大反降 EV，見參數分析 2026-05-05）
        "BS": bool(gap <= -2.0 and vol_ratio >= 1.3 and chg_pct < 0),
        "C":  bool(rsi_p < 35 and rsi_t > rsi_p and close > ma5),
        "CS": bool(rsi_p > 65 and rsi_t < rsi_p and close < ma5),
        "D":  bool(close > hi5 and ma5 > ma20),                      # 需在均線多頭才有效突破
        "DS": bool(close < lo5 and vol_ratio >= vol_mult),
        "E":  bool(three_up and ma5 > ma10 > ma20 and chg_pct > 0),
        "ES": bool(three_dn and ma5 < ma10 < ma20 and chg_pct < 0),
        # 均量類
        "F":  bool(vol_expand and chg_pct > 0),
        "FS": bool(vol_shrink and vol_ratio >= vol_mult and chg_pct < 0),
        "G":  bool(tide_shrink and ma5 > ma20 and chg_pct > 0),      # 縮量反彈需在多頭趨勢
        "GS": bool(tide_shrink and ma5 < ma20 and vol_ratio >= vol_mult and chg_pct < 0),  # 同理空頭
        # K棒型態
        "H":  bool(hammer    and close < ma20),  # 鎚子需在 MA20 以下（超賣情境才有反轉意義）
        "HS": bool(shoot_star and close > ma20),  # 射擊之星需在 MA20 以上（超買情境）
        "I":  bool(engulf_bull),
        "IS": bool(engulf_bear),
        # 技術指標類（新增）
        "J":  bool(macd_golden and ma5 > ma20 and chg_pct > 0),      # MACD 金叉需順趨勢
        "JS": bool(macd_death  and ma5 < ma20 and chg_pct < 0),      # MACD 死叉需順趨勢
        "K":  bool(bb_bounce),
        "KS": bool(bb_rejection),
        "L":  bool(kd_oversold_cross),
        "LS": bool(kd_overbought_cross),
        "M":  bool(wr_oversold),
        "MS": bool(wr_overbought),
        "N":  bool(ma_pull_bull),
        "NS": bool(ma_pull_bear),
        # 進階K棒型態（新增）
        "O":  bool(morning_star  and close < ma20),  # 晨星需在 MA20 以下
        "OS": bool(evening_star  and close > ma20),  # 黃昏星需在 MA20 以上
        "P":  bool(three_soldiers and ma5 > ma20),   # 紅三兵需在多頭趨勢
        "PS": bool(three_crows    and ma5 < ma20),   # 黑三兵需在空頭趨勢
        "Q":  bool(inside_breakout  and ma5 > ma20), # IB 突破需在多頭趨勢
        "QS": bool(inside_breakdown and ma5 < ma20), # IB 跌破需在空頭趨勢
        # 均值回歸（新增）
        "R":  bool(bias_oversold),
        "RS": bool(bias_overbought),
        # 大跳空低量（新增）
        "B2": bool(gap >= 5.0 and vol_ratio >= 0.8 and chg_pct > 0),
        # 頸線型態（新增，非向量化）
        "T":  bool(w_neckline),
        "TS": bool(m_neckline),
    }


# ══════════════════════════════════════════════
# 向量化訊號預計算（加速回測核心）
# ══════════════════════════════════════════════
def _precompute_signals_vec(sub: pd.DataFrame,
                             vol_mult: float, min_avg_vol: int) -> np.ndarray:
    """
    一次性向量化計算所有 18 個策略在每一根 K 棒的訊號。
    等同於對每個 i 呼叫 _check(sub, i, ...)，但速度快 100~1000 倍。

    回傳 numpy bool 陣列，shape = (n_rows, 18)，
    欄位順序與 _STRAT_KEYS / STRATEGY_NAMES 一致。
    """
    c = sub["Close"].astype(float).values
    o = sub["Open"].astype(float).values
    h = sub["High"].astype(float).values
    l = sub["Low"].astype(float).values
    v = sub["Vol_K"].astype(float).values
    n = len(sub)

    # pandas Series（用於 rolling）
    c_s = pd.Series(c)
    h_s = pd.Series(h)
    l_s = pd.Series(l)
    v_s = pd.Series(v)

    # ── Moving averages ──────────────────────────────────────────
    ma5_v  = c_s.rolling(5).mean().values
    ma10_v = c_s.rolling(10).mean().values
    ma20_v = c_s.rolling(20).mean().values

    # ── Volume averages（前 N 日均量，不含今日）────────────────
    avg5_v  = v_s.shift(1).rolling(5).mean().values
    avg20_v = v_s.shift(1).rolling(20).mean().values

    # ── Base filter：需 ≥20 根歷史 + 5日均量達標 ───────────────
    base_ok = np.zeros(n, dtype=bool)
    if n > 20:
        base_ok[20:] = True
    with np.errstate(invalid="ignore"):
        vol_ok = base_ok & (avg5_v >= min_avg_vol)

    # ── Vol ratio ─────────────────────────────────────────────────
    with np.errstate(divide="ignore", invalid="ignore"):
        vol_ratio = np.where(avg5_v > 0, v / avg5_v, np.nan)

    # ── Price change & gap ───────────────────────────────────────
    prev_close = np.empty(n); prev_close[0] = np.nan; prev_close[1:] = c[:-1]
    with np.errstate(divide="ignore", invalid="ignore"):
        chg_pct = (c - prev_close) / prev_close * 100
        gap     = (o - prev_close) / prev_close * 100

    # ── RSI (14-period) ──────────────────────────────────────────
    rsi_s    = _rsi(c_s, 14)
    rsi_v    = rsi_s.values
    rsi_prev = np.empty(n); rsi_prev[0] = np.nan; rsi_prev[1:] = rsi_v[:-1]

    # ── 5-day high/low（前 5 日，不含今日）─────────────────────
    hi5_v = h_s.shift(1).rolling(5).max().values
    lo5_v = l_s.shift(1).rolling(5).min().values

    # ── Three consecutive candles（前 3 日）─────────────────────
    bull_c = (c > o)
    bear_c = (c < o)
    # shift(k) → 前 k 根
    def _shift(arr, k):
        out = np.empty(n, dtype=bool); out[:k] = False; out[k:] = arr[:-k]; return out
    three_up = _shift(bull_c, 1) & _shift(bull_c, 2) & _shift(bull_c, 3)
    three_dn = _shift(bear_c, 1) & _shift(bear_c, 2) & _shift(bear_c, 3)

    # ── Volume expansion / contraction ───────────────────────────
    with np.errstate(invalid="ignore"):
        vol_expand = (avg5_v > avg20_v * 1.2)
        vol_shrink = (avg5_v < avg20_v * 0.8)

    # ── Tide shrink（連續 3 日縮量）────────────────────────────
    def _shiftf(arr, k):
        out = np.empty(n); out[:k] = np.nan; out[k:] = arr[:-k]; return out
    tide_shrink = (
        (_shiftf(v, 1) < _shiftf(v, 2)) &
        (_shiftf(v, 2) < _shiftf(v, 3)) &
        (_shiftf(v, 3) < _shiftf(v, 4))
    )

    # ── K-bar patterns ───────────────────────────────────────────
    body      = np.abs(c - o)
    upper     = h - np.maximum(c, o)
    lower     = np.minimum(c, o) - l
    prev_body = _shiftf(body, 1)
    prev_bull = _shift(bull_c, 1)
    prev_bear = _shift(bear_c, 1)
    prev_c    = _shiftf(c, 1)
    prev_o    = _shiftf(o, 1)

    with np.errstate(invalid="ignore"):
        hammer = (
            (body > 0) & (lower >= 2 * body) & (upper <= body) &
            (chg_pct < 0) & prev_bear
        )
        shoot_star = (
            (body > 0) & (upper >= 2 * body) & (lower <= body) &
            (chg_pct > 0) & prev_bull
        )
        engulf_bull = (
            bull_c & prev_bear &
            (o <= prev_c) & (c >= prev_o) &
            (body > prev_body * 0.8)
        )
        engulf_bear = (
            bear_c & prev_bull &
            (o >= prev_c) & (c <= prev_o) &
            (body > prev_body * 0.8)
        )

    # ── 技術指標類（新增）─────────────────────────────────
    # J/JS: MACD 黃金/死亡交叉 (12/26/9)
    ema12_v    = c_s.ewm(span=12, adjust=False).mean().values
    ema26_v    = c_s.ewm(span=26, adjust=False).mean().values
    macd_v     = ema12_v - ema26_v
    macd_sig_v = pd.Series(macd_v).ewm(span=9, adjust=False).mean().values
    macd_prev  = _shiftf(macd_v,     1)
    msig_prev  = _shiftf(macd_sig_v, 1)
    with np.errstate(invalid="ignore"):
        macd_golden_v = (macd_prev < msig_prev) & (macd_v > macd_sig_v)
        macd_death_v  = (macd_prev > msig_prev) & (macd_v < macd_sig_v)

    # K/KS: 布林通道下軌反彈 / 上軌反壓 (20, 2)
    bb_std_v    = c_s.rolling(20).std().values
    bb_upper_v  = ma20_v + 2 * bb_std_v
    bb_lower_v  = ma20_v - 2 * bb_std_v
    prev_bb_lo  = _shiftf(bb_lower_v, 1)
    prev_bb_up  = _shiftf(bb_upper_v, 1)
    with np.errstate(invalid="ignore"):
        bb_bounce_v    = (prev_c <= prev_bb_lo) & (c > bb_lower_v)
        bb_rejection_v = (prev_c >= prev_bb_up) & (c < bb_upper_v)

    # L/LS: KD 超賣黃金交叉(<30) / 超買死亡交叉(>70)
    hi9_v  = h_s.rolling(9).max().values
    lo9_v  = l_s.rolling(9).min().values
    with np.errstate(divide="ignore", invalid="ignore"):
        rsv_v  = np.where((hi9_v - lo9_v) > 0,
                          (c - lo9_v) / (hi9_v - lo9_v) * 100, 50.0)
    kd_k_v  = pd.Series(rsv_v).ewm(com=2, adjust=False).mean().values
    kd_d_v  = pd.Series(kd_k_v).ewm(com=2, adjust=False).mean().values
    kd_kp   = _shiftf(kd_k_v, 1)
    kd_dp   = _shiftf(kd_d_v, 1)
    with np.errstate(invalid="ignore"):
        kd_oversold_v   = (kd_k_v < 20) & (kd_kp < kd_dp) & (kd_k_v > kd_d_v)   # ★ 30→20（見參數分析 2026-05-05）
        kd_overbought_v = (kd_k_v > 70) & (kd_kp > kd_dp) & (kd_k_v < kd_d_v)

    # M/MS: 威廉%R 超賣(<-80)反彈 / 超買(>-20)回落
    hi14_v = h_s.rolling(14).max().values
    lo14_v = l_s.rolling(14).min().values
    with np.errstate(divide="ignore", invalid="ignore"):
        wr_v = np.where((hi14_v - lo14_v) > 0,
                        (hi14_v - c) / (hi14_v - lo14_v) * -100, -50.0)
    wr_prev_v = _shiftf(wr_v, 1)
    with np.errstate(invalid="ignore"):
        wr_oversold_v   = (wr_prev_v < -80) & (wr_v > wr_prev_v) & (chg_pct > 0)
        wr_overbought_v = (wr_prev_v > -20) & (wr_v < wr_prev_v) & (chg_pct < 0)

    # N/NS: 均線多頭排列+昨收跌破ma5今收站回 / 空頭排列+昨收突破ma5今收跌破
    ma_bull_v    = (ma5_v > ma10_v) & (ma10_v > ma20_v)
    ma_bear_v    = (ma5_v < ma10_v) & (ma10_v < ma20_v)
    with np.errstate(invalid="ignore"):
        ma_pull_bull_v = ma_bull_v & (prev_c < ma5_v) & (c >= ma5_v)
        ma_pull_bear_v = ma_bear_v & (prev_c > ma5_v) & (c <= ma5_v)

    # ── 進階K棒型態（新增）─────────────────────────────────
    # O/OS: 晨星 / 黃昏之星（三K棒，需至少3根）
    d1_c_v    = _shiftf(c, 2);   d1_o_v    = _shiftf(o, 2)
    d1_body_v = np.abs(d1_c_v - d1_o_v)
    d2_body_v = _shiftf(body, 1)
    d1_mid_v  = (d1_c_v + d1_o_v) / 2
    with np.errstate(invalid="ignore"):
        morning_star_v = (
            (d1_c_v < d1_o_v) &                   # Day1 陰線
            (d1_body_v > 0) &                      # Day1 有實體
            (d2_body_v < d1_body_v * 0.4) &        # Day2 小實體
            bull_c &                               # Day3 陽線
            (c > d1_mid_v) &                       # Day3 收超過Day1中點
            (body >= d1_body_v * 0.5)              # Day3 實體夠大
        )
        evening_star_v = (
            (d1_c_v > d1_o_v) &                   # Day1 陽線
            (d1_body_v > 0) &
            (d2_body_v < d1_body_v * 0.4) &
            bear_c &                               # Day3 陰線
            (c < d1_mid_v) &                       # Day3 收低於Day1中點
            (body >= d1_body_v * 0.5)
        )

    # P/PS: 紅三兵 / 黑三兵（完整5條件：陽線×3＋收盤遞進＋開在前體內＋實體厚＋影線短）
    d1_bull_v  = _shift(bull_c, 2); d2_bull_v = _shift(bull_c, 1)
    d1_bear_v  = _shift(bear_c, 2); d2_bear_v = _shift(bear_c, 1)
    d1_c2_v    = _shiftf(c, 2);     d2_c2_v   = _shiftf(c, 1)
    d1_o2_v    = _shiftf(o, 2);     d2_o2_v   = _shiftf(o, 1)
    d1_h2_v    = _shiftf(h, 2);     d2_h2_v   = _shiftf(h, 1)
    d1_l2_v    = _shiftf(l, 2);     d2_l2_v   = _shiftf(l, 1)
    # 各根實體與全範圍
    d1_bull_body_v = d1_c2_v - d1_o2_v          # 陽線實體（正值）
    d2_bull_body_v = d2_c2_v - d2_o2_v
    d1_rng2_v      = np.maximum(d1_h2_v - d1_l2_v, 1e-9)
    d2_rng2_v      = np.maximum(d2_h2_v - d2_l2_v, 1e-9)
    today_rng_v    = np.maximum(h - l, 1e-9)
    body_abs_v     = np.abs(c - o)
    upper_v        = h - np.maximum(c, o)
    lower_v        = np.minimum(c, o) - l
    with np.errstate(invalid="ignore"):
        three_soldiers_v = (
            d1_bull_v & d2_bull_v & bull_c &
            # ③ 收盤遞進
            (d2_c2_v > d1_c2_v) & (c > d2_c2_v) &
            # ② 開盤梯步遞升（每根開盤不低於前根開盤，允許跳空）
            (d2_o2_v >= d1_o2_v) & (o >= d2_o2_v) &
            # ④ 實體 ≥ 50% 範圍
            (d1_bull_body_v / d1_rng2_v >= 0.5) &
            (d2_bull_body_v / d2_rng2_v >= 0.5) &
            (body_abs_v / today_rng_v >= 0.5) &
            # ⑤ 上影線 < 實體 × 50%
            ((d1_h2_v - d1_c2_v) < d1_bull_body_v * 0.5) &
            ((d2_h2_v - d2_c2_v) < d2_bull_body_v * 0.5) &
            (upper_v < body_abs_v * 0.5)
        )
        d1_bear_body_v = d1_o2_v - d1_c2_v      # 陰線實體（正值）
        d2_bear_body_v = d2_o2_v - d2_c2_v
        three_crows_v = (
            d1_bear_v & d2_bear_v & bear_c &
            # ③ 收盤遞跌
            (d2_c2_v < d1_c2_v) & (c < d2_c2_v) &
            # ② 開盤梯步遞降（每根開盤不高於前根開盤）
            (d2_o2_v <= d1_o2_v) & (o <= d2_o2_v) &
            # ④ 實體 ≥ 50%
            (d1_bear_body_v / d1_rng2_v >= 0.5) &
            (d2_bear_body_v / d2_rng2_v >= 0.5) &
            (body_abs_v / today_rng_v >= 0.5) &
            # ⑤ 下影線 < 實體 × 50%
            ((d1_c2_v - d1_l2_v) < d1_bear_body_v * 0.5) &
            ((d2_c2_v - d2_l2_v) < d2_bear_body_v * 0.5) &
            (lower_v < body_abs_v * 0.5)
        )

    # Q/QS: Inside Bar 突破/跌破
    prev2_h_v  = _shiftf(h, 2); prev2_l_v = _shiftf(l, 2)
    prev1_h_v  = _shiftf(h, 1); prev1_l_v = _shiftf(l, 1)
    is_inside_v = (prev1_h_v < prev2_h_v) & (prev1_l_v > prev2_l_v)
    with np.errstate(invalid="ignore"):
        inside_breakout_v  = is_inside_v & (c > prev2_h_v)                          # Q多：移除爆量
        inside_breakdown_v = is_inside_v & (c < prev2_l_v) & (vol_ratio >= vol_mult)  # QS空：保留爆量

    # ── 均值回歸（新增）────────────────────────────────────
    # R/RS: BIAS 乖離率（對 MA20）
    with np.errstate(divide="ignore", invalid="ignore"):
        bias_v = np.where(ma20_v > 0, (c - ma20_v) / ma20_v * 100, 0.0)
    bias_oversold_v   = (bias_v < -10) & (chg_pct > 0)   # ★ -8→-10（見參數分析 2026-05-05）
    bias_overbought_v = (bias_v > +8) & (chg_pct < 0)

    # ── 大跳空低量（B2）────────────────────────────────────────
    # B2多：開盤跳空 ≥ 5%（強力缺口）+ 成交量 ≥ 0.8x 均量（不要求爆量）
    # 設計初衷：補抓後節假日消息面跳空（如 2489 04/01，VolR=0.85x 符合但 B多需 1.3x）
    b2_long_v = (gap >= 5.0) & (vol_ratio >= 0.8) & (chg_pct > 0)

    # ── 組合成訊號陣列（shape: n × 37，欄位順序 = _STRAT_KEYS）──
    def _f(cond):
        """NaN-safe bool：NaN → False"""
        return np.where(np.isnan(cond.astype(float)), False, cond).astype(bool) & vol_ok

    with np.errstate(invalid="ignore"):
        sig = np.column_stack([
            # ★ 多方（A/B/D/F/G/Q）已移除 vol_ratio 爆量條件（2026-04-25）
            # 量多反而降低多方 EV；空方保留爆量條件（量大對空方有正向確認）
            _f((ma5_v > ma20_v)                               & (chg_pct > 0)),   # A  移除 vol_mult
            _f((ma5_v < ma20_v)   & (vol_ratio >= vol_mult)  & (chg_pct < 0)),   # AS 保留
            _f((gap >= 2.0)                                  & (chg_pct > 0)),   # B  移除 vol_ratio（量大反降 EV）
            _f((gap <= -2.0)       & (vol_ratio >= 1.3)       & (chg_pct < 0)),   # BS 保留
            _f((rsi_prev < 35)     & (rsi_v > rsi_prev)       & (c > ma5_v)),     # C
            _f((rsi_prev > 65)     & (rsi_v < rsi_prev)       & (c < ma5_v)),     # CS
            _f((c > hi5_v)         & (ma5_v > ma20_v)),                            # D  需多頭趨勢
            _f((c < lo5_v)         & (vol_ratio >= vol_mult)),                     # DS
            _f(three_up            & (ma5_v > ma10_v) & (ma10_v > ma20_v) & (chg_pct > 0)),  # E
            _f(three_dn            & (ma5_v < ma10_v) & (ma10_v < ma20_v) & (chg_pct < 0)),  # ES
            _f(vol_expand                                     & (chg_pct > 0)),   # F
            _f(vol_shrink          & (vol_ratio >= vol_mult)  & (chg_pct < 0)),   # FS
            _f(tide_shrink         & (ma5_v > ma20_v)         & (chg_pct > 0)),   # G  需多頭趨勢
            _f(tide_shrink         & (ma5_v < ma20_v) & (vol_ratio >= vol_mult) & (chg_pct < 0)),  # GS 需空頭趨勢
            _f(hammer              & (c < ma20_v)),                                # H  需在 MA20 以下
            _f(shoot_star          & (c > ma20_v)),                                # HS 需在 MA20 以上
            _f(engulf_bull),                                                       # I
            _f(engulf_bear),                                                       # IS
            _f(macd_golden_v       & (ma5_v > ma20_v) & (chg_pct > 0)),           # J  需多頭趨勢
            _f(macd_death_v        & (ma5_v < ma20_v) & (chg_pct < 0)),           # JS 需空頭趨勢
            _f(bb_bounce_v),                                                       # K
            _f(bb_rejection_v),                                                    # KS
            _f(kd_oversold_v),                                                     # L
            _f(kd_overbought_v),                                                   # LS
            _f(wr_oversold_v),                                                     # M
            _f(wr_overbought_v),                                                   # MS
            _f(ma_pull_bull_v),                                                    # N
            _f(ma_pull_bear_v),                                                    # NS
            _f(morning_star_v      & (c < ma20_v)),                                # O  需在 MA20 以下
            _f(evening_star_v      & (c > ma20_v)),                                # OS 需在 MA20 以上
            _f(three_soldiers_v    & (ma5_v > ma20_v)),                            # P  需多頭趨勢
            _f(three_crows_v       & (ma5_v < ma20_v)),                            # PS 需空頭趨勢
            _f(inside_breakout_v   & (ma5_v > ma20_v)),                            # Q  需多頭趨勢
            _f(inside_breakdown_v  & (ma5_v < ma20_v)),                            # QS 需空頭趨勢
            _f(bias_oversold_v),                                                   # R
            _f(bias_overbought_v),                                                 # RS
            _f(b2_long_v),                                                         # B2
            np.zeros(n, dtype=bool),                                               # T  (頸線W底，非向量化，回測階段保留佔位)
            np.zeros(n, dtype=bool),                                               # TS (頸線M頭，非向量化，回測階段保留佔位)
        ])  # shape (n, 39), dtype bool

    return sig


# ══════════════════════════════════════════════
# 選股
# ══════════════════════════════════════════════
def run_scan(data: dict, min_hit: int, vol_mult: float,
             min_avg_vol: int, strategy_filter: list = None,
             show_combos: list = None, workers: int = 4) -> list:
    """
    對每檔股票計算今日訊號，回傳通過命中門檻的候選清單。
    使用 ThreadPoolExecutor 平行處理各股票以加速掃描。

    show_combos : 若指定（如 ['C+F+I','B+E+F']），只保留至少命中其中一個組合的股票，
                  並在結果中附上「命中組合」欄位。
    """
    # 預先解析每個組合的策略集合
    combo_specs = None
    if show_combos:
        combo_specs = [(c, c.split("+")) for c in show_combos]

    def _process_stock(item):
        code, df = item
        if len(df) < 22:
            return None

        sigs = _check(df, len(df) - 1, min_avg_vol, vol_mult)

        # ── MA20 方向判斷 ────────────────────────────────────────
        # 比較今日 MA20 與 10 個交易日前的 MA20
        # 需要至少 30 根 K 棒（MA20 + 10 天位移）
        _SIDEWAYS_THRESH = 0.01   # 1% → 橫盤門檻
        if len(df) >= 31:
            ma20_now = float(df["Close"].iloc[-20:].mean())
            ma20_10d = float(df["Close"].iloc[-30:-10].mean())
            ma20_slope = (ma20_now - ma20_10d) / ma20_10d if ma20_10d > 0 else 0
            if abs(ma20_slope) < _SIDEWAYS_THRESH:
                ma20_dir = "橫"
            elif ma20_slope > 0:
                ma20_dir = "升"
            else:
                ma20_dir = "降"
        else:
            ma20_dir   = "—"
            ma20_slope = 0.0

        # ── MA20 訊號過濾 ────────────────────────────────────────
        # 升：只用 C/K/L/M 多；CS/KS/LS/MS/R/RS 全部關閉
        # 降：只用 CS/KS/LS/MS 空；C/K/L/M/R/RS 全部關閉
        # 橫：R/RS 啟用（BIAS 策略）；C/K/L/M 兩方向均可保留
        # —（資料不足）：不過濾
        if ma20_dir == "升":
            for k in ("CS", "KS", "LS", "MS", "R", "RS",
                      "JS", "PS", "QS"):       # J/P/Q 空方版本關閉
                sigs[k] = False
        elif ma20_dir == "降":
            for k in ("C", "K", "L", "M", "R", "RS",
                      "J", "P", "Q"):          # J/P/Q 多方版本關閉
                sigs[k] = False
        elif ma20_dir == "橫":
            for k in ("C", "CS", "K", "KS", "L", "LS", "M", "MS", "R", "RS",
                      "J", "JS", "P", "PS", "Q", "QS"):  # 橫盤全關
                sigs[k] = False

        # MA20方向顯示標籤
        slope_pct = ma20_slope * 100
        if ma20_dir == "升":
            ma20_label = f"升↑{slope_pct:+.1f}%"
        elif ma20_dir == "降":
            ma20_label = f"降↓{slope_pct:+.1f}%"
        elif ma20_dir == "橫":
            ma20_label = f"橫─{slope_pct:+.1f}%"
        else:
            ma20_label = "—"

        # ── 組合過濾 ─────────────────────────────────────────────
        matched_combos = []
        if combo_specs is not None:
            for combo_key, strats in combo_specs:
                if all(sigs.get(s, False) for s in strats):
                    matched_combos.append(combo_key)
            if not matched_combos:
                return None

        # ── 命中數計算 ───────────────────────────────────────────
        if strategy_filter:
            long_hit  = sum(1 for k, v in sigs.items() if v and k in strategy_filter and not k.endswith("S"))
            short_hit = sum(1 for k, v in sigs.items() if v and k in strategy_filter and k.endswith("S"))
        else:
            long_hit  = sum(1 for k, v in sigs.items() if v and not k.endswith("S"))
            short_hit = sum(1 for k, v in sigs.items() if v and k.endswith("S"))
        hit = max(long_hit, short_hit)

        if combo_specs is None and hit < min_hit:
            return None

        direction = "多" if long_hit >= short_hit else "空"
        today     = df.iloc[-1]
        prev      = df.iloc[-2] if len(df) >= 2 else today
        close     = round(float(today["Close"]), 2)
        prev_cl   = float(prev["Close"])
        chg       = round((close - prev_cl) / prev_cl * 100, 2) if prev_cl else 0
        vol_k     = int(round(float(today["Vol_K"])))
        avg5      = float(df["Vol_K"].iloc[-6:-1].mean())
        vol_r     = f"{vol_k/avg5:.1f}x" if avg5 > 0 else "—"
        strats    = ",".join(k for k, v in sigs.items() if v)
        # 策略型態分類（依方向取對應訊號集合 → 均值回歸 / 趨勢跟風 / 混合 / 缺口）
        if direction == "多":
            typed_keys = [k for k, v in sigs.items() if v and not k.endswith("S")]
        else:
            typed_keys = [k for k, v in sigs.items()
                          if v and (k.endswith("S") or k in _GAP_SIGS)]
        style_tag = _classify_style(typed_keys)
        try:
            data_date = str(df.index[-1].date())
        except Exception:
            data_date = "—"

        return {
            "代號":      code,
            "名稱":      get_name(code),
            "資料日期":  data_date,
            "收盤":      close,
            "漲跌幅(%)": chg,
            "成交量(張)": vol_k,
            "量/均量":   vol_r,
            "A多":  "✅" if sigs["A"]  else "❌",
            "AS空": "✅" if sigs["AS"] else "❌",
            "B多":  "✅" if sigs["B"]  else "❌",
            "BS空": "✅" if sigs["BS"] else "❌",
            "C多":  "✅" if sigs["C"]  else "❌",
            "CS空": "✅" if sigs["CS"] else "❌",
            "D多":  "✅" if sigs["D"]  else "❌",
            "DS空": "✅" if sigs["DS"] else "❌",
            "E多":  "✅" if sigs["E"]  else "❌",
            "ES空": "✅" if sigs["ES"] else "❌",
            "F均量多":  "✅" if sigs["F"]  else "❌",
            "FS均量空": "✅" if sigs["FS"] else "❌",
            "G潮汐多":  "✅" if sigs["G"]  else "❌",
            "GS潮汐空": "✅" if sigs["GS"] else "❌",
            "H鎚子":    "✅" if sigs["H"]  else "❌",
            "HS射擊":   "✅" if sigs["HS"] else "❌",
            "I吞多":    "✅" if sigs["I"]  else "❌",
            "IS吞空":   "✅" if sigs["IS"] else "❌",
            "J MACD多": "✅" if sigs["J"]  else "❌",
            "JS MACD空":"✅" if sigs["JS"] else "❌",
            "K布林多":  "✅" if sigs["K"]  else "❌",
            "KS布林空": "✅" if sigs["KS"] else "❌",
            "L KD多":   "✅" if sigs["L"]  else "❌",
            "LS KD空":  "✅" if sigs["LS"] else "❌",
            "M威廉多":  "✅" if sigs["M"]  else "❌",
            "MS威廉空": "✅" if sigs["MS"] else "❌",
            "N均線多":  "✅" if sigs["N"]  else "❌",
            "NS均線空": "✅" if sigs["NS"] else "❌",
            "O晨星":    "✅" if sigs["O"]  else "❌",
            "OS昏星":   "✅" if sigs["OS"] else "❌",
            "P三兵":    "✅" if sigs["P"]  else "❌",
            "PS三鴉":   "✅" if sigs["PS"] else "❌",
            "Q破IB多":  "✅" if sigs["Q"]  else "❌",
            "QS破IB空": "✅" if sigs["QS"] else "❌",
            "R偏低多":  "✅" if sigs["R"]  else "❌",
            "RS偏高空": "✅" if sigs["RS"] else "❌",
            "B2大跳空": "✅" if sigs["B2"] else "❌",
            "T頸線W底": "✅" if sigs["T"]  else "❌",
            "TS頸線M頭":"✅" if sigs["TS"] else "❌",
            "命中數":    hit,
            "方向":      direction,
            "MA20方向":  ma20_label,
            "策略型態":  style_tag,
            "策略清單":  strats,
            "命中組合":  " | ".join(matched_combos) if matched_combos else "—",
        }

    with ThreadPoolExecutor(max_workers=workers) as pool:
        results = [r for r in pool.map(_process_stock, data.items()) if r is not None]

    results.sort(key=lambda x: (-x["命中數"], -x["漲跌幅(%)"]))
    return results


# ══════════════════════════════════════════════
# 回測
# ══════════════════════════════════════════════
def run_backtest(data: dict, days: int, min_hit: int,
                 vol_mult: float, min_avg_vol: int,
                 stop_loss: float,
                 take_profit: float = 0.0,
                 exit_day: int = 2,
                 short_only: bool = False,
                 long_only: bool = False,
                 strategy_filter: list = None,
                 workers: int = 4) -> list:
    """
    對所有股票跑 days 天歷史回測。
    使用向量化訊號預計算 + ThreadPoolExecutor 平行處理以大幅加速。

    進出場邏輯（exit_day 控制）：
      D0 收盤訊號確認 → D1 開盤進場
      exit_day=1：D1 收盤出場（隔日當沖）
      exit_day=2：D2 開盤出場（持有一日，預設）
      exit_day=3：D3 開盤出場（持有兩日）

    stop_loss  > 0：D1~D{exit_day} 任一天觸及停損線即出場（估算）
    take_profit > 0：D1~D{exit_day} 任一天觸及止盈線即出場（估算）
    short_only：True 時只統計空方策略
    """

    def _process_stock(item):
        code, df = item
        buf     = days + 35
        df_sl   = df.iloc[-min(buf, len(df)):]
        n       = len(df_sl)
        start_i = max(20, n - days)
        if n < 22 or start_i >= n - exit_day:
            return {k: [] for k in STRATEGY_NAMES}

        # 保留原始日期
        try:
            row_dates = [str(idx.date()) for idx in df_sl.index]
        except Exception:
            row_dates = [str(j) for j in range(n)]

        sub     = df_sl.reset_index(drop=True)
        sig_arr = _precompute_signals_vec(sub, vol_mult, min_avg_vol)  # (n, 18) bool

        local = {k: [] for k in STRATEGY_NAMES}

        for i in range(start_i, n - exit_day):
            row = sig_arr[i]

            long_hits  = int(sum(row[j] for j in range(_N_STRATS) if not _IS_SHORT[j]))
            short_hits = int(sum(row[j] for j in range(_N_STRATS) if _IS_SHORT[j]))

            if max(long_hits, short_hits) < min_hit:
                continue

            d1_open = float(sub.iloc[i + 1]["Open"])
            if d1_open == 0:
                continue

            # 出場價
            if exit_day == 1:
                exit_price = float(sub.iloc[i + 1]["Close"])
            else:
                if i + exit_day >= n:
                    continue
                exit_price = float(sub.iloc[i + exit_day]["Open"])
                if exit_price == 0:
                    continue

            hold_high = max(float(sub.iloc[i + j]["High"]) for j in range(1, exit_day + 1))
            hold_low  = min(float(sub.iloc[i + j]["Low"])  for j in range(1, exit_day + 1))

            date_str = row_dates[i + 1] if i + 1 < len(row_dates) else str(i + 1)

            for ji, k in enumerate(_STRAT_KEYS):
                if not row[ji]:
                    continue
                is_short = _IS_SHORT[ji]
                if short_only and not is_short:
                    continue
                if long_only and is_short:
                    continue
                if strategy_filter and k not in strategy_filter:
                    continue
                hit_count = short_hits if is_short else long_hits
                if hit_count < min_hit:
                    continue

                stopped = False
                if stop_loss > 0:
                    if is_short:
                        stopped = hold_high >= d1_open * (1 + stop_loss / 100)
                    else:
                        stopped = hold_low  <= d1_open * (1 - stop_loss / 100)

                taken = False
                if not stopped and take_profit > 0:
                    if is_short:
                        taken = hold_low  <= d1_open * (1 - take_profit / 100)
                    else:
                        taken = hold_high >= d1_open * (1 + take_profit / 100)

                if stopped:
                    ret_net = -(stop_loss   + TRADE_COST * 100)
                elif taken:
                    ret_net =  take_profit  - TRADE_COST * 100
                elif is_short:
                    ret_net = (d1_open - exit_price) / d1_open * 100 - TRADE_COST * 100
                else:
                    ret_net = (exit_price - d1_open) / d1_open * 100 - TRADE_COST * 100

                local[k].append({
                    "code":    code,
                    "date":    date_str,
                    "ret_net": round(ret_net, 3),
                    "win":     ret_net > 0,
                    "stopped": stopped,
                    "taken":   taken,
                })

        return local

    # ── 平行處理各股票 ────────────────────────────────────────
    trades = {k: [] for k in STRATEGY_NAMES}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for loc in pool.map(_process_stock, data.items()):
            for k in STRATEGY_NAMES:
                trades[k].extend(loc.get(k, []))

    # ── 彙整統計 ──────────────────────────────────────────────
    rows = []
    for k, name in STRATEGY_NAMES.items():
        t = trades[k]
        if not t:
            rows.append({
                "策略": f"{k}｜{name}", "訊號次數": 0,
                "勝率(%)": "-", "平均獲利(%)": "-",
                "平均虧損(%)": "-", "期望值(%)": "-",
                "最大單筆虧損(%)": "-", "停損觸發(%)": "-", "止盈觸發(%)": "-",
            })
            continue
        wins    = [x["ret_net"] for x in t if x["win"]]
        losses  = [x["ret_net"] for x in t if not x["win"]]
        stopped = sum(1 for x in t if x["stopped"])
        taken   = sum(1 for x in t if x.get("taken", False))
        wr      = len(wins) / len(t) * 100
        avg_w   = sum(wins)   / len(wins)   if wins   else 0
        avg_l   = sum(losses) / len(losses) if losses else 0
        ev      = wr/100 * avg_w + (1-wr/100) * avg_l
        max_dd  = min((x["ret_net"] for x in t), default=0)
        stp_pct = round(stopped / len(t) * 100, 1) if stop_loss   > 0 else "-"
        tkp_pct = round(taken   / len(t) * 100, 1) if take_profit > 0 else "-"
        rows.append({
            "策略":              f"{k}｜{name}",
            "訊號次數":          len(t),
            "勝率(%)":           round(wr,     1),
            "平均獲利(%)":       round(avg_w,  3),
            "平均虧損(%)":       round(avg_l,  3),
            "期望值(%)":         round(ev,     3),
            "最大單筆虧損(%)":   round(max_dd, 3),
            "停損觸發(%)":       stp_pct,
            "止盈觸發(%)":       tkp_pct,
        })
    return rows


# ══════════════════════════════════════════════
# 策略組合分析
# ══════════════════════════════════════════════
def run_combo_analysis(data: dict, days: int, min_hit: int,
                       vol_mult: float, min_avg_vol: int,
                       stop_loss: float = 0.0,
                       direction: str = "long",
                       min_signals: int = 100,
                       max_combo: int = 3,
                       exit_day: int = 1,
                       workers: int = 4,
                       forced_combos: list = None) -> list:
    """
    計算所有 2 到 max_combo 策略組合的勝率與期望值。
    每個組合的每筆交易只計算一次，正確反映「同時符合所有條件才進場」的績效。
    使用向量化訊號預計算 + ThreadPoolExecutor 平行處理以大幅加速。

    direction     : "long"（多方組合）或 "short"（空方組合）
    min_signals   : 低於此訊號數的組合標記為 ⚠️ 樣本不足
    forced_combos : 指定必須強制計算的組合清單（即使觸發次數極少），通常來自 show_combos
    max_combo  : 最大組合策略數（預設 3；太大時組合數爆炸且樣本極少）
    exit_day   : 1=D1收盤出場（含漲跌停板）2=D2開盤出場（預設）3=D3開盤出場
    """
    from itertools import combinations as _comb

    is_short_dir = (direction == "short")
    candidates   = [k for k in _STRAT_KEYS
                    if (k.endswith("S") if is_short_dir else not k.endswith("S"))]
    # 候選策略在 _STRAT_KEYS 中的位置索引
    cand_idx     = [_STRAT_KEYS.index(k) for k in candidates]

    # 強制計算組合：解析為 (key, [idx, ...]) 清單
    _forced: list = []
    if forced_combos:
        for fc in forced_combos:
            parts = fc.strip().split("+")
            try:
                idxs = [_STRAT_KEYS.index(p) for p in parts]
                _forced.append((fc, idxs))
            except ValueError:
                pass  # 含有不存在的策略名稱，忽略

    def _process_stock(item):
        code, df = item
        buf     = days + 35
        df_sl   = df.iloc[-min(buf, len(df)):]
        n       = len(df_sl)
        start_i = max(20, n - days)
        if n < 22 or start_i >= n - exit_day:
            return {}

        sub     = df_sl.reset_index(drop=True)
        sig_arr = _precompute_signals_vec(sub, vol_mult, min_avg_vol)  # (n, 18) bool

        local_combo: dict = {}  # key → list[float]

        for i in range(start_i, n - exit_day):
            row = sig_arr[i]
            hit = [candidates[j] for j, cidx in enumerate(cand_idx) if row[cidx]]
            if len(hit) < 2:
                continue

            d0_close = float(sub.iloc[i]["Close"])
            d1       = sub.iloc[i + 1]
            d1_open  = float(d1["Open"])
            d1_high  = float(d1["High"])
            d1_low   = float(d1["Low"])
            d1_close = float(d1["Close"])
            if d1_open == 0 or d0_close == 0:
                continue

            lim_up   = calc_limit_up(d0_close)
            lim_down = calc_limit_down(d0_close)

            # ── 出場計算（只算一次，供所有組合共用）────
            # 停損：持有期間任一天觸線
            hold_high = max(float(sub.iloc[i + j]["High"]) for j in range(1, exit_day + 1))
            hold_low  = min(float(sub.iloc[i + j]["Low"])  for j in range(1, exit_day + 1))
            stopped = False
            if stop_loss > 0:
                if is_short_dir:
                    stopped = hold_high >= d1_open * (1 + stop_loss / 100)
                else:
                    stopped = hold_low  <= d1_open * (1 - stop_loss / 100)

            if stopped:
                ret_net = -(stop_loss + TRADE_COST * 100)
            elif exit_day == 1:
                # D1 收盤出場，含漲跌停板
                if not is_short_dir:
                    exit_p = lim_up   if d1_high >= lim_up   else d1_close
                    ret_net = (exit_p - d1_open) / d1_open * 100 - TRADE_COST * 100
                else:
                    exit_p = lim_down if d1_low  <= lim_down else d1_close
                    ret_net = (d1_open - exit_p) / d1_open * 100 - TRADE_COST * 100
            else:
                # D{exit_day} 開盤出場
                exit_price = float(sub.iloc[i + exit_day]["Open"])
                if exit_price == 0:
                    continue
                if not is_short_dir:
                    ret_net = (exit_price - d1_open) / d1_open * 100 - TRADE_COST * 100
                else:
                    ret_net = (d1_open - exit_price) / d1_open * 100 - TRADE_COST * 100

            ret_net = round(ret_net, 3)

            cap = min(max_combo, len(hit))
            for size in range(2, cap + 1):
                for combo in _comb(hit, size):
                    key = "+".join(combo)
                    if key not in local_combo:
                        local_combo[key] = []
                    local_combo[key].append(ret_net)

            # ── 強制計算 show_combos 指定的組合 ────────────────────
            # 若指定組合的所有策略都在本 bar 觸發，但策略數超過 max_combo
            # 或恰好沒被自然枚舉到（例如只有 B2 一個策略觸發時被跳過），
            # 仍強制計算，確保 show_combos 指定的組合一定有結果。
            for fc_key, fc_idxs in _forced:
                if fc_key in local_combo:
                    continue  # 已由自然枚舉計算過
                if all(bool(row[idx]) for idx in fc_idxs):
                    if fc_key not in local_combo:
                        local_combo[fc_key] = []
                    local_combo[fc_key].append(ret_net)

        return local_combo

    # ── 平行處理各股票，主執行緒合併結果 ─────────────────────
    combo_trades: dict = {}
    total = len(data)
    done  = 0
    print(f"  🔄 策略組合分析：處理 {total} 檔股票（{max_combo} 策略最多同時命中）...")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for local_combo in pool.map(_process_stock, data.items()):
            for key, rets in local_combo.items():
                if key not in combo_trades:
                    combo_trades[key] = []
                combo_trades[key].extend(rets)
            done += 1
            if done % 50 == 0 or done == total:
                print(f"     進度：{done}/{total} 檔  組合數：{len(combo_trades)}", end="\r", flush=True)
    print()  # 換行

    # ── 彙整統計 ──────────────────────────────────
    rows = []
    for key, rets in combo_trades.items():
        n_sig  = len(rets)
        wins   = [r for r in rets if r > 0]
        losses = [r for r in rets if r <= 0]
        wr     = len(wins) / n_sig * 100
        avg_w  = sum(wins)   / len(wins)   if wins   else 0.0
        avg_l  = sum(losses) / len(losses) if losses else 0.0
        ev     = wr / 100 * avg_w + (1 - wr / 100) * avg_l
        max_dd = min(rets)
        rows.append({
            "組合":          key,
            "訊號次數":      n_sig,
            "勝率(%)":       round(wr,    1),
            "平均獲利(%)":   round(avg_w, 3),
            "平均虧損(%)":   round(avg_l, 3),
            "期望值(%)":     round(ev,    3),
            "最大單筆虧損(%)": round(max_dd, 3),
            "樣本":          "⚠️少" if n_sig < min_signals else "OK",
        })

    rows.sort(key=lambda x: (-x["勝率(%)"], -x["期望值(%)"]))
    return rows


def print_combo_result(rows: list, days: int, direction: str,
                       stop_loss: float, min_signals: int = 100,
                       max_combo: int = 3,
                       show_combos: list = None):
    """
    show_combos : 若指定（如 ['C+F+I','B+E+F','B+E','B+D+E']），只顯示這幾個組合。
                  None = 全部顯示。
    """
    dir_str  = "多方" if direction == "long" else "空方"
    sl_str   = f"-{stop_loss}%（估算）" if stop_loss > 0 else "未設定"
    size_str = f"2 ~ {max_combo} 策略同時命中"
    print("\n" + "═"*85)
    print(f"  🔗 策略組合分析（{dir_str}，{size_str}）  回測天數：{days} 日  停損：{sl_str}")
    print(f"  邏輯：組合內所有策略同時觸發才進場，每筆交易只計算一次（無重複計入）")
    if show_combos:
        print(f"  📌 僅顯示指定組合：{', '.join(show_combos)}")
    print(f"  ⚠️少 = 樣本數 < {min_signals} 筆，統計上不可靠，供參考")
    print("═"*85)

    if not rows:
        print("  （無符合條件的策略組合）")
        print("═"*85 + "\n")
        return

    # 若有 show_combos，依指定清單篩選並保持指定順序
    display_rows = rows
    if show_combos:
        want = [c.strip() for c in show_combos]
        row_map = {r["組合"]: r for r in rows}
        display_rows = []
        for key in want:
            if key in row_map:
                display_rows.append(row_map[key])
            else:
                # 組合不存在（樣本為 0）→ 補空行而非僅印警告
                display_rows.append({
                    "組合": key,
                    "訊號次數": 0,
                    "勝率(%)": 0.0,
                    "平均獲利(%)": 0.0,
                    "平均虧損(%)": 0.0,
                    "期望值(%)": 0.0,
                    "最大單筆虧損(%)": 0.0,
                    "樣本": "n=0",
                })
        if not display_rows:
            print("  （指定組合均無資料）")
            print("═"*85 + "\n")
            return
    else:
        # 無指定組合時：分別印出「期望值 Top 30」與「勝率 Top 30」
        reliable = [r for r in rows if r["樣本"] == "OK"]
        total_combos = len(rows)

        print(f"  共計算 {total_combos} 個組合"
              f"（樣本充足 {len(reliable)} 個，⚠️少 {total_combos - len(reliable)} 個）")
        print()

        # ── 期望值 Top 30（樣本充足優先）──────────────────
        top_ev = sorted(reliable, key=lambda x: -x["期望值(%)"])[:30]
        if not top_ev:
            top_ev = sorted(rows, key=lambda x: -x["期望值(%)"])[:30]
        print(f"  📊 期望值 Top {len(top_ev)}（樣本充足，依期望值排序）")
        df_ev = pd.DataFrame(top_ev)
        cols  = ["組合","訊號次數","勝率(%)","平均獲利(%)","平均虧損(%)",
                 "期望值(%)","最大單筆虧損(%)","樣本"]
        print_table(df_ev, cols)

        # ── 勝率 Top 20（樣本充足優先）──────────────────
        top_wr = sorted(reliable, key=lambda x: (-x["勝率(%)"], -x["期望值(%)"]))[:20]
        if top_wr and top_wr[0]["組合"] != top_ev[0]["組合"]:
            print(f"\n  🏆 勝率 Top {len(top_wr)}（樣本充足，依勝率排序）")
            df_wr = pd.DataFrame(top_wr)
            print_table(df_wr, cols)

        print("═"*85 + "\n")
        if reliable:
            best_ev_r = top_ev[0] if top_ev else None
            best_wr_r = top_wr[0] if top_wr else None
            if best_ev_r:
                print(f"  💰 期望值最高：{best_ev_r['組合']}"
                      f"  EV {best_ev_r['期望值(%)']:+.3f}%  勝率 {best_ev_r['勝率(%)']}%"
                      f"  n={best_ev_r['訊號次數']}")
            if best_wr_r and best_wr_r['組合'] != (best_ev_r['組合'] if best_ev_r else ""):
                print(f"  🏆 勝率最高：{best_wr_r['組合']}"
                      f"  勝率 {best_wr_r['勝率(%)']}%  EV {best_wr_r['期望值(%)']:+.3f}%"
                      f"  n={best_wr_r['訊號次數']}")
        print("═"*85 + "\n")
        return

    df = pd.DataFrame(display_rows)
    cols = ["組合","訊號次數","勝率(%)","平均獲利(%)","平均虧損(%)",
            "期望值(%)","最大單筆虧損(%)","樣本"]
    print_table(df, cols)

    reliable = [r for r in display_rows if r["樣本"] == "OK"]
    if reliable:
        best_wr = max(reliable, key=lambda x: x["勝率(%)"])
        best_ev = max(reliable, key=lambda x: x["期望值(%)"])
        print(f"\n  🏆 勝率最高（樣本充足）：{best_wr['組合']}"
              f"  勝率 {best_wr['勝率(%)']}%  期望值 {best_wr['期望值(%)']:+.3f}%"
              f"  訊號 {best_wr['訊號次數']} 次")
        if best_ev["組合"] != best_wr["組合"]:
            print(f"  💰 期望值最高（樣本充足）：{best_ev['組合']}"
                  f"  期望值 {best_ev['期望值(%)']:+.3f}%  勝率 {best_ev['勝率(%)']}%"
                  f"  訊號 {best_ev['訊號次數']} 次")
    print("═"*85 + "\n")


# ══════════════════════════════════════════════
# 漲跌停出場回測
# ══════════════════════════════════════════════
def run_backtest_limit(data: dict, days: int, min_hit: int,
                       vol_mult: float, min_avg_vol: int,
                       stop_loss: float = 0.0,
                       strategy_filter: list = None,
                       return_trades: bool = False,
                       workers: int = 4):
    """
    漲跌停出場回測（隔日沖模式）。
    使用向量化訊號預計算 + ThreadPoolExecutor 平行處理以大幅加速。

    進出場邏輯：
      D0 收盤訊號確認 → D1 開盤進場
      多方：D1 盤中 High ≥ 漲停價 → 漲停價出場；否則 → D1 收盤出場
      空方：D1 盤中 Low  ≤ 跌停價 → 跌停價出場；否則 → D1 收盤出場

    stop_loss > 0（可選）：
      停損判斷優先於漲跌停；以 D1 High/Low 估算是否觸及停損。

    漲停/跌停：依台股最小升降單位精確計算（≈ ±10%，非剛好 10%）。
    """

    def _process_stock(item):
        code, df = item
        buf     = days + 35
        df_sl   = df.iloc[-min(buf, len(df)):]
        n       = len(df_sl)
        start_i = max(20, n - days)
        if n < 22 or start_i >= n - 1:
            return {k: [] for k in STRATEGY_NAMES}

        # 保留原始日期
        try:
            row_dates = [str(idx.date()) for idx in df_sl.index]
        except Exception:
            row_dates = [str(j) for j in range(n)]

        sub     = df_sl.reset_index(drop=True)
        sig_arr = _precompute_signals_vec(sub, vol_mult, min_avg_vol)  # (n, 18) bool

        local = {k: [] for k in STRATEGY_NAMES}

        for i in range(start_i, n - 1):
            row = sig_arr[i]

            long_hits  = int(sum(row[j] for j in range(_N_STRATS) if not _IS_SHORT[j]))
            short_hits = int(sum(row[j] for j in range(_N_STRATS) if _IS_SHORT[j]))

            if max(long_hits, short_hits) < min_hit:
                continue

            d0_close = float(sub.iloc[i]["Close"])
            d1       = sub.iloc[i + 1]
            d1_open  = float(d1["Open"])
            d1_high  = float(d1["High"])
            d1_low   = float(d1["Low"])
            d1_close = float(d1["Close"])

            if d1_open == 0 or d0_close == 0:
                continue

            lim_up   = calc_limit_up(d0_close)
            lim_down = calc_limit_down(d0_close)
            date_str = row_dates[i + 1] if i + 1 < len(row_dates) else str(i + 1)

            for ji, k in enumerate(_STRAT_KEYS):
                if not row[ji]:
                    continue
                is_short = _IS_SHORT[ji]
                if strategy_filter and k not in strategy_filter:
                    continue
                hit_count = short_hits if is_short else long_hits
                if hit_count < min_hit:
                    continue

                # ── 停損判斷（優先） ─────────────────────────
                stopped = False
                if stop_loss > 0:
                    if is_short:
                        stopped = d1_high >= d1_open * (1 + stop_loss / 100)
                    else:
                        stopped = d1_low  <= d1_open * (1 - stop_loss / 100)

                # ── 出場計算 ──────────────────────────────────
                if stopped:
                    ret_net   = -(stop_loss + TRADE_COST * 100)
                    exit_type = "停損"
                    exit_p    = (d1_open * (1 - stop_loss / 100) if not is_short
                                 else d1_open * (1 + stop_loss / 100))
                elif not is_short:
                    if d1_high >= lim_up:
                        exit_p    = lim_up
                        exit_type = "漲停"
                    else:
                        exit_p    = d1_close
                        exit_type = "收盤"
                    ret_net = (exit_p - d1_open) / d1_open * 100 - TRADE_COST * 100
                else:
                    if d1_low <= lim_down:
                        exit_p    = lim_down
                        exit_type = "跌停"
                    else:
                        exit_p    = d1_close
                        exit_type = "收盤"
                    ret_net = (d1_open - exit_p) / d1_open * 100 - TRADE_COST * 100

                local[k].append({
                    "code":      code,
                    "name":      get_name(code),
                    "date":      date_str,
                    "d0_close":  round(d0_close, 2),
                    "lim_up":    round(lim_up,   2),
                    "lim_down":  round(lim_down, 2),
                    "d1_open":   round(d1_open,  2),
                    "d1_high":   round(d1_high,  2),
                    "d1_low":    round(d1_low,   2),
                    "d1_close":  round(d1_close, 2),
                    "exit_p":    round(exit_p,   2),
                    "exit_type": exit_type,
                    "ret_net":   round(ret_net, 3),
                    "win":       ret_net > 0,
                    "stopped":   stopped,
                    "lim_exit":  (not stopped) and exit_type in ("漲停", "跌停"),
                    "is_short":  is_short,
                    "strategy":  k,
                })

        return local

    # ── 平行處理各股票 ────────────────────────────────────────
    trades = {k: [] for k in STRATEGY_NAMES}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for loc in pool.map(_process_stock, data.items()):
            for k in STRATEGY_NAMES:
                trades[k].extend(loc.get(k, []))

    # ── 彙整統計 ──────────────────────────────────
    rows = []
    for k, name in STRATEGY_NAMES.items():
        t = trades[k]
        if not t:
            rows.append({
                "策略":          f"{k}｜{name}",
                "訊號次數":      0,
                "勝率(%)":       "-",
                "平均獲利(%)":   "-",
                "平均虧損(%)":   "-",
                "期望值(%)":     "-",
                "最大單筆虧損(%)": "-",
                "停損觸發(%)":   "-",
                "漲跌停出場(%)": "-",
            })
            continue

        wins    = [x["ret_net"] for x in t if x["win"]]
        losses  = [x["ret_net"] for x in t if not x["win"]]
        stopped = sum(1 for x in t if x["stopped"])
        lim_ex  = sum(1 for x in t if x.get("lim_exit", False))
        wr      = len(wins) / len(t) * 100
        avg_w   = sum(wins)   / len(wins)   if wins   else 0.0
        avg_l   = sum(losses) / len(losses) if losses else 0.0
        ev      = wr / 100 * avg_w + (1 - wr / 100) * avg_l
        max_dd  = min((x["ret_net"] for x in t), default=0)
        stp_pct = round(stopped / len(t) * 100, 1) if stop_loss > 0 else "-"
        lim_pct = round(lim_ex  / len(t) * 100, 1)

        rows.append({
            "策略":              f"{k}｜{name}",
            "訊號次數":          len(t),
            "勝率(%)":           round(wr,     1),
            "平均獲利(%)":       round(avg_w,  3),
            "平均虧損(%)":       round(avg_l,  3),
            "期望值(%)":         round(ev,     3),
            "最大單筆虧損(%)":   round(max_dd, 3),
            "停損觸發(%)":       stp_pct,
            "漲跌停出場(%)":     lim_pct,
        })

    if return_trades:
        # 回傳 (統計彙整, 明細交易清單)
        all_trades = []
        for k, t in trades.items():
            all_trades.extend(t)
        all_trades.sort(key=lambda x: x["date"])
        return rows, all_trades
    return rows


# ══════════════════════════════════════════════
# 顯示工具
# ══════════════════════════════════════════════
def _dw(s: str) -> int:
    w = 0
    for c in str(s):
        w += 2 if unicodedata.east_asian_width(c) in ('W','F','A') else 1
    return w

def _pad(s: str, width: int, align: str = "left") -> str:
    s   = str(s)
    pad = max(0, width - _dw(s))
    return (" " * pad + s) if align == "right" else (s + " " * pad)

RIGHT = {"收盤","漲跌幅(%)","成交量(張)","命中數","訊號次數",
         "勝率(%)","平均獲利(%)","平均虧損(%)","期望值(%)",
         "最大單筆虧損(%)","停損觸發(%)"}

def _wrap_header(text: str, width: int) -> list:
    """將標題文字切成每行最多 width 顯示寬度的多行清單。"""
    lines, cur, cur_w = [], "", 0
    for ch in text:
        ch_w = _dw(ch)
        if cur_w + ch_w > width:
            lines.append(cur)
            cur, cur_w = ch, ch_w
        else:
            cur += ch
            cur_w += ch_w
    if cur:
        lines.append(cur)
    return lines or [""]

def print_table(df: pd.DataFrame, cols: list, title: str = ""):
    if title:
        print(f"\n  {title}")
    widths = {}
    for col in cols:
        if col not in df.columns:
            continue
        # 欄寬只由「資料」決定，標題過長時自動換行 —— 最小 3 避免太窄
        data_w = max((_dw(str(v)) for v in df[col]), default=0)
        widths[col] = max(data_w + 1, 3)
    cols = [c for c in cols if c in widths]

    # 建立多行標題
    wrapped = [_wrap_header(c, widths[c]) for c in cols]
    n_lines = max(len(w) for w in wrapped)
    # 短標題補空行至頂端對齊
    for w in wrapped:
        while len(w) < n_lines:
            w.insert(0, "")

    for i in range(n_lines):
        print("  " + " ".join(
            _pad(wrapped[j][i], widths[c], "right" if c in RIGHT else "left")
            for j, c in enumerate(cols)
        ))

    sep = "  " + "─" * (sum(widths.values()) + len(cols))
    print(sep)
    for _, row in df.iterrows():
        line = "  " + " ".join(
            _pad(str(row[c]), widths[c], "right" if c in RIGHT else "left")
            for c in cols
        )
        print(line)


def save_scan_xlsx(df: pd.DataFrame, fname: str) -> None:
    """
    將選股結果 DataFrame 輸出為美化的 Excel 檔案。
    - 標頭：深藍底白字，凍結首列，Auto Filter
    - 資料列：依「價位區間」欄位著色
    - 訊號欄：✅ 綠底綠字 / ❌ 保留列底色灰字
    - 漲跌幅：正紅負綠；方向：多紅空綠
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        csv_fname = fname.replace(".xlsx", ".csv")
        df.to_csv(csv_fname, index=False, encoding="utf-8-sig")
        print(f"  ⚠️  找不到 openpyxl，已改存 CSV：{csv_fname}")
        print(f"  💡 安裝方式：pip install openpyxl")
        return

    # ── 顏色定義（openpyxl 需要 ARGB 8 碼，FF = 完全不透明）────
    HDR_BG   = "FF1F4E79"   # 標頭背景（深藍）
    HDR_FG   = "FFFFFFFF"   # 標頭文字（白）

    GROUP_BG = {            # 各價位區間列底色
        "<50":      "FFDEEBF7",   # 淡藍
        "50~100":   "FFE2EFDA",   # 淡綠
        "100~300":  "FFFFF2CC",   # 淡黃
        "300~500":  "FFFCE4D6",   # 淡橙
        "500~1000": "FFFFE6E6",   # 淡紅
        ">1000":    "FFE8D5F0",   # 淡紫
    }
    SIG_YES_BG = "FFC6EFCE"   # ✅ 底色（淡綠）
    SIG_YES_FG = "FF375623"   # ✅ 字色（深綠）
    SIG_NO_FG  = "FFBFBFBF"   # ❌ 字色（灰）
    BULL_FG    = "FFC00000"   # 多/正漲跌幅 紅
    BEAR_FG    = "FF375623"   # 空/負漲跌幅 綠

    # ── 欄寬設定 ────────────────────────────────────────────
    FIXED_WIDTHS = {
        "代號": 8, "名稱": 11, "資料日期": 11,
        "收盤": 8, "價位區間": 10,
        "漲跌幅(%)": 9, "成交量(張)": 10, "量/均量": 7,
        "命中數": 6, "方向": 5, "MA20方向": 11,
        "策略型態": 13, "策略清單": 28, "命中組合": 36,
    }
    MA20_UP_FG   = "FF375623"   # 升 → 深綠
    MA20_DN_FG   = "FFC00000"   # 降 → 深紅
    MA20_FL_FG   = "FF7F6000"   # 橫 → 深黃褐
    SIG_WIDTH = 5   # 訊號欄統一寬度

    # ── 訊號欄自動偵測並移到最右邊 ──────────────────────────────
    _all_cols = list(df.columns)
    sig_cols: set = {
        c for c in _all_cols
        if df[c].dropna().isin(["✅", "❌"]).all() and not df[c].dropna().empty
    }
    # 非訊號欄保持原順序，訊號欄整批附到最後
    non_sig = [c for c in _all_cols if c not in sig_cols]
    sig_ordered = [c for c in _all_cols if c in sig_cols]
    headers = non_sig + sig_ordered
    df = df[headers]

    wb = Workbook()
    ws = wb.active
    ws.title = "選股結果"

    # ── 標頭列 ──────────────────────────────────────────────
    hdr_font  = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    hdr_fill  = PatternFill("solid", start_color=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── 資料列 ──────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        grp    = str(row.get("價位區間", ""))
        row_bg = GROUP_BG.get(grp, "FFFFFF")
        row_fill = PatternFill("solid", start_color=row_bg)

        for ci, col in enumerate(headers, 1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci, value=val)

            if col in sig_cols:
                # 訊號欄
                if val == "✅":
                    cell.fill  = PatternFill("solid", start_color=SIG_YES_BG)
                    cell.font  = Font(name="Arial", size=9, color=SIG_YES_FG)
                else:
                    cell.fill  = row_fill
                    cell.font  = Font(name="Arial", size=9, color=SIG_NO_FG)
                cell.alignment = Alignment(horizontal="center")

            elif col == "漲跌幅(%)":
                try:
                    fg = BULL_FG if float(val) >= 0 else BEAR_FG
                except (TypeError, ValueError):
                    fg = "000000"
                cell.fill      = row_fill
                cell.font      = Font(name="Arial", size=9, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="right")

            elif col == "方向":
                fg = BULL_FG if val == "多" else BEAR_FG
                cell.fill      = row_fill
                cell.font      = Font(name="Arial", size=9, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center")

            elif col == "MA20方向":
                sv = str(val)
                if sv.startswith("升"):
                    fg = MA20_UP_FG
                elif sv.startswith("降"):
                    fg = MA20_DN_FG
                elif sv.startswith("橫"):
                    fg = MA20_FL_FG
                else:
                    fg = "FF808080"
                cell.fill      = row_fill
                cell.font      = Font(name="Arial", size=9, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center")

            elif col == "命中數":
                cell.fill      = row_fill
                cell.font      = Font(name="Arial", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center")

            elif col == "策略型態":
                sv = str(val)
                if "回歸" in sv:
                    bg, fg = "FFE1F5EE", "FF0F6E56"   # 綠底深綠字（均值回歸）
                elif "跟風" in sv:
                    bg, fg = "FFE6F1FB", "FF185FA5"   # 藍底深藍字（趨勢跟風）
                elif "混合" in sv:
                    bg, fg = "FFFAEEDA", "FF854F0B"   # 橘底深橘字（混合型）
                elif "缺口" in sv:
                    bg, fg = "FFFBEAF0", "FF993556"   # 粉底深粉字（缺口動能）
                else:
                    bg, fg = "00000000", "FF808080"
                cell.fill      = PatternFill("solid", start_color=bg) if bg != "00000000" else row_fill
                cell.font      = Font(name="Arial", size=9, bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center")

            elif col in ("代號", "名稱"):
                cell.fill = row_fill
                cell.font = Font(name="Arial", size=9, bold=True)

            else:
                cell.fill = row_fill
                cell.font = Font(name="Arial", size=9)
                if col in ("收盤", "成交量(張)", "量/均量"):
                    cell.alignment = Alignment(horizontal="right")

    # ── 欄寬 ────────────────────────────────────────────────
    for ci, col in enumerate(headers, 1):
        letter = get_column_letter(ci)
        if col in sig_cols:
            ws.column_dimensions[letter].width = SIG_WIDTH
        else:
            ws.column_dimensions[letter].width = FIXED_WIDTHS.get(col, 12)

    wb.save(fname)


# ── 股價分組設定（共用）──────────────────────────────────────
_PRICE_BUCKETS = [
    (0,    50,   "<50",       "🔵  低價股  ( 收盤 < 50 )"),
    (50,   100,  "50~100",    "🟢  中低價  ( 50 ~ 100 )"),
    (100,  300,  "100~300",   "🟡  中價股  ( 100 ~ 300 )"),
    (300,  500,  "300~500",   "🟠  中高價  ( 300 ~ 500 )"),
    (500,  1000, "500~1000",  "🔴  高價股  ( 500 ~ 1000 )"),
    (1000, None, ">1000",     "🟣  超高價  ( > 1000 )"),
]

def price_group(close) -> str:
    """依收盤價回傳簡短的價位區間標籤，供 CSV 欄位 & 終端機分組共用。"""
    try:
        c = float(close)
    except (TypeError, ValueError):
        return "N/A"
    for lo, hi, label, _ in _PRICE_BUCKETS:
        if hi is None:
            return label
        if c < hi:
            return label
    return ">1000"


def print_scan_result(rows: list, date_str: str, min_hit: int,
                      show_combos: list = None):

    RESET = "\033[0m"

    print("\n" + "═"*100)
    print(f"  📋 隔日沖選股彙整  {date_str}  命中門檻：≥{min_hit}")
    if show_combos:
        print(f"  📌 組合篩選：{', '.join(show_combos)}（至少命中其中一個）")
    print("═"*100)
    print(f"  通過門檻：{len(rows)} 檔")
    print("═"*100)
    if not rows:
        print("  （今日無股票達到命中門檻）")
        return

    df = pd.DataFrame(rows)
    df["_close_f"] = pd.to_numeric(df["收盤"], errors="coerce").fillna(0)

    # ── 依股價分組輸出（文字版，不印訊號矩陣）────────────────────
    any_printed = False
    for lo, hi, _label_short, label_long in _PRICE_BUCKETS:
        if hi is None:
            mask = df["_close_f"] >= lo
        else:
            mask = (df["_close_f"] >= lo) & (df["_close_f"] < hi)
        grp = df[mask].sort_values("_close_f").copy()
        if grp.empty:
            continue
        any_printed = True
        print(f"\n  {label_long}  （{len(grp)} 檔）")
        print("  " + "─"*98)
        for _, r in grp.iterrows():
            # 從策略清單取出當日觸發訊號，依方向過濾
            strats_raw = str(r.get("策略清單", ""))
            all_sigs   = [s for s in strats_raw.split(",") if s.strip()]
            direction  = str(r.get("方向", "多"))
            if direction == "多":
                active = [s for s in all_sigs if not s.endswith("S")]
            else:
                active = [s for s in all_sigs if s.endswith("S") or s in ("B","B2")]
            sig_str  = " ".join(active) if active else "—"
            dir_icon = "▲" if direction == "多" else "▽"
            chg      = float(r.get("漲跌幅(%)", 0))
            chg_color = "\033[31m" if chg > 0 else ("\033[32m" if chg < 0 else "")
            vol_k    = int(r.get("成交量(張)", 0))
            vol_r    = str(r.get("量/均量", "—"))
            style    = str(r.get("策略型態", ""))
            ma20     = str(r.get("MA20方向", ""))
            combo    = str(r.get("命中組合", "—"))
            combo_str = f"  組合：{combo}" if combo != "—" else ""
            style_str = f"  {style}" if style else ""
            ma20_str  = f"  {ma20}" if ma20 and ma20 not in ("—", "") else ""
            vol_field = _pad(f"{vol_k}張({vol_r})", 15)
            print(f"  │  {_pad(str(r['代號']), 7)}{_pad(str(r['名稱']), 12)}"
                  f"{float(r['收盤']):>9.2f}  "
                  f"{chg_color}{chg:>+7.2f}%{RESET}  "
                  f"{vol_field}  "
                  f"{dir_icon} [{sig_str}]  命中{r['命中數']}"
                  f"{style_str}{ma20_str}{combo_str}")

    if not any_printed:
        print("  （今日無股票達到命中門檻）")

    print("═"*100)
    print("  訊號：A=均線突破  B=跳空↑  B2=大跳空≥5%  C=RSI超賣  D=突破前高(多頭排列)  "
          "E=強勢連漲  F=均量擴張  G=縮後上漲")
    print("        H=鎚子K  I=吞噬陽  J=MACD金叉  K=布林下軌  L=KD<20金叉  "
          "M=威廉超賣  N=多頭排列回測  O=晨星")
    print("        P=紅三兵  Q=IB突破  R=BIAS<-10%反彈  空方加 S 後綴（AS/BS/CS...）")
    print("═"*100 + "\n")


def print_backtest_result(stats: list, days: int, min_hit: int, stop_loss: float,
                          take_profit: float = 0.0, exit_day: int = 2,
                          short_only: bool = False, long_only: bool = False,
                          strategy_filter: list = None):
    sl_str = f"-{stop_loss}%（估算）" if stop_loss > 0 else "未設定"
    print("\n" + "═"*80)
    print(f"  📊 隔日沖回測  回測天數：{days} 日  命中門檻：≥{min_hit}  停損：{sl_str}")
    exit_str = {1:"D1收盤（隔日當沖）", 2:"D2開盤（持一日）", 3:"D3開盤（持兩日）"}.get(exit_day, f"D{exit_day}開盤")
    tp_str   = f"  止盈：+{take_profit}%（估算）" if take_profit > 0 else ""
    if strategy_filter:
        so_str = f"  │  策略篩選：{','.join(strategy_filter)}"
    elif short_only:
        so_str = "  │  只計算空方"
    elif long_only:
        so_str = "  │  只計算多方"
    else:
        so_str = ""
    print(f"  進場：D1開盤  │  出場：{exit_str}  │  手續費：{TRADE_COST*100:.3f}%{so_str}")
    print(f"  ⚠️  停損/止盈為估算值（日線 High/Low）{tp_str}")
    print("═"*80)
    df = pd.DataFrame(stats)
    # long_only 模式：只顯示多方策略
    if strategy_filter:
        # 只保留指定策略（格式如 "I｜多 吞噬陽線" 開頭的代號）
        pattern = "|".join(f"^{k}｜" for k in strategy_filter)
        df = df[df["策略"].str.match(pattern)].copy()
    elif long_only:
        df = df[~df["策略"].str.contains("空")].copy()
    elif short_only:
        df = df[~df["策略"].str.contains("多")].copy()
    cols = ["策略","訊號次數","勝率(%)","平均獲利(%)","平均虧損(%)",
            "期望值(%)","最大單筆虧損(%)","停損觸發(%)"]
    print_table(df, cols)

    # 最佳策略
    numeric = df[df["勝率(%)"] != "-"].copy()
    if not numeric.empty:
        numeric["期望值(%)"] = numeric["期望值(%)"].astype(float)
        best = numeric.loc[numeric["期望值(%)"].idxmax()]
        print(f"\n  🏆 最佳：{best['策略']}（期望值 {best['期望值(%)']:.3f}%）")
        long_best  = numeric[numeric["策略"].str.contains("多")].nlargest(1, "期望值(%)")
        short_best = numeric[numeric["策略"].str.contains("空")].nlargest(1, "期望值(%)")
        if not long_best.empty and not short_only:
            r = long_best.iloc[0]
            print(f"  📈 多方最佳：{r['策略']}（{r['期望值(%)']:.3f}%）")
        if not short_best.empty and not long_only:
            r = short_best.iloc[0]
            print(f"  📉 空方最佳：{r['策略']}（{r['期望值(%)']:.3f}%）")
    print("═"*80 + "\n")


def print_trade_detail(all_trades: list, strategy_filter: list = None,
                       long_only: bool = False, short_only: bool = False,
                       show_n: int = 20):
    """
    印出回測明細，每筆交易顯示：
      D0 收盤 → 漲/跌停價 → D1 開/高/低/收 → 出場價 → 損益
    """
    trades = all_trades
    if strategy_filter:
        trades = [t for t in trades if t["strategy"] in strategy_filter]
    if long_only:
        trades = [t for t in trades if not t["is_short"]]
    if short_only:
        trades = [t for t in trades if t["is_short"]]

    if not trades:
        print("  （無符合條件的明細交易）")
        return

    print(f"\n  🔍 回測明細（共 {len(trades)} 筆，顯示前 {min(show_n, len(trades))} 筆）")
    print("  " + "─" * 105)
    hdr = (f"  {'日期(D1)':<12} {'代號':<6} {'名稱':<8} {'策略':<4}"
           f" {'D0收盤':>7} {'漲停價':>7} {'跌停價':>7}"
           f" {'D1開':>7} {'D1高':>7} {'D1低':>7} {'D1收':>7}"
           f" {'出場價':>7} {'出場方式':<6} {'損益%':>7}")
    print(hdr)
    print("  " + "─" * 105)
    for tr in trades[:show_n]:
        direction = "空" if tr["is_short"] else "多"
        win_mark  = "✅" if tr["win"] else "❌"
        print(
            f"  {tr['date']:<12} {tr['code']:<6} {tr['name']:<8} "
            f"{tr['strategy']:<4}({direction})"
            f" {tr['d0_close']:>7.2f} {tr['lim_up']:>7.2f} {tr['lim_down']:>7.2f}"
            f" {tr['d1_open']:>7.2f} {tr['d1_high']:>7.2f} {tr['d1_low']:>7.2f} {tr['d1_close']:>7.2f}"
            f" {tr['exit_p']:>7.2f} {tr['exit_type']:<6} {tr['ret_net']:>+7.3f}% {win_mark}"
        )
    if len(trades) > show_n:
        print(f"  ... 還有 {len(trades) - show_n} 筆（使用 --show-trades N 增加顯示數量）")
    print("  " + "─" * 105)
    print(f"  驗證方式：取上表任一列，以『日期(D1)』當天去 K 線圖確認")
    print(f"    D1開盤=當日開盤價  D1高=當日最高  D1低=當日最低  D1收=當日收盤")
    print(f"    漲停價=前日收盤×1.1（捨去）  跌停價=前日收盤×0.9（進位）\n")


def print_backtest_limit_result(stats: list, days: int, min_hit: int,
                                stop_loss: float,
                                strategy_filter: list = None):
    """顯示漲跌停出場回測結果"""
    sl_str = f"-{stop_loss}%（估算）" if stop_loss > 0 else "未設定"
    print("\n" + "═"*90)
    print(f"  📊 漲跌停出場回測  回測天數：{days} 日  命中門檻：≥{min_hit}  停損：{sl_str}")
    print(f"  進場：D1 開盤  │  多方出場：漲停價 或 D1 收盤"
          f"  │  空方出場：跌停價 或 D1 收盤")
    print(f"  手續費：{TRADE_COST*100:.3f}%  │  漲跌停依台股最小升降單位精確計算（≈±10%）")
    if strategy_filter:
        print(f"  策略篩選：{','.join(strategy_filter)}")
    print("═"*90)

    df = pd.DataFrame(stats)
    if strategy_filter:
        pattern = "|".join(f"^{k}｜" for k in strategy_filter)
        df = df[df["策略"].str.match(pattern)].copy()

    cols = ["策略","訊號次數","勝率(%)","平均獲利(%)","平均虧損(%)",
            "期望值(%)","最大單筆虧損(%)","停損觸發(%)","漲跌停出場(%)"]
    print_table(df, cols)

    numeric = df[df["勝率(%)"] != "-"].copy()
    if not numeric.empty:
        numeric["期望值(%)"] = numeric["期望值(%)"].astype(float)
        best = numeric.loc[numeric["期望值(%)"].idxmax()]
        print(f"\n  🏆 最佳策略：{best['策略']}  期望值 {best['期望值(%)']:+.3f}%"
              f"  勝率 {best['勝率(%)']}%  訊號 {best['訊號次數']} 次")
        long_best  = numeric[numeric["策略"].str.contains("多")].nlargest(1, "期望值(%)")
        short_best = numeric[numeric["策略"].str.contains("空")].nlargest(1, "期望值(%)")
        if not long_best.empty:
            r = long_best.iloc[0]
            print(f"  📈 多方最佳：{r['策略']}  期望值 {r['期望值(%)']:+.3f}%  勝率 {r['勝率(%)']}%")
        if not short_best.empty:
            r = short_best.iloc[0]
            print(f"  📉 空方最佳：{r['策略']}  期望值 {r['期望值(%)']:+.3f}%  勝率 {r['勝率(%)']}%")
    print("═"*90)
    print("  ⚠️  停損以 D1 High/Low 估算；同日若同時觸及停損與漲跌停，停損優先。")
    print("  ⚠️  漲跌停出場率低不代表策略差，代表多數交易在尾盤才收盤出場。\n")


def print_history(df: pd.DataFrame):
    if df.empty:
        print("  （無歷史記錄）")
        return
    print_table(df, df.columns.tolist(), title="歷史選股記錄")


# ══════════════════════════════════════════════
# 診斷工具：比對兩個資料源
# ══════════════════════════════════════════════
def diagnose_sources(code: str, days: int,
                     sj_key: str = "", sj_secret: str = "",
                     tail: int = 10):
    """
    比對 yfinance 與 Shioaji 對同一檔股票的最近 N 根 K 棒。
    找出日期缺口、價格差異、成交量差異。
    """
    print(f"\n{'═'*70}")
    print(f"  🔬 資料源診斷：{code}  最近 {tail} 根日K")
    print(f"{'═'*70}")

    # ── 取 yfinance 資料 ──
    yf_df = None
    try:
        yf_data = fetch_data([code], days)
        if code in yf_data:
            yf_df = yf_data[code].tail(tail)[["Open","High","Low","Close","Vol_K"]].copy()
            yf_df.index = pd.to_datetime(yf_df.index).date
    except Exception as e:
        print(f"  ⚠️  yfinance 取得失敗：{e}")

    # ── 取 Shioaji 資料 ──
    sj_df = None
    if sj_key and sj_secret:
        try:
            sj_data = fetch_data_sinopac([code], days, sj_key, sj_secret)
            if code in sj_data:
                sj_df = sj_data[code].tail(tail)[["Open","High","Low","Close","Vol_K"]].copy()
                sj_df.index = pd.to_datetime(sj_df.index).date
        except Exception as e:
            print(f"  ⚠️  Shioaji 取得失敗：{e}")
    else:
        print(f"  ℹ️  未提供 Shioaji 憑證，只顯示 yfinance 資料")

    # ── 比對輸出 ──
    if yf_df is not None:
        print(f"\n  【yfinance】最後一筆日期：{yf_df.index[-1]}  收盤：{yf_df['Close'].iloc[-1]}")
        print(f"  {'日期':<12} {'開':>7} {'高':>7} {'低':>7} {'收':>7} {'量(張)':>9}")
        print(f"  {'─'*55}")
        for dt, row in yf_df.iterrows():
            print(f"  {str(dt):<12} {row['Open']:>7.1f} {row['High']:>7.1f} "
                  f"{row['Low']:>7.1f} {row['Close']:>7.1f} {row['Vol_K']:>9.0f}")

    if sj_df is not None:
        print(f"\n  【Shioaji】最後一筆日期：{sj_df.index[-1]}  收盤：{sj_df['Close'].iloc[-1]}")
        print(f"  {'日期':<12} {'開':>7} {'高':>7} {'低':>7} {'收':>7} {'量(張)':>9}")
        print(f"  {'─'*55}")
        for dt, row in sj_df.iterrows():
            print(f"  {str(dt):<12} {row['Open']:>7.1f} {row['High']:>7.1f} "
                  f"{row['Low']:>7.1f} {row['Close']:>7.1f} {row['Vol_K']:>9.0f}")

    # ── 差異摘要 ──
    if yf_df is not None and sj_df is not None:
        common = sorted(set(yf_df.index) & set(sj_df.index))
        yf_only = sorted(set(yf_df.index) - set(sj_df.index))
        sj_only = sorted(set(sj_df.index) - set(yf_df.index))
        print(f"\n  📊 差異摘要：")
        print(f"     共同日期：{len(common)} 天")
        if yf_only:
            print(f"     ⚠️  yfinance 有但 Shioaji 沒有：{yf_only}")
        if sj_only:
            print(f"     ⚠️  Shioaji 有但 yfinance 沒有：{sj_only}")
        if common:
            max_close_diff = max(
                abs(yf_df.loc[d, "Close"] - sj_df.loc[d, "Close"])
                for d in common if d in yf_df.index and d in sj_df.index
            )
            max_vol_ratio = max(
                yf_df.loc[d, "Vol_K"] / sj_df.loc[d, "Vol_K"]
                if sj_df.loc[d, "Vol_K"] > 0 else 0
                for d in common if d in yf_df.index and d in sj_df.index
            )
            print(f"     最大收盤價差：{max_close_diff:.2f} 元")
            print(f"     yfinance/Shioaji 成交量比：{max_vol_ratio:.3f}x"
                  f"  {'（接近 1 = 單位相同）' if 0.9 < max_vol_ratio < 1.1 else '（⚠️ 單位可能不同）'}")

    print(f"{'═'*70}\n")


# ══════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="台股隔日沖選股與回測系統",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例：
  python swing_trade.py                         # 選股（預設，等同 --scan）
  python swing_trade.py --min-hit 2 --save      # 選股，命中 ≥2，儲存結果
  python swing_trade.py --backtest --days 60    # 回測 60 日
  python swing_trade.py --history --limit 10   # 查詢最近 10 筆歷史
  python swing_trade.py --history --code 2330  # 查詢 2330 歷史
        """,
    )

    # ── 執行模式 ──────────────────────────────────
    mode_group = parser.add_argument_group("執行模式（預設：--scan）")
    mode_ex = mode_group.add_mutually_exclusive_group()
    mode_ex.add_argument("--list",           action="store_true",
                         help="列出所有訊號代碼（含條件說明）與 preset 組合清單，無需資料源直接印出")
    mode_ex.add_argument("--scan",           action="store_true",
                         help="執行選股（預設，無參數時自動啟用）")
    mode_ex.add_argument("--backtest",       action="store_true",
                         help="執行回測（D+N 開盤出場）")
    mode_ex.add_argument("--backtest-limit", action="store_true",
                         help="執行漲跌停出場回測（D1 盤中漲停/跌停 或 D1 收盤出場）")
    mode_ex.add_argument("--combo",          action="store_true",
                         help="策略組合分析：計算所有兩策略雙重確認的勝率與期望值")
    mode_ex.add_argument("--history",        action="store_true",
                         help="查詢歷史選股記錄")
    mode_ex.add_argument("--diagnose",       type=str, default=None,
                         metavar="CODE",
                         help="診斷：比對兩個資料源對指定股票的 K 棒差異，如 --diagnose 4540")

    # ── 選股 / 量能篩選 ───────────────────────────
    scan_group = parser.add_argument_group("選股 / 量能篩選")
    scan_group.add_argument("--top-n",   type=int,   default=300,
                            help="掃描成交量前 N 名，預設 300")
    scan_group.add_argument("--min-vol", type=int,   default=MIN_AVG_VOL,
                            help=f"5日均量最低門檻（張），預設 {MIN_AVG_VOL}")
    scan_group.add_argument("--vol-mult",type=float, default=VOL_MULT,
                            help=f"爆量倍數，預設 {VOL_MULT}")
    scan_group.add_argument("--min-hit", type=int,   default=1,
                            help="策略命中門檻，預設 1，建議設 2")
    scan_group.add_argument("--save",    action="store_true",
                            help="儲存選股結果到 SQLite 和 CSV")
    scan_group.add_argument("--date",    type=str, default=None, metavar="YYYY-MM-DD",
                            help="指定掃描基準日（預設：今日），格式 YYYY-MM-DD，"
                                 "例：--date 2026-04-22 可回顧該日訊號")

    # ── 進出場設定 ────────────────────────────────
    trade_group = parser.add_argument_group("進出場設定（選股 / 回測共用）")
    trade_group.add_argument("--stop-loss",   type=float, default=7.0,
                             help="停損門檻（%%），0=不設停損，預設 7.0"
                                  "（2026-04 停損敏感度測試建議值：-2%% 太緊會毀掉 EV；"
                                  "-7%% 幾乎等同無停損，但能擋掉尾端風險）")
    trade_group.add_argument("--take-profit", type=float, default=0.0,
                             help="止盈門檻（%%），0=不設止盈，如 2.0=漲 2%% 出場（估算值）")
    trade_group.add_argument("--exit-day",    type=int,   default=None,
                             help="出場日：1=隔日當沖(D1收盤)  2=持一日(D2開盤)  3=持兩日(D3開盤)  "
                                  "未指定時：--combo 模式從 PRESET_EXIT_DAY 查表，其他模式預設 2")

    # ── 策略篩選 ──────────────────────────────────
    strat_group = parser.add_argument_group("策略篩選")
    strat_ex = strat_group.add_mutually_exclusive_group()
    strat_ex.add_argument("--long-only",  action="store_true",
                          help="只顯示多方策略（市場偏多時適用）")
    strat_ex.add_argument("--short-only", action="store_true",
                          help="只顯示空方策略（市場偏空時適用）")
    strat_group.add_argument("--strategy", type=str, default=None,
                             help="只看指定策略，逗號分隔，如 I,D,C 或 I,BS（大小寫均可）")

    # ── 回測專用 ──────────────────────────────────
    bt_group = parser.add_argument_group("回測專用")
    bt_group.add_argument("--days", type=int, default=DEFAULT_DAYS,
                          help=f"回測天數，預設 {DEFAULT_DAYS}")
    bt_group.add_argument("--show-trades", type=int, default=0, metavar="N",
                          help="顯示前 N 筆個別交易明細（搭配 --backtest-limit 使用，"
                               "用於人工驗證回測邏輯，如 --show-trades 20）")
    bt_group.add_argument("--min-signals", type=int, default=100, metavar="N",
                          help="組合分析：低於 N 筆的組合標記為樣本不足（預設 100，統計顯著門檻）")
    bt_group.add_argument("--max-combo",  type=int, default=3, metavar="N",
                          help="組合分析：最大組合策略數（預設 3；設 2 只看兩兩配對）")
    bt_group.add_argument("--workers",     type=int, default=4, metavar="N",
                          help="回測平行執行緒數（預設 4；設 1 停用多執行緒）")
    bt_group.add_argument("--show-combos", type=str, default=None, metavar="COMBOS",
                          help="只顯示指定組合（逗號分隔），"
                               "如 --show-combos \"BS+GS,AS+BS+GS\"")
    bt_group.add_argument("--preset",      type=str, default=None, metavar="NAME[,NAME2...]",
                          help=f"使用預設組合清單（等同 --show-combos 展開版），"
                               f"支援逗號分隔多個 preset，例如：long3_lean,short3_lean。"
                               f"可用：{', '.join(COMBO_PRESETS.keys())}")

    # ── 資料來源 ──────────────────────────────────
    src_group = parser.add_argument_group("資料來源")
    src_group.add_argument("--datasource", type=str, default="sinopac",
                           choices=["yfinance", "sinopac", "twse"],
                           help="行情資料來源："
                                "sinopac（預設，永豐金，盤後 5 分鐘即更新，需帳號）／"
                                "twse（台灣證交所官方公開 API，免帳號、含上櫃 TPEX）／"
                                "yfinance（免帳號，目前不穩定）")
    src_group.add_argument("--sj-api-key",    type=str, default=None,
                           metavar="KEY",
                           help="永豐金 API Key（亦可設環境變數 SJ_API_KEY）")
    src_group.add_argument("--sj-secret-key", type=str, default=None,
                           metavar="SECRET",
                           help="永豐金 Secret Key（亦可設環境變數 SJ_SECRET_KEY）")
    # ── SQLite 快取控制 ──────────────────────────
    src_group.add_argument("--no-cache",      action="store_true",
                           help="不讀寫 SQLite 快取（每次都打遠端，較慢）")
    src_group.add_argument("--cache-only",    action="store_true",
                           help="只讀 SQLite 快取，不打遠端（快取缺資料的 code 會被略過）")
    src_group.add_argument("--refresh-cache", action="store_true",
                           help="忽略既有快取，全部重抓並覆蓋"
                                "（首次用 sinopac 抓滿時建議搭配此旗標）")
    src_ex = src_group.add_mutually_exclusive_group()
    src_ex.add_argument("--codes", type=str, default=None,
                        help="指定股票代號，逗號分隔，如 2330,2317,2454")
    src_ex.add_argument("--csv",   type=str, default=None,
                        help="從選股 CSV 載入代號，如 --csv 隔日沖選股_20260415.csv")

    # ── 歷史查詢 ──────────────────────────────────
    hist_group = parser.add_argument_group("歷史查詢（搭配 --history）")
    hist_group.add_argument("--code",  type=str, default=None,
                            help="查詢特定股票的歷史記錄")
    hist_group.add_argument("--limit", type=int, default=20,
                            help="查詢筆數，預設 20")

    args = parser.parse_args()

    # ── --list：直接印出策略清單，不需要資料源 ──────────────────
    if args.list:
        print_strategy_list()
        sys.exit(0)

    # 預設模式：無其他模式時自動啟用 --scan
    if not any([args.scan, args.backtest, args.backtest_limit,
                args.combo, args.history, args.diagnose]):
        args.scan = True

    # ── --preset 展開（優先於 --show-combos）────────────────────
    if args.preset:
        expanded = resolve_preset(args.preset)
        if not expanded:
            sys.exit(1)
        if args.show_combos:
            print(f"  ℹ️  --preset 與 --show-combos 同時指定，以 --preset {args.preset} 為準")
        args.show_combos = ",".join(expanded)

    # ── exit-day 解析：未指定時從 PRESET_EXIT_DAY 查表，否則 fallback 2 ──
    _exit_day_explicit = (args.exit_day is not None)
    if not _exit_day_explicit:
        if args.combo and args.preset:
            # 多 preset 時收集各自建議值（忽略 None 的 all3_lean 等）
            _preset_names  = [p.strip().lower() for p in args.preset.split(",") if p.strip()]
            _preset_exits  = [PRESET_EXIT_DAY[n] for n in _preset_names
                              if n in PRESET_EXIT_DAY and PRESET_EXIT_DAY[n] is not None]
            _unique_exits  = list(dict.fromkeys(_preset_exits))   # 去重保序
            if len(_unique_exits) == 1:
                # 所有 preset 建議值一致
                args.exit_day = _unique_exits[0]
                print(f"  ℹ️  --exit-day 未指定，依 preset '{args.preset}' 建議值自動設為 {args.exit_day}")
            elif len(_unique_exits) > 1:
                # 建議值衝突：取最大值並提示
                args.exit_day = max(_unique_exits)
                print(f"  ⚠️  多 preset 的建議 exit-day 不同 {dict(zip(_preset_names, _preset_exits))}，"
                      f"自動取最大值 {args.exit_day}；可用 --exit-day 手動指定")
            else:
                args.exit_day = 2
        else:
            args.exit_day = 2

    # 啟動提示
    date_str = datetime.date.today().strftime("%Y/%m/%d")
    print("\n" + "═"*65)
    print(f"  📈 台股隔日沖選股與回測系統  {date_str}")
    print("═"*65)

    # 初始化資料庫
    conn = init_db()

    # ── 歷史查詢 ─────────────────────────────────
    if args.history:
        print(f"  查詢歷史選股記錄（最近 {args.limit} 筆）")
        df = query_history(conn, code=args.code, limit=args.limit)
        print_history(df)
        conn.close()
        return

    # ── 資料源診斷 ────────────────────────────────
    if args.diagnose:
        sj_key    = args.sj_api_key    or os.environ.get("SJ_API_KEY",    "")
        sj_secret = args.sj_secret_key or os.environ.get("SJ_SECRET_KEY", "")
        diagnose_sources(args.diagnose, days=60,
                         sj_key=sj_key, sj_secret=sj_secret, tail=10)
        conn.close()
        return

    # ── 下載資料 ─────────────────────────────────
    if args.codes or args.csv:
        src_str = f"自訂清單（--codes/--csv）"
    else:
        src_str = f"成交量前 {args.top_n} 名"
    _ds_label = {
        "sinopac":  "永豐金 Shioaji",
        "twse":     "TWSE/TPEX 官方 API",
        "yfinance": "yfinance",
    }
    ds_str = _ds_label.get(args.datasource, args.datasource)
    print(f"  掃描範圍：{src_str}  │  最低均量：{args.min_vol} 張  │  資料源：{ds_str}")
    print(f"  命中門檻:≥{args.min_hit}  │  爆量倍數：{args.vol_mult}x（空方用；多方策略已移除爆量條件）")
    if args.backtest:
        sl_str  = f"-{args.stop_loss}%（估算）" if args.stop_loss > 0 else "未設定"
        tp_str  = f"  止盈：+{args.take_profit}%" if args.take_profit > 0 else ""
        if args.short_only:
            so_str = "  只算空方"
        elif args.long_only:
            so_str = "  只算多方"
        else:
            so_str = ""
        ed_map  = {1:"隔日當沖",2:"持一日",3:"持兩日"}
        ed_str  = ed_map.get(args.exit_day, f"D{args.exit_day}")
        print(f"  回測天數：{args.days} 日  │  出場：{ed_str}  │  停損：{sl_str}{tp_str}{so_str}")
    if args.backtest_limit:
        sl_str = f"-{args.stop_loss}%（估算）" if args.stop_loss > 0 else "未設定"
        print(f"  回測天數：{args.days} 日  │  出場：漲停/跌停 或 D1收盤  │  停損：{sl_str}")
    print(f"  平行執行緒：{args.workers} 個（可用 --workers N 調整）")
    print("═"*65 + "\n")

    # 優先用 --codes / --csv，否則從 TWSE 取清單
    codes = []
    if args.codes:
        codes = [c.strip() for c in args.codes.split(",") if c.strip()]
        print(f"  📋 自訂清單：{len(codes)} 檔")
    elif args.csv:
        try:
            df_csv = pd.read_csv(args.csv)
            col    = next((c for c in df_csv.columns if "代號" in c), df_csv.columns[0])
            codes  = df_csv[col].astype(str).str.strip().tolist()
            print(f"  📋 從 CSV 載入：{len(codes)} 檔（{args.csv}）")
        except Exception as e:
            print(f"  ❌ CSV 讀取失敗：{e}")
            conn.close()
            sys.exit(1)
    else:
        codes = _get_top_codes(args.top_n)

    if not codes:
        print("  ❌ 無法取得股票清單，請使用 --codes 2330,2317 指定股票")
        conn.close()
        sys.exit(1)

    # ── 選擇資料來源 ──────────────────────────────
    sj_key    = args.sj_api_key    or os.environ.get("SJ_API_KEY",    "")
    sj_secret = args.sj_secret_key or os.environ.get("SJ_SECRET_KEY", "")

    # sinopac 必須有金鑰；twse / yfinance 不需要
    if args.datasource == "sinopac" and not (sj_key and sj_secret):
        if not args.cache_only:
            print("  ❌ 使用永豐金資料源需提供 API 金鑰，有三種方式：")
            print("     方式一（命令列）：--sj-api-key YOUR_KEY --sj-secret-key YOUR_SECRET")
            print("     方式二（環境變數）：set SJ_API_KEY=... && set SJ_SECRET_KEY=...")
            print("     方式三（不需金鑰）：--datasource twse")
            print("     ＊ API 金鑰請至永豐金證券後台申請")
            conn.close()
            sys.exit(1)

    if args.no_cache:
        # 不走快取，每次直接打遠端（舊行為）
        if args.datasource == "sinopac":
            data = fetch_data_sinopac(codes, args.days + 30, sj_key, sj_secret)
        elif args.datasource == "twse":
            data = fetch_data_twse(codes, args.days + 30)
        else:
            data = fetch_data(codes, args.days + 30)
    else:
        # 走 SQLite 快取 + 增量補（推薦流程）
        data = fetch_data_cached(
            codes, args.days + 30,
            datasource=args.datasource,
            sj_key=sj_key, sj_secret=sj_secret,
            cache_only=args.cache_only,
            refresh=args.refresh_cache,
        )

    if not data:
        print("  ❌ 無法取得歷史資料")
        conn.close()
        sys.exit(1)

    # ── --date：把所有 DataFrame 截到指定日期 ────────
    if args.date:
        try:
            _target_date = datetime.date.fromisoformat(args.date)
        except ValueError:
            print(f"  ❌ --date 格式錯誤（須為 YYYY-MM-DD）：{args.date}")
            conn.close()
            sys.exit(1)
        _sliced: dict = {}
        for _code, _df in data.items():
            _df2 = _df[_df.index.date <= _target_date]
            if len(_df2) >= 22:
                _sliced[_code] = _df2
        if not _sliced:
            print(f"  ❌ 指定日期 {args.date} 找不到任何股票的 K 棒資料（快取不足？）")
            conn.close()
            sys.exit(1)
        data     = _sliced
        date_str = _target_date.strftime("%Y/%m/%d")
        print(f"  📅 回顧模式：以 {args.date} 為基準日，命中股票以當日收盤訊號計算")

    # ── 選股 ─────────────────────────────────────
    if args.scan:
        strat_filter = [s.strip().upper() for s in args.strategy.split(",")] if args.strategy else None
        show_combos  = ([c.strip() for c in args.show_combos.split(",") if c.strip()]
                        if args.show_combos else None)
        rows = run_scan(data, args.min_hit, args.vol_mult, args.min_vol,
                        strategy_filter=strat_filter,
                        show_combos=show_combos,
                        workers=args.workers)
        print_scan_result(rows, date_str, args.min_hit, show_combos=show_combos)

        if args.save and rows:
            scan_date = args.date if args.date else datetime.date.today().isoformat()
            save_scan(conn, scan_date, rows, args.min_hit)
            print(f"  💾 選股結果已存入 SQLite：{DB_PATH}")

            date_tag = args.date.replace("-", "") if args.date else datetime.date.today().strftime("%Y%m%d")
            df_out = pd.DataFrame(rows)
            # 插入「價位區間」欄（緊接在「收盤」欄後面）
            close_idx = df_out.columns.get_loc("收盤") + 1 if "收盤" in df_out.columns else 0
            df_out.insert(close_idx, "價位區間", df_out["收盤"].apply(price_group))
            # 排序：先依價位區間（依定義順序），再依命中數（高→低）
            _group_order = {lo_hi_label[2]: i
                            for i, lo_hi_label in enumerate(_PRICE_BUCKETS)}
            df_out["_grp_ord"] = df_out["價位區間"].map(_group_order).fillna(99)
            df_out = df_out.sort_values(["_grp_ord", "命中數"],
                                        ascending=[True, False]).drop(columns=["_grp_ord"])
            df_out = df_out.reset_index(drop=True)
            # ── CSV（供 daytrade_live.py --csv 使用）
            csv_fname = f"隔日沖選股_{date_tag}.csv"
            df_out.to_csv(csv_fname, index=False, encoding="utf-8-sig")
            print(f"  💾 CSV  ：{csv_fname}")
            # ── Excel（美化版，供人工閱覽）
            xlsx_fname = f"隔日沖選股_{date_tag}.xlsx"
            save_scan_xlsx(df_out, xlsx_fname)
            print(f"  💾 Excel：{xlsx_fname}")

    # ── 回測 ─────────────────────────────────────
    if args.backtest:
        strat_filter = [s.strip().upper() for s in args.strategy.split(",")] \
                       if args.strategy else None
        stats = run_backtest(data, args.days, args.min_hit,
                             args.vol_mult, args.min_vol, args.stop_loss,
                             take_profit=args.take_profit,
                             exit_day=args.exit_day,
                             short_only=args.short_only,
                             long_only=args.long_only,
                             strategy_filter=strat_filter,
                             workers=args.workers)
        print_backtest_result(stats, args.days, args.min_hit, args.stop_loss,
                              take_profit=args.take_profit,
                              exit_day=args.exit_day,
                              short_only=args.short_only,
                              long_only=args.long_only,
                              strategy_filter=strat_filter)

        if args.save:
            ed_tag = f"_d{args.exit_day}"
            tp_tag = f"_tp{args.take_profit}" if args.take_profit > 0 else ""
            so_tag = "_short" if args.short_only else ""
            fname  = save_backtest_csv(stats, args.days, args.min_hit, args.stop_loss)
            print(f"  💾 回測結果已儲存：{fname}")

    # ── 漲跌停出場回測 ───────────────────────────
    if args.backtest_limit:
        strat_filter = [s.strip().upper() for s in args.strategy.split(",")]\
                       if args.strategy else None
        need_trades  = args.show_trades > 0

        result = run_backtest_limit(
            data, args.days, args.min_hit,
            args.vol_mult, args.min_vol,
            stop_loss=args.stop_loss,
            strategy_filter=strat_filter,
            return_trades=need_trades,
            workers=args.workers,
        )
        if need_trades:
            stats, all_trades = result
        else:
            stats = result

        print_backtest_limit_result(
            stats, args.days, args.min_hit, args.stop_loss,
            strategy_filter=strat_filter,
        )

        if need_trades:
            print_trade_detail(
                all_trades,
                strategy_filter=strat_filter,
                long_only=args.long_only,
                short_only=args.short_only,
                show_n=args.show_trades,
            )
        if args.save:
            sl_tag = f"_sl{args.stop_loss}" if args.stop_loss > 0 else "_nosl"
            date_s = datetime.date.today().strftime("%Y%m%d")
            fname  = f"漲跌停回測_{date_s}_d{args.days}_hit{args.min_hit}{sl_tag}.csv"
            df_out = pd.DataFrame(stats)
            df_out.insert(0, "回測日期", datetime.date.today().isoformat())
            df_out.insert(1, "回測天數", args.days)
            df_out.insert(2, "命中門檻", args.min_hit)
            df_out.insert(3, "停損(%)",  args.stop_loss)
            df_out.to_csv(fname, index=False, encoding="utf-8-sig")
            print(f"  💾 回測結果已儲存：{fname}")

    # ── 策略組合分析 ─────────────────────────────
    if args.combo:
        # 解析 show_combos 並自動偵測長空混合
        _raw_show = ([c.strip() for c in args.show_combos.split(",") if c.strip()]
                     if args.show_combos else None)

        def _is_short_combo(c: str) -> bool:
            return any(s.endswith("S") for s in c.split("+"))

        if _raw_show:
            _long_show  = [c for c in _raw_show if not _is_short_combo(c)] or None
            _short_show = [c for c in _raw_show if     _is_short_combo(c)] or None
        else:
            _long_show  = None
            _short_show = None

        # 決定要跑哪些方向
        # --short_only：只跑空方；有混合 preset：長空都跑；否則跟原本邏輯
        _run_long  = (not args.short_only) and (_raw_show is None or _long_show  is not None)
        _run_short = args.short_only       or  (_raw_show is not None and _short_show is not None)

        # ── exit_day：all3_lean 時長空各自查 PRESET_EXIT_DAY ────────
        def _resolve_exit_day(direction: str) -> int:
            """手動指定優先；all3_lean 混合時長空各查子 preset；否則用 args.exit_day"""
            if _exit_day_explicit:
                return args.exit_day
            if args.preset and args.preset.strip().lower() == "all3_lean":
                sub = "long3_lean" if direction == "long" else "short3_lean"
                return PRESET_EXIT_DAY.get(sub, 2)
            return args.exit_day  # 已在前面由 PRESET_EXIT_DAY 或 fallback 2 設好

        _combo_base = dict(
            days=args.days, min_hit=args.min_hit,
            vol_mult=args.vol_mult, min_avg_vol=args.min_vol,
            stop_loss=args.stop_loss,
            min_signals=args.min_signals,
            max_combo=args.max_combo,
            workers=args.workers,
        )

        if _run_long:
            _long_exit = _resolve_exit_day("long")
            rows_long = run_combo_analysis(data, direction="long",
                                           exit_day=_long_exit,
                                           forced_combos=_long_show,
                                           **_combo_base)
            print_combo_result(rows_long, args.days, "long",
                               args.stop_loss, args.min_signals,
                               max_combo=args.max_combo,
                               show_combos=_long_show)
            if args.save and rows_long:
                sl_tag = f"_sl{args.stop_loss}" if args.stop_loss > 0 else "_nosl"
                date_s = datetime.date.today().strftime("%Y%m%d")
                fname  = f"組合分析_long_{date_s}_d{args.days}{sl_tag}_ed{_long_exit}.csv"
                pd.DataFrame(rows_long).to_csv(fname, index=False, encoding="utf-8-sig")
                print(f"  💾 組合分析已儲存：{fname}")

        if _run_short:
            _short_exit = _resolve_exit_day("short")
            rows_short = run_combo_analysis(data, direction="short",
                                            exit_day=_short_exit,
                                            forced_combos=_short_show,
                                            **_combo_base)
            print_combo_result(rows_short, args.days, "short",
                               args.stop_loss, args.min_signals,
                               max_combo=args.max_combo,
                               show_combos=_short_show)
            if args.save and rows_short:
                sl_tag = f"_sl{args.stop_loss}" if args.stop_loss > 0 else "_nosl"
                date_s = datetime.date.today().strftime("%Y%m%d")
                fname  = f"組合分析_short_{date_s}_d{args.days}{sl_tag}_ed{_short_exit}.csv"
                pd.DataFrame(rows_short).to_csv(fname, index=False, encoding="utf-8-sig")
                print(f"  💾 組合分析已儲存：{fname}")

    conn.close()


if __name__ == "__main__":
    main()
